# Расчет наведенного напряжения на проводах и троссах иследуемой ВЛ

# Открываем библиотеку для работы с файлом Excel
# https://habrahabr.ru/post/232291/
import openpyxl
import numpy as np

from scipy.integrate import simps
from scipy.constants import epsilon_0

import scipy.sparse as sparse
import scipy.sparse.linalg as linalg
from time import time

# Открываем каталог с данными по опорам и проводникам
Katal = openpyxl.load_workbook(filename = 'Katal.xlsx')
# выбирем необходимый рабочий лист
Kat = Katal['Kat']

# Функция выбора фазировки участка и зеркало опоры
def VF(i,j,F,Katalog,Kat):
    XA=Kat[Katalog.get(i)+str(2)].value
    XB=Kat[Katalog.get(i)+str(3)].value
    XC=Kat[Katalog.get(i)+str(4)].value
    XT1=Kat[Katalog.get(i)+str(5)].value
    XT2=Kat[Katalog.get(i)+str(6)].value

    YA=Kat[Katalog.get(i)+str(9)].value
    YB=Kat[Katalog.get(i)+str(10)].value
    YC=Kat[Katalog.get(i)+str(11)].value
    YT1=Kat[Katalog.get(i)+str(12)].value
    YT2=Kat[Katalog.get(i)+str(13)].value
    
    G=Kat[Katalog.get(i)+str(15)].value
    Pos=Kat[Katalog.get(i)+str(14)].value

    def ZandG(XA,XB,XC,XT1,XT2,YA,YB,YC,G,a):
        XA=XA*a
        XB=XB*a
        XC=XC*a
        XT1=XT1*a
        XT2=XT2*a

        YA-=G
        YB-=G
        YC-=G
        return XA,XB,XC,XT1,XT2,YA,YB,YC
        
    a=abs(j)/j

    if abs(j)==1:
        XA,XB,XC,XT1,XT2,YA,YB,YC = ZandG(XA,XB,XC,XT1,XT2,YA,YB,YC,G,a)
    elif abs(j)==2:
        XA,XB,XC=XA,XC,XB
        YA,YB,YC=YA,YC,YB
        XA,XB,XC,XT1,XT2,YA,YB,YC = ZandG(XA,XB,XC,XT1,XT2,YA,YB,YC,G,a)
    elif abs(j)==3:
        XA,XB,XC=XC,XA,XB
        YA,YB,YC=YC,YA,YB
        XA,XB,XC,XT1,XT2,YA,YB,YC = ZandG(XA,XB,XC,XT1,XT2,YA,YB,YC,G,a)
    elif abs(j)==4:
        XA,XB,XC=XB,XA,XC
        YA,YB,YC=YB,YA,YC
        XA,XB,XC,XT1,XT2,YA,YB,YC = ZandG(XA,XB,XC,XT1,XT2,YA,YB,YC,G,a)
    elif abs(j)==5:
        XA,XB,XC=XB,XC,XA
        YA,YB,YC=YB,YC,YA
        XA,XB,XC,XT1,XT2,YA,YB,YC = ZandG(XA,XB,XC,XT1,XT2,YA,YB,YC,G,a)
    elif abs(j)==6:
        XA,XB,XC=XC,XB,XA
        YA,YB,YC=YC,YB,YA
        XA,XB,XC,XT1,XT2,YA,YB,YC = ZandG(XA,XB,XC,XT1,XT2,YA,YB,YC,G,a)

    if F=='X0':
        return XA
    if F=='X1':
        return XB
    if F=='X2':
        return XC
    if F=='XT1':
        return XT1
    if F=='XT2':
        return XT2
    if F=='Y0':
        return YA
    if F=='Y1':
        return YB
    if F=='Y2':
        return YC
    if F=='YT1':
        return YT1
    if F=='YT2':
        return YT2
    if F=='OpFig':
        return (XA,XB,XC,YA+G,YB+G,YC+G,G,XT1,XT2,YT1,YT2,Pos)

# Функция формирования словоря имён столбцов таблицы Excel
def ExDict():
    # Создаём словарь по которому будем обращаться к элементам Каталога в файле
    Alf = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    Katalog = {}
    sch = 1
    for i in range(0,2):
        if i !=0:
                for j in Alf:
                        for k in Alf:
                                Katalog.setdefault(sch, j*i+k)
                                sch  += 1
                        if j == 'B': break
        else:
                for j in Alf:
                        if j == 'A' : continue
                        Katalog.setdefault(sch, j)
                        sch  += 1
    return Katalog


def Vvod_inf(ivl, vvl, zy, zytr, zvl, per, koef_graf,pz,Bool_hesh,File_hesh,trig_r):


    if Bool_hesh[0] and Bool_hesh[1] and Bool_hesh[2] and trig_r:
        try: 
            FiAl = np.load('Hesh_files/FiA'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy')
            FiBl = np.load('Hesh_files/FiB'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy')
            FiCl = np.load('Hesh_files/FiC'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy')
            FiT1l = np.load('Hesh_files/FiT1'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy')
            p_l = np.load('Hesh_files/p'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy')
            cols = np.load('Hesh_files/cols'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy')
            l = np.shape(cols)[0]
            return [list(FiAl[i,:cols[i]]*koef_graf) for i in range(l)],\
                    [list(FiBl[i,:cols[i]]*koef_graf) for i in range(l)],\
                    [list(FiCl[i,:cols[i]]*koef_graf) for i in range(l)],\
                    [list(FiT1l[i,:cols[i]]*koef_graf) for i in range(l)],\
                    [list(p_l[i,:cols[i]]) for i in range(l)]
        except Exception:
            1
    
    VVL = np.array(vvl)
    IVL = np.array(ivl)
    Zy = np.array(zy)
    Zytr = np.array(zytr)
    Zvl = np.array(zvl)
    # Часть, которая создаём матрицу графов
    # Создаём пустой массив для будующего шаблона
    Aeds = sparse.lil_matrix((5,5),dtype=np.float64)
    # Заполняем его
    for i in range(5):
        for j in range(5):
                if i == j :
                        Aeds[i,j] = -1
    Alepn = -1* Aeds
    Alepk = Aeds
    Alep = sparse.vstack([Alepn, Alepk],format='lil')
    nagr = sparse.lil_matrix([[-1, -1, -1, 0, 0]],dtype=np.float64)
    Anagr = sparse.vstack([Alepn, nagr],format='lil')
    
    # Функция для определения количества ветвей в исследуемой ВЛ
    def ivv(IVL):
        ij = np.shape(IVL)
        v=0
        iv = [ ]
        for i in range(ij[0]):
                if IVL[i][0] != 0:
                        v = v+1
                        iv.append(i)
        return iv
    iv = ivv(IVL)
    # Функция для определения количестава учасков в ИВЛ
    def ychh(IVL,iv):
        ij = np.shape(iv)
        rr = {}
        ych = [ ]
        for i in range(ij[0]):
                if IVL[iv[i]][2] in rr :
                        ych.append(int(IVL[iv[i]][3]-IVL[iv[i-1]][3]))
                else:
                        ych.append(int(IVL[iv[i]][3]-IVL[iv[i]][2]))
                rr[IVL[iv[i]][2]] = 1
        return ych
    ych = ychh(IVL,iv)
    # Функция определения крличества ветвей в ИВЛ
    def vvetvi(ych,iv):
        ij = np.shape(iv)
        vetvi = 0
        for i in range(ij[0]):
                vetvi = vetvi + ych[i]*5
        return vetvi
    vetvi = int(vvetvi(ych,iv))
    # Определяем количество узлов в ИВЛ
    yzli = vetvi + 5
    # Функция для привязки ветвей к узлам
    def opmm(IVL,iv):
        rr = {} 
        ij = np.shape(iv)
        opm=[]
        # Создаём массив, но он с не равными по длинне строками
        for i in range(ij[0]):
                o = []
                if IVL[iv[i]][2] in rr :
                        for j in range(int(IVL[iv[i]][3]-IVL[iv[i-1]][3])+1):
                                if j == 0:
                                        o.append(int(IVL[iv[i]][2]))
                                else:
                                        o.append(int(IVL[iv[i-1]][3])+j)

                else:
                        for j in range(int(IVL[iv[i]][3]-IVL[iv[i]][2])+1):
                                o.append(int(IVL[iv[i]][2])+j)

                rr[IVL[iv[i]][2]] = 1
                opm.append(o)
        # Ищем самую длинную строку

        h = 0
        for i in range(ij[0]):
                if h < (np.shape(opm[i]))[0]:
                        h = (np.shape(opm[i]))[0]
        # Создаём нулевую матрицу
        op = np.zeros((ij[0], h), dtype=np.int64)        
        # Выравниваем марицу
        for i in range(ij[0]):
                for j in range((np.shape(opm[i]))[0]):
                        op[i][j] = opm[i][j]
        return op
    opm = opmm(IVL,iv)
    
    ij = np.shape(iv)
    iv.append((np.shape(IVL))[0])
    # Функция, котороя формирует матрицу графов для ИВЛ
    def AA1(yzli,vetvi,IVL,ych,iv):
        A1 = sparse.lil_matrix((yzli, vetvi),dtype=np.float64)
        ynn=np.zeros((np.shape(ych)[0],max(ych)),dtype=np.int64)
        ykk=np.zeros((np.shape(ych)[0], max(ych)),dtype=np.int64)
        
        k = 0
        rr = {}

        for n in range((np.shape(ych))[0]):

                for p in range(k,ych[n]+k):
                        if (IVL[iv[n]][2] in rr) and (p == k):
                                ynn[n][p-k]=(rr.get(IVL[iv[n]][2])+1)*5-5
                                for i in range((rr.get(IVL[iv[n]][2])+1)*5-5,(rr.get(IVL[iv[n]][2])+1)*5):
                                        for j in range((p+1)*5-5, (p+1)*5):
                                                A1[i,j] = Alepn[i-((rr.get(IVL[iv[n]][2])+1)*5-5),j-((p+1)*5-5)]
                                ykk[n][p-k]=(p+1)*5
                                for i in range((p+1)*5-5, (p+1)*5):
                                        for j in range((p+1)*5-5, (p+1)*5):
                                                A1[i+5,j] = Alepk[i-((p+1)*5-5),j-((p+1)*5-5)]
                        else:
                                ynn[n][p-k]=(p+1)*5-5
                                for i in range((p+1)*5-5, (p+1)*5):
                                        for j in range((p+1)*5-5, (p+1)*5):
                                                A1[i,j] = Alepn[i-((p+1)*5-5),j-((p+1)*5-5)]
                                ykk[n][p-k]=(p+1)*5
                                for i in range((p+1)*5-5, (p+1)*5):
                                        for j in range((p+1)*5-5, (p+1)*5):
                                                A1[i+5,j] = Alepk[i-((p+1)*5-5),j-((p+1)*5-5)]
                if IVL[iv[n]][2] not in rr:
                    rr[IVL[iv[n]][2]] = k
                k = p+1
        return A1, ynn, ykk
    if not Bool_hesh[0]:
        A1, ynn, ykk  = AA1(yzli,vetvi,IVL,ych,iv)
        np.save('Hesh_files/ynn'+str(File_hesh[0])+'.npy',ynn)
        np.save('Hesh_files/ykk'+str(File_hesh[0])+'.npy',ykk)

    # Определяем колличество ВВЛ
    vv = ivv(VVL)
    vv.append((np.shape(VVL))[0])
    # Функция которая определяет колличество пролётов ВВЛ
    def Prolett(vv,opm,VVL):
        Prolet = []
        for i in range((np.shape(vv))[0]-1):
                x=0
                for k in range((np.shape(opm))[0]):
                        for n in range((np.shape(opm))[1]):
                            for c in range(vv[i]+1,vv[i+1]):
                                if VVL[c][2] == opm[k][n]:
                                        for b in range((np.shape(opm))[1]):
                                                if VVL[c][3] == opm[k][b]:
                                                    x += b-n
                Prolet.append(x)
                                                
        return Prolet

    Prolet = Prolett(vv,opm,VVL)

    # Определяем количество ветвей в ВВЛ
    def vetvi11(vv, Prolet):
        vetvi1 = 0
        for i in range((np.shape(vv))[0]-1):
                vetvi1 = vetvi1 + 10 + Prolet[i] * 5
        return vetvi1
    vetvi1 = vetvi11(vv, Prolet)
    # Определяем количество узлов в ВВЛ
    def yzli11(vv, Prolet):
        yzli1 = 0
        for i in range((np.shape(vv))[0]-1):
                yzli1 = yzli1 + 6+ Prolet[i] * 5
        return yzli1
    yzli1 = yzli11(vv, Prolet)
    # Функция, котороя формирует матрицу графов для ВВЛ
    def AA2(vv, Prolet, yzli1, vetvi1):
        A2 = sparse.lil_matrix((yzli1, vetvi1),dtype=np.float64)
        vvlnagr = np.zeros(((np.shape(vv))[0]-1,2),dtype=np.int64)
        ki = 0
        kj = 0
        for v in range((np.shape(vv))[0]-1):
                
                for i in range(ki,ki+5):
                        for j in range(kj,kj+5):
                                A2[i,j] = Aeds[i-ki,j-kj]
                kj = kj+5
                for p in range(Prolet[v]):
                        for i in range(ki,ki+10):
                                for j in range(kj,kj+5):
                                        A2[i,j] = Alep[i-ki,j-kj]
                        ki = ki + 5
                        kj = kj + 5
                for i in range(ki,ki+6):
                        for j in range(kj,kj+5):
                                A2[i,j] = Anagr[i-ki,j-kj]
                vvlnagr[v][0]=ki
                vvlnagr[v][1]=kj
                ki = ki+6
                kj = kj+5
        return A2,vvlnagr
    if not Bool_hesh[0]:
        A2,vvlnagr = AA2(vv, Prolet, yzli1, vetvi1)
    
    if not Bool_hesh[0]:
        A4 = sparse.lil_matrix((yzli1, vetvi),dtype=np.float64)
        A3 = sparse.lil_matrix((yzli, vetvi1),dtype=np.float64)
        A = sparse.hstack((sparse.vstack([A1, A4]), sparse.vstack([A3, A2])),format='lil')
        # Обнулим ненужные массивы
        del A1, A2, A3, A4
        sparse.save_npz('Hesh_files/A'+str(File_hesh[0])+'.npz', A.tocsc())
        np.save('Hesh_files/vvlnagr'+str(File_hesh[0])+'.npy',vvlnagr)
    else:
        A = sparse.load_npz('Hesh_files/A'+str(File_hesh[0])+'.npz').tolil()
        ynn = np.load('Hesh_files/ynn'+str(File_hesh[0])+'.npy')
        ykk = np.load('Hesh_files/ykk'+str(File_hesh[0])+'.npy')
        vvlnagr = np.load('Hesh_files/vvlnagr'+str(File_hesh[0])+'.npy')
    
    
    # Заполняем матрици привязок инофой по ИВЛ
    vetvi0 = vetvi + vetvi1
    yzli0 = yzli + yzli1
    WZI = {}
    WHX = {}
    WF = {}
    WR = {}
    WS = {}
    YIn = {}
    YIk = {}
    WL = {}
    WT1 = {}
    WT2 = {}
    WRAZV = {}
    E = sparse.lil_matrix((vetvi0, 1), dtype=np.complex128)
    W = sparse.lil_matrix((vetvi0, vetvi0), dtype=np.complex128)
    Wn = sparse.lil_matrix((vetvi0, vetvi0), dtype=np.complex128)
    k = 0
    if not Bool_hesh[0]:
        for v in range((np.shape(iv))[0]-1):
            for b in range(iv[v]+1,iv[v+1]):
                    for c in range((np.shape(opm))[1]):
                            if opm[v][c] == IVL[b][2]:
                                    for g in range((np.shape(opm))[1]):
                                            if opm[v][g] == IVL[b][3]:
                                                    WLL = (IVL[b][1]*1000)/(g-c)*1.0
                                                    break
                                    break
                    for i in range((np.shape(opm))[1]):
                            if IVL[b][2] == opm[v][i]:
                                    for j in range((np.shape(opm))[1]):
                                            if IVL[b][3] == opm[v][j]:
                                                    for n in range(i+1, j+1):
                                                            WZI[(k,k)] = opm[v][n]
                                                            YIn[(k,k)] = ynn[v][n-1]
                                                            YIk[(k,k)] = ykk[v][n-1]
                                                            WHX[(k,k)] = int(IVL[b][6])
                                                            WF[(k,k)] = int(IVL[b][5])
                                                            WR[(k,k)] = int(IVL[b][7])
                                                            WS[(k,k)] = IVL[b][4]
                                                            WT1[(k,k)] = int(IVL[b][8])
                                                            WT2[(k,k)] = int(IVL[b][9])
                                                            WL[(k,k)] = WLL
                                                            k = k+5
                                                    break
                                    break
    else:
        for v in range((np.shape(iv))[0]-1):
            for b in range(iv[v]+1,iv[v+1]):
                for i in range((np.shape(opm))[1]): 
                    if IVL[b][2] == opm[v][i]:
                        for j in range((np.shape(opm))[1]):
                            if IVL[b][3] == opm[v][j]:   
                                k+=(j-i)*5
                                break
                        break
    
    # Создаём вспомогательные матрици для привязки ВВЛ к ИВЛ
    opvk = np.zeros(((np.shape(VVL))[0],(np.shape(opm))[1]), dtype=np.int64)
    opvn = np.zeros(((np.shape(VVL))[0],(np.shape(opm))[1]), dtype=np.int64)
    ws = np.zeros(((np.shape(VVL))[0],(np.shape(opm))[1]), dtype=np.float64)
    for v in range((np.shape(vv))[0]-1):
        for a in range(vv[v]+1,vv[v+1]):
                for b in range((np.shape(opm))[0]):
                        for c in range((np.shape(opm))[1]):
                                if VVL[a][2] == opm[b][c]:
                                        for f in range((np.shape(opm))[1]):
                                                if VVL[a][3] == opm[b][f]:
                                                        dS = (VVL[a][10]-VVL[a][4])/(f-c)
                                                        s = VVL[a][4]
                                                        for g in range(c+1,f+1):
                                                                s += dS
                                                                ws[a][g-c-1] = s - dS/2
                                                                opvk[a][g-c-1] = opm[b][g]
                                                                opvn[a][g-c-1] = opm[b][g-1]             
                                                        break
    
    # Заполняем матрици привязок инфой по ВВЛ
    NotSim=[]
    y = yzli
    alf1=np.cos(240*np.pi/180)+1j*np.sin(240*np.pi/180)
    alf2=np.cos(120*np.pi/180)+1j*np.sin(120*np.pi/180)
    vvleds = np.zeros(((np.shape(vv))[0]-1,9),dtype=np.complex128)
    for v in range((np.shape(vv))[0]-1):
        if not Bool_hesh[2]:
            if VVL[vv[v]][6] == 0:
                KefU = 0
                KefPQ = 1
            else:
                KefU=VVL[vv[v]][6]/abs(VVL[vv[v]][6])
                KefPQ=abs(VVL[vv[v]][6])
            E[k,0] = (VVL[vv[v]][0]*KefU*1000)/np.sqrt(3)
            E[k+1,0] = (VVL[vv[v]][0]*KefU*1000*alf1)/np.sqrt(3)
            E[k+2,0] = (VVL[vv[v]][0]*KefU*1000*alf2)/np.sqrt(3)
            Wn[k,k] = Wn[k+1,k+1] = Wn[k+2,k+2] = 0.001
            Wn[k+3,k+3] = Wn[k+4,k+4] = 10**15

            vvleds[v][0] = E[k,0]
            vvleds[v][1] = E[k+1,0]
            vvleds[v][2] = E[k+2,0]

        k +=5
        kyrs = 0
        if not Bool_hesh[0]:
            for a in range(vv[v]+1,vv[v+1]):
                    for b in range((np.shape(opvk))[1]):
                            if opvk[a][b] != 0:
                                    for c in range(kyrs, vetvi):
                                            if opvk[a][b] == WZI.get((c,c),0):
                                                    WZI[(c,k)]=WZI[(k,k)]=WZI[(k,c)]=opvk[a][b]
                                                    YIn[(k,k)]=YIn[(c,k)]=y
                                                    YIn[(k,c)]=YIn[(c,c)]
                                                    YIk[(k,k)]=YIk[(c,k)]=y+5
                                                    YIk[(k,c)]=YIk[(c,c)]
                                                    WHX[(c,k)]=WHX[(k,k)]=int(VVL[a][6])
                                                    WHX[(k,c)]=WHX[(c,c)]
                                                    WF[(c,k)]=WF[(k,k)]=int(VVL[a][5])
                                                    WF[(k,c)]=WF[(c,c)]
                                                    WR[(c,k)]=WR[(k,k)]=int(VVL[a][7])
                                                    WR[(k,c)]=WR[(c,c)]
                                                    WS[(c,k)]=WS[(k,k)]=ws[a][b]
                                                    WS[(k,c)]=WS[(c,k)]
                                                    WL[(k,k)]=WL[(c,c)]
                                                    WT1[(k,k)]=WT1[(c,k)]=int(VVL[a][8])
                                                    WT1[(k,c)]=WT1[(c,c)]
                                                    WT2[(k,k)]=WT2[(c,k)]=int(VVL[a][9])
                                                    WT2[(k,c)]=WT2[(c,c)]
                                                    WRAZV[(c,k)]=VVL[a][11]
                                                    WRAZV[(k,c)]=VVL[a][11]
                                                    kyrs = c
                                    k +=5
                                    y +=5
        else:
            for a in range(vv[v]+1,vv[v+1]):
                    for b in range((np.shape(opvk))[1]):
                            if opvk[a][b] != 0:
                                    for c in range(kyrs, vetvi):
                                            if opvk[a][b] == WZI.get((c,c),0):
                                                    kyrs = c
                                    k +=5
                                    y +=5
        NotSim.append(VVL[vv[v]][11])
        if not Bool_hesh[2]:
            if VVL[vv[v]][11] ==0:
                Wn[k,k] = ((VVL[vv[v]][0]*1000)**2)/((VVL[vv[v]][4]-VVL[vv[v]][5]*1j)*(10**6)*KefPQ)
                Wn[k+1,k+1] = ((VVL[vv[v]][0]*1000)**2)/((VVL[vv[v]][7]-VVL[vv[v]][9]*1j)*(10**6)*KefPQ)
                Wn[k+2,k+2] = ((VVL[vv[v]][0]*1000)**2)/((VVL[vv[v]][8]-VVL[vv[v]][10]*1j)*(10**6)*KefPQ)

                vvleds[v][3] = (VVL[vv[v]][4]/3+VVL[vv[v]][5]*1j/3)*(10**6)*abs(KefU)
                vvleds[v][4] = (VVL[vv[v]][7]/3+VVL[vv[v]][9]*1j/3)*(10**6)*abs(KefU)
                vvleds[v][5] = (VVL[vv[v]][8]/3+VVL[vv[v]][10]*1j/3)*(10**6)*abs(KefU)

                vvleds[v][6] = Wn[k,k]
                vvleds[v][7] = Wn[k+1,k+1]
                vvleds[v][8] = Wn[k+2,k+2]
            else:
                #print("Сопр")
                Wn[k,k] = ((VVL[vv[v]][0]*1000/np.sqrt(3))**2)/((VVL[vv[v]][4]-VVL[vv[v]][5]*1j)*(10**6)*KefPQ)
                Wn[k+1,k+1] = ((VVL[vv[v]][0]*1000/np.sqrt(3))**2)/((VVL[vv[v]][7]-VVL[vv[v]][9]*1j)*(10**6)*KefPQ)
                Wn[k+2,k+2] = ((VVL[vv[v]][0]*1000/np.sqrt(3))**2)/((VVL[vv[v]][8]-VVL[vv[v]][10]*1j)*(10**6)*KefPQ)

                vvleds[v][3] = (VVL[vv[v]][4]+VVL[vv[v]][5]*1j)*(10**6)*abs(KefU)
                vvleds[v][4] = (VVL[vv[v]][7]+VVL[vv[v]][9]*1j)*(10**6)*abs(KefU)
                vvleds[v][5] = (VVL[vv[v]][8]+VVL[vv[v]][10]*1j)*(10**6)*abs(KefU)

                vvleds[v][6] = Wn[k,k]
                vvleds[v][7] = Wn[k+1,k+1]
                vvleds[v][8] = Wn[k+2,k+2]

                """ Wn[k,k] = 0.001
                Wn[k+1,k+1] = 0.001
                Wn[k+2,k+2] = 0.001 """
            Wn[k+3,k+3] = Wn[k+4,k+4] = 10**15
        k +=5
        y +=6
    # Создаём словарь имён столбцов таблицы Excel 
    if not Bool_hesh[2]:
        sparse.save_npz('Hesh_files/Wn'+str(File_hesh[2])+'.npz', Wn.tocsc())
        sparse.save_npz('Hesh_files/E'+str(File_hesh[2])+'.npz', E.tocsc())
        np.save('Hesh_files/vvleds'+str(File_hesh[2])+'.npy',vvleds)
    else:
        Wn = sparse.load_npz('Hesh_files/Wn'+str(File_hesh[2])+'.npz').tolil()
        E = sparse.load_npz('Hesh_files/E'+str(File_hesh[2])+'.npz').tolil()
        vvleds = np.load('Hesh_files/vvleds'+str(File_hesh[2])+'.npy')


    Katalog=ExDict()
    # Заполняем матрици взаимных сопротивлений и проводимостей
    #print("test")

    # Подинтегральные выражения интеграла Карсона
    def Kars_ii(x,hi,ri,mp):
        return (np.exp(-2*x*hi)*np.cos(x*ri))/(x+(x**2+mp**2)**0.5)
    def Kars_ij(x,hi,hj,xi,xj,mp):
        return (np.exp(-x*(hi+hj))*np.cos(x*abs(xi-xj)))/(x+(x**2+mp**2)**0.5)



    mu0=np.pi*4*10**(-7)
    eps=epsilon_0
    jp=1j*2*np.pi*50
    #pz=100
    epz=1
    mp=(mu0*jp*(1/pz))**0.5
    sh_int=np.arange(0, 1, 0.00001)

    d_matZ={}
    d_matC={}
    SpEmk={} # Cловарь признаков участков сближения для расчета ёмкости
    SpVz={} # Словарь адресов взаимных ёмкостей
    Yu = sparse.lil_matrix((yzli0,yzli0), dtype=np.complex128)
    if not Bool_hesh[0]:
        for i in range(0,vetvi0,5):
            if WZI.get((i,i), 0) != 0:
                SpEmk[i]=['',[]]
                SpVz[i]=[]
                SpVz[i].append((i,i))
                for j in range(0,vetvi0,5):
                    if WZI.get((i,j), 0) != 0:
                        z = 0
                        z1 = 0
                        if i+1 > vetvi:
                            u = -1
                        else:
                            u = 1

                        if i == j:
                            ll = 5
                            ysl=((WHX[(i,i)],WHX[(i,j)]),(WF[(i,i)],WF[(i,j)]),(WR[(i,i)],WR[(i,j)]),\
                            (WT1[(i,i)],WT1[(i,j)]),(WT2[(i,i)],WT2[(i,j)]),WL[(i,i)],WS[(i,j)],u,True)
                            SpEmk[i][0]=ysl
                        else:
                            ll = 10
                            ysl=((WHX[(i,i)],WHX[(i,j)]),(WF[(i,i)],WF[(i,j)]),(WR[(i,i)],WR[(i,j)]),\
                            (WT1[(i,i)],WT1[(i,j)]),(WT2[(i,i)],WT2[(i,j)]),WL[(i,i)],WS[(i,j)],u,False)
                            SpEmk[i][1].append(ysl)
                            SpVz[i].append((i,j))
                       
                        if ysl not in d_matZ:                    
                            XX = sparse.lil_matrix((ll,1),dtype=np.float64)
                            HH = sparse.lil_matrix((ll,1),dtype=np.float64)
                            RR = sparse.lil_matrix((ll,1),dtype=np.float64)
                            # Формируем столбцы координат фаз и их сечения
                            for k in range(3):
                                XX[k,0] = VF(WHX[(i,i)],WF[(i,i)],'X'+str(k),Katalog,Kat)
                                HH[k,0] = VF(WHX[(i,i)],WF[(i,i)],'Y'+str(k),Katalog,Kat)
                                RR[k,0] = Kat[Katalog.get(WR[(i,i)])+str(18)].value
                            if WT1[(i,i)] != 0:
                                k +=1
                                z +=1
                                XX[k,0] = VF(WHX[(i,i)],WF[(i,i)],'XT1',Katalog,Kat)
                                HH[k,0] = VF(WHX[(i,i)],WF[(i,i)],'YT1',Katalog,Kat)
                                RR[k,0] = Kat[Katalog.get(WT1[(i,i)])+str(18)].value
                            if WT2[(i,i)] != 0:
                                k +=1
                                z +=1
                                XX[k,0] = VF(WHX[(i,i)],WF[(i,i)],'XT2',Katalog,Kat)
                                HH[k,0] = VF(WHX[(i,i)],WF[(i,i)],'YT2',Katalog,Kat)
                                RR[k,0] = Kat[Katalog.get(WT2[(i,i)])+str(18)].value

                            k +=1
                            if i != j:
                                for l in range(k,k+3):
                                        XX[l,0] = VF(WHX[(i,j)],WF[(i,j)],'X'+str(l-k),Katalog,Kat) + WS[(i,j)]*u
                                        HH[l,0] = VF(WHX[(i,j)],WF[(i,j)],'Y'+str(l-k),Katalog,Kat)
                                        RR[l,0] = Kat[Katalog.get(WR[(i,j)])+str(18)].value                
                                k=l
                                if WT1[(i,j)] != 0:
                                        k +=1
                                        z1 +=1
                                        XX[k,0] = VF(WHX[(i,j)],WF[(i,j)],'XT1',Katalog,Kat)+WS[(i,j)]*u
                                        HH[k,0] = VF(WHX[(i,j)],WF[(i,j)],'YT1',Katalog,Kat)
                                        RR[k,0] = Kat[Katalog.get(WT1[(i,j)])+str(18)].value
                                if WT2[(i,j)] != 0:
                                        k +=1
                                        z1 +=1
                                        XX[k,0] = VF(WHX[(i,j)],WF[(i,j)],'XT2',Katalog,Kat)+WS[(i,j)]*u
                                        HH[k,0] = VF(WHX[(i,j)],WF[(i,j)],'YT2',Katalog,Kat)
                                        RR[k,0] = Kat[Katalog.get(WT2[(i,j)])+str(18)].value
                                k +=1
                            M = sparse.lil_matrix((k,k), dtype=np.complex128)
                            #C = np.zeros((k,k), dtype=np.complex128)

                            # Формируем массив взаимных сопротивлений и ёмкостей
                            for a in range(k):
                                for b in range(a,k):
                                    if a == b:
                                        
                                        M[a,b]=jp*mu0/2/np.pi*(np.log(2*abs(HH[a,0])/RR[a,0])+2*simps(Kars_ii(sh_int,HH[a,0],RR[a,0],mp),sh_int))*WL[(i,i)]
                                        #C[a][b]=1/(eps*2*np.pi)*np.log(2*abs(HH[a][0])/RR[a][0])
                                    else:
                                        M[a,b]=M[b,a]=jp*mu0/4/np.pi*(np.log(((HH[a,0]+HH[b,0])**2+(XX[a,0]-XX[b,0])**2)/((HH[a,0]-HH[b,0])**2+(XX[a,0]-XX[b,0])**2))\
                                                +4*simps(Kars_ij(sh_int,HH[a,0],HH[b,0],XX[a,0],XX[b,0],mp),sh_int))*WL[(i,i)]
                                        #C[a][b]=C[b][a]=1/(eps*2*np.pi)*np.log((((XX[a][0]-XX[b][0])**2+(HH[a][0]+HH[b][0])**2)**0.5)\
                                                                            #/(((XX[a][0]-XX[b][0])**2+(HH[a][0]-HH[b][0])**2)**0.5))
                            #C = jp*WL[(i,i)]/2*np.linalg.inv(C)

                            if i == j:
                                a1 = 3+z
                                b1 = 3+z
                                c1 = 0
                            else:
                                a1 = 3+z
                                b1 = 3+z1
                                c1 = 3+z
                            d_matZ[ysl]=M
                            #d_matC[ysl]=C
                        else:
                            k=3
                            if WT1[(i,i)] != 0:
                                k +=1
                                z +=1
                            if WT2[(i,i)] != 0:
                                k +=1
                                z +=1
                            if i != j:
                                k+=3
                                if WT1[(i,j)] != 0:
                                        k +=1
                                        z1 +=1
                                if WT2[(i,j)] != 0:
                                        k +=1
                                        z1 +=1
                                k+=1
                            if i == j:
                                a1 = 3+z
                                b1 = 3+z
                                c1 = 0
                            else:
                                a1 = 3+z
                                b1 = 3+z1
                                c1 = 3+z
                            M=d_matZ[ysl]
                            #C=d_matC[ysl]
                            # Записываем значения взаимных сопротивлений в основной массив
                        for a in range(i,i+a1):
                            for b in range(j,j+b1):
                                    W[a,b]=M[a-i,b-j+c1]*(1 if i==j else WRAZV[(i,j)])                                
                            if c1 == 0 and 0<=a-i<=2:
                                    W[a,a]=W[a,a]\
                                            +Kat[Katalog.get(WR[(i,j)])+str(19)].value\
                                            *WL[(i,i)]*10**(-3)
                        if c1 == 0:
                            if W[i+4,i+4]==0:
                                W[i+4,i+4]=10**15
                            if W[i+3,i+3]==0:   
                                W[i+3,i+3]=10**15
                    
                        if WT1[(i,j)] !=0 and c1 == 0:
                            W[i+3,a]=W[i+3,i+3]\
                                    +Kat[Katalog.get(WT1[(i,j)])+str(19)].value\
                                    *WL[(i,i)]*10**(-3)

                        if WT2[(i,j)] !=0 and c1 == 0:
                            if z==1:
                                W[i+3,i+3]=W[i+3,i+3]\
                                        +Kat[Katalog.get(WT2[(i,j)])+str(19)].value\
                                        *WL[(i,i)]*10**(-3)
                            elif z==2:
                                W[i+4,i+4]=W[i+4,i+4]\
                                        +Kat[Katalog.get(WT2[(i,j)])+str(19)].value\
                                        *WL[(i,i)]*10**(-3)

                        # Записываем значения ёмкастных проводимостей в основной массив
                        
                        #for a in range(YIn[(i,i)],YIn[(i,i)]+a1):
                            #for b in range(YIn[(i,j)],YIn[(i,j)]+b1):
                                    #Yu[a][b]=Yu[a][b]+C[a-YIn[(i,i)]][b-YIn[(i,j)]+c1]
                        #for a in range(YIk[(i,i)],YIk[(i,i)]+a1):
                            #for b in range(YIk[(i,j)],YIk[(i,j)]+b1):
                                    #Yu[a][b]=Yu[a][b]+C[a-YIk[(i,i)]][b-YIk[(i,j)]+c1]
                        
                        if WT1[(i,j)] ==0 and c1 == 0:     
                            Yu[YIn[(i,i)]+3,YIn[(i,j)]+3]+=10**(-15)
                            Yu[YIk[(i,i)]+3,YIk[(i,j)]+3]+=10**(-15)
                        if WT2[(i,j)] ==0 and c1 == 0:
                            Yu[YIn[(i,i)]+4,YIn[(i,j)]+4]+=10**(-15)
                            Yu[YIk[(i,i)]+4,YIk[(i,j)]+4]+=10**(-15)
                        
                # Преобразовываем списки в кортежи, чтобы над ними можно было проводить операции сравнения
                SpEmk[i][1]=tuple(SpEmk[i][1]) 
                SpEmk[i]=tuple(SpEmk[i])

    # Функция формирования массивов исходных данных для расчета взаимной ёмкости
    def Param(whx,wf,wr,wt1,wt2,ws,u,xx,hh,rr):
        b=len(xx)
        for k in range(b,b+3):
            xx.append(VF(whx,wf,'X'+str(k-b),Katalog,Kat) + ws*u)
            hh.append(VF(whx,wf,'Y'+str(k-b),Katalog,Kat))
            rr.append(Kat[Katalog.get(wr)+str(18)].value)                
        if wt1 != 0:
            k+=1
            xx.append(VF(whx,wf,'XT1',Katalog,Kat)+ws*u)
            hh.append(VF(whx,wf,'YT1',Katalog,Kat))
            rr.append(Kat[Katalog.get(wt1)+str(18)].value)
        if wt2 != 0:
            k+=1
            xx.append(VF(whx,wf,'XT2',Katalog,Kat)+ws*u)
            hh.append(VF(whx,wf,'YT2',Katalog,Kat))
            rr.append(Kat[Katalog.get(wt2)+str(18)].value)
        k+=1
        return xx, hh, rr, b, k
    
    # Расчитываем ёмкостную проводимость для всего участка сдлижения             
    sp=list(SpEmk.keys())
    #print(1)
    d_m={}
    #d_matC={}
    if not Bool_hesh[0]:
        for i in range(len(sp)):
            if SpEmk[sp[i]] not in d_matC:
                xx=[]
                hh=[]
                rr=[]
                m=[]
                ((whx1,whx2),(wf1,wf2),(wr1,wr2),(wt11,wt12),(wt21,wt22),wl,ws,u,qq)=SpEmk[sp[i]][0]
                xx,hh,rr,a,b=Param(whx1,wf1,wr1,wt11,wt21,0,u,xx,hh,rr)
                m.append((a,b))
                for j in range(len(SpEmk[sp[i]][1])):
                    ((whx1,whx2),(wf1,wf2),(wr1,wr2),(wt11,wt12),(wt21,wt22),wl,ws,u,qq)=SpEmk[sp[i]][1][j]
                    xx,hh,rr,a,b=Param(whx2,wf2,wr2,wt12,wt22,ws,u,xx,hh,rr)
                    m.append((a,b))   
                XX=np.array(xx,dtype=np.float64)
                HH=np.array(hh,dtype=np.float64)
                RR=np.array(rr,dtype=np.float64)
                C =sparse.lil_matrix((np.shape(XX)[0],np.shape(XX)[0]), dtype=np.complex128)
                
                for a in range(np.shape(XX)[0]):
                    for b in range(a,np.shape(XX)[0]):
                        if a == b:
                            C[a,b]=1/(eps*2*np.pi)*np.log(2*HH[a]/RR[a])
                        else:
                            if (XX[a]-XX[b])**2+(HH[a]-HH[b])**2==0:
                                XX[b]+=0.01
                            C[a,b]=C[b,a]=1/(eps*2*np.pi)*np.log((((XX[a]-XX[b])**2+(HH[a]+HH[b])**2)**0.5)\
                                                                /(((XX[a]-XX[b])**2+(HH[a]-HH[b])**2)**0.5))

                C = jp*wl/2*linalg.inv(C.tocsc())
                C = C[:m[0][1],:np.shape(XX)[0]]
                d_matC[SpEmk[sp[i]]]=C
                d_m[SpEmk[sp[i]]]=m
            else:
                C=d_matC[SpEmk[sp[i]]]
                m=d_m[SpEmk[sp[i]]]

            for j in range(len(m)):
                Yu[YIn[SpVz[sp[i]][0]]:YIn[SpVz[sp[i]][0]]+m[0][1],YIn[SpVz[sp[i]][j]]:YIn[SpVz[sp[i]][j]]+m[j][1]-m[j][0]]+=C[:m[0][1],m[j][0]:m[j][1]]
                Yu[YIk[SpVz[sp[i]][0]]:YIk[SpVz[sp[i]][0]]+m[0][1],YIk[SpVz[sp[i]][j]]:YIk[SpVz[sp[i]][j]]+m[j][1]-m[j][0]]+=C[:m[0][1],m[j][0]:m[j][1]]
        
    del d_matZ, d_matC, d_m

    if not Bool_hesh[0]:
        sparse.save_npz('Hesh_files/Yu'+str(File_hesh[0])+'.npz', Yu.tocsc())
    else:
        Yu = sparse.load_npz('Hesh_files/Yu'+str(File_hesh[0])+'.npz').tolil()
   

     
    # Создаём шаблоны проводимостей для заземления
    a=10**6
    b=-10**6
    # Шаблон для заземления 3х фаз с тросами и без
    Y3str=sparse.lil_matrix([[a,b,0,0,0],[b,4*a,b,b,b],[0,b,a,0,0],[0,b,0,a,0],[0,b,0,0,a]],dtype=np.float64)
    Y3btr=sparse.lil_matrix([[a,b,0,0,0],[b,2*a,b,0,0],[0,b,a,0,0],[0,0,0,0,0],[0,0,0,0,0]],dtype=np.float64)
    # Шаблон для заземления 2х фаз с тросами
    Y21str=sparse.lil_matrix([[3*a,b,0,b,b],[b,a,0,0,0],[0,0,0,0,0],[b,0,0,a,0],[b,0,0,0,a]],dtype=np.float64)
    Y22str=sparse.lil_matrix([[3*a,0,b,b,b],[0,0,0,0,0],[b,0,a,0,0],[b,0,0,a,0],[b,0,0,0,a]],dtype=np.float64)
    Y23str=sparse.lil_matrix([[0,0,0,0,0],[0,3*a,b,b,b],[0,b,a,0,0],[0,b,0,a,0],[0,b,0,0,a]],dtype=np.float64)
    # Шаблон для заземления 2х фаз без тросов
    Y21btr=sparse.lil_matrix([[a,b,0,0,0],[b,a,0,0,0],[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0]],dtype=np.float64)
    Y22btr=sparse.lil_matrix([[a,0,b,0,0],[0,0,0,0,0],[b,0,a,0,0],[0,0,0,0,0],[0,0,0,0,0]],dtype=np.float64)
    Y23btr=sparse.lil_matrix([[0,0,0,0,0],[0,a,b,0,0],[0,b,a,0,0],[0,0,0,0,0],[0,0,0,0,0]],dtype=np.float64)
    # Шаблон для заземления 1й фазы с тросами
    Y11str=sparse.lil_matrix([[2*a,0,0,b,b],[0,0,0,0,0],[0,0,0,0,0],[b,0,0,a,0],[b,0,0,0,a]],dtype=np.float64)
    Y12str=sparse.lil_matrix([[0,0,0,0,0],[0,2*a,0,b,b],[0,0,0,0,0],[0,b,0,a,0],[0,b,0,0,a]],dtype=np.float64)
    Y13str=sparse.lil_matrix([[0,0,0,0,0],[0,0,0,0,0],[0,0,2*a,b,b],[0,0,b,a,0],[0,0,b,0,a]],dtype=np.float64)
    # Шаблон для заземления 1й фазы без тросов
    Y11btr=sparse.lil_matrix([[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0]],dtype=np.float64)
    Y12btr=sparse.lil_matrix([[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0]],dtype=np.float64)
    Y13btr=sparse.lil_matrix([[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0],[0,0,0,0,0]],dtype=np.float64)
    # Шаблон для заземления тросов на ВЛ
    Y45=sparse.lil_matrix([[a,b],[b,a]],dtype=np.float64)
    
    # Раставляем зазеление на опоры ИВЛ                                
    Y1 = np.zeros((yzli,1),dtype=np.int64)
    rr = -1
    rv = {}
    k = 0
    if not Bool_hesh[1]:
        Yuz = sparse.lil_matrix((yzli0,yzli0), dtype=np.complex128) 
        for i in range((np.shape(opm))[0]):
            for j in range((np.shape(opm))[1]):
                    if opm[i][j] != 0 and opm[i][j] != rr and (opm[i][j] not in  rv):
                            Y1[k][0]=rr=opm[i][j]
                            k +=5
            rv[opm[i][0]]=1

        # Создаём функцию присвоения значений шаблона на координаты
        def Zyfunk(Yu,Y,Zyy,i,j):
            for a in range(i,i+(np.shape(Y))[0]):
                    for b in range(i,i+(np.shape(Y))[0]):
                            Yu[a,b]+=Y[a-i,b-i]
            Yu[i+j,i+j]+=1/Zyy
            return Yu

        # Дополняем матрицу проводимостей проводимостями заземления
        for i in range((np.shape(Zy))[0]):
            for j in range(0,yzli,5):
                    if Zy[i][0]==Y1[j][0]:
                            # 3ф заземление с тросами
                            if Zy[i][2]==1:
                                    Yuz=Zyfunk(Yuz,Y3str,Zy[i][1],j,1)
                                    Y1[j][0]=0
                            # 3ф заземление без тросов
                            elif Zy[i][2]==2:
                                    Yuz=Zyfunk(Yuz,Y3btr,Zy[i][1],j,1)
                                    Y1[j][0]=0
                            # 2ф заземление с тросами
                            elif Zy[i][2]==3 and Zy[i][3]==1 and Zy[i][4]==2:
                                    Yuz=Zyfunk(Yuz,Y21str,Zy[i][1],j,0)
                                    Y1[j][0]=0
                            elif Zy[i][2]==3 and Zy[i][3]==1 and Zy[i][4]==3:
                                    Yuz=Zyfunk(Yuz,Y22str,Zy[i][1],j,0)
                                    Y1[j][0]=0
                            elif Zy[i][2]==3 and Zy[i][3]==2 and Zy[i][4]==3:
                                    Yuz=Zyfunk(Yuz,Y23str,Zy[i][1],j,1)
                                    Y1[j][0]=0
                            # 2ф заземление без тросов
                            elif Zy[i][2]==4 and Zy[i][3]==1 and Zy[i][4]==2:
                                    Yuz=Zyfunk(Yuz,Y21btr,Zy[i][1],j,0)
                                    Y1[j][0]=0
                            elif Zy[i][2]==4 and Zy[i][3]==1 and Zy[i][4]==3:
                                    Yuz=Zyfunk(Yuz,Y22btr,Zy[i][1],j,0)
                                    Y1[j][0]=0
                            elif Zy[i][2]==4 and Zy[i][3]==2 and Zy[i][4]==3:
                                    Yuz=Zyfunk(Yuz,Y23btr,Zy[i][1],j,1)
                                    Y1[j][0]=0
                            # 1ф заземление с тросами       
                            elif Zy[i][2]==5 and Zy[i][3]==1:
                                    Yuz=Zyfunk(Yuz,Y11str,Zy[i][1],j,0)
                                    Y1[j][0]=0
                            elif Zy[i][2]==5 and Zy[i][3]==2:
                                    Yuz=Zyfunk(Yuz,Y12str,Zy[i][1],j,1)
                                    Y1[j][0]=0
                            elif Zy[i][2]==5 and Zy[i][3]==3:
                                    Yuz=Zyfunk(Yuz,Y13str,Zy[i][1],j,2)
                                    Y1[j][0]=0
                            # 1ф заземление без тросов
                            elif Zy[i][2]==6 and Zy[i][3]==1:
                                    Yuz=Zyfunk(Yuz,Y11btr,Zy[i][1],j,0)
                                    Y1[j][0]=0
                            elif Zy[i][2]==6 and Zy[i][3]==2:
                                    Yuz=Zyfunk(Yuz,Y12btr,Zy[i][1],j,1)
                                    Y1[j][0]=0
                            elif Zy[i][2]==6 and Zy[i][3]==3:
                                    Yuz=Zyfunk(Yuz,Y13btr,Zy[i][1],j,2)
                                    Y1[j][0]=0
                        
        # Заземляем троссы на опорах ИВЛ
        for i in range((np.shape(Zytr))[0]):
            for j in range(0,yzli,5):
                    if Zytr[i][0] <= Y1[j][0] <=Zytr[i][1]:
                            Yuz=Zyfunk(Yuz,Y45,Zytr[i][2],j+3,0)
                            Y1[j][0]=0
        
        # Заземляем троссы на опора ВВЛ
        k = yzli
        for v in range((np.shape(vv))[0]-1):
            for c in range(Prolet[v]+1):
                if np.shape(Zvl)[0]-1 >= v:
                    Yuz=Zyfunk(Yuz,Y45,Zvl[v][0],k+3,0)
                    k +=5
            k +=1

        sparse.save_npz('Hesh_files/Yz'+str(File_hesh[1])+'.npz', Yuz.tocsc())
    else:
        Yuz = sparse.load_npz('Hesh_files/Yz'+str(File_hesh[1])+'.npz').tolil()
    
    # Заземление нейтрали
    k = yzli
    for v in range((np.shape(vv))[0]-1):
        for c in range(Prolet[v]+1):
                k +=5
        if True:#NotSim[v]==1
            #print("Зазамлено")
            Yu[k,k]=+10**(6) # 2
        k +=1
        
    
    
    # Функция эквивалентирования хвостов ВВЛ
    def Ekv1(Zvl,v,nk,t):
        l0=Zvl[v][nk*2+1]/Zvl[v][nk*2+2]
        H=VF(Zvl[v][6],Zvl[v][5],'YT'+str(t+1),Katalog,Kat)
        R=Kat[Katalog.get(Zvl[v][7+t])+str(18)].value
        z0=Kat[Katalog.get(Zvl[v][7+t])+str(19)].value
        Z=(z0+1000*jp*mu0/2/np.pi*(np.log(2*H/R)+2*simps(Kars_ii(sh_int,H,R,mp),sh_int)))*l0
        Ee=np.zeros((int(Zvl[v][nk*2+2])), dtype=np.complex128)
        Ze=np.zeros((int(Zvl[v][nk*2+2])), dtype=np.complex128)
        Ee[0]=1
        Ze[0]=Z+Zvl[v][0]
        if Zvl[v][nk*2+2] >1:
              for i in range(1,int(Zvl[v][nk*2+2])):
                      Ee[i]=(Ee[i-1]*Zvl[v][0])/(Ze[i-1]+Zvl[v][0])+1
                      Ze[i]=(Ze[i-1]*Zvl[v][0])/(Ze[i-1]+Zvl[v][0])
        p=Ee[i]/(i+1)
        return Ze[i], p

    # Функция для формирования матрици взаим сопр для экв участка
    def Prisv(Zvl,v,nk):
        XX = np.zeros((5,1))
        HH = np.zeros((5,1))
        RR = np.zeros((5,1))
        for k in range(3):
                XX[k][0] = VF(Zvl[v][6],Zvl[v][5],'X'+str(k),Katalog,Kat)
                HH[k][0] = VF(Zvl[v][6],Zvl[v][5],'Y'+str(k),Katalog,Kat)
                RR[k][0] = Kat[Katalog.get(Zvl[v][7])+str(18)].value
        if Zvl[v][7] != 0:
                k +=1
                XX[k][0] = VF(Zvl[v][6],Zvl[v][5],'XT1',Katalog,Kat)
                HH[k][0] = VF(Zvl[v][6],Zvl[v][5],'YT1',Katalog,Kat)
                RR[k][0] = Kat[Katalog.get(Zvl[v][7])+str(18)].value
        if Zvl[v][8] != 0:
                k +=1
                XX[k][0] = VF(Zvl[v][6],Zvl[v][5],'XT2',Katalog,Kat)
                HH[k][0] = VF(Zvl[v][6],Zvl[v][5],'YT2',Katalog,Kat)
                RR[k][0] = Kat[Katalog.get(Zvl[v][8])+str(18)].value
        k +=1

        M = sparse.lil_matrix((k,k), dtype=np.complex128)

                # Формируем массив взаимных сопротивлений и ёмкостей

        for a in range(k):
                for b in range(a,k):
                        if a == b:
                                M[a,b]=jp*mu0/2/np.pi*(np.log(2*HH[a][0]/RR[a][0])+2*simps(Kars_ii(sh_int,HH[a][0],RR[a][0],mp),sh_int))*1000

                        else:
                                M[a,b]=M[b,a]=jp*mu0/4/np.pi*(np.log(((HH[a][0]+HH[b][0])**2+(XX[a][0]-XX[b][0])**2)/((HH[a][0]-HH[b][0])**2+(XX[a][0]-XX[b][0])**2))\
                                               +4*simps(Kars_ij(sh_int,HH[a][0],HH[b][0],XX[a][0],XX[b][0],mp),sh_int))*1000

        return M
                                 

    # Начинаем эквивалентировать хвосты ВВЛ

    k=vetvi
    if not Bool_hesh[0]:
        for v in range(np.shape(Zvl)[0]):
            if Zvl[v][1] !=0:
                    r=np.shape(Prisv(Zvl,v,0))
                    M = sparse.lil_matrix(r, dtype=np.complex128)
                    if (4,4)==r:
                            M[:,3]=1
                            M[3,:]=1
                    else:
                            M[:,3]=1
                            M[3,:]=1
                            M[:,4]=1
                            M[4,:]=1
                    M=M*Prisv(Zvl,v,0)
                    P1=Ekv1(Zvl,v,0,0)[1]
                    Ze1=Ekv1(Zvl,v,0,0)[0]
                    if Zvl[v][8] !=0:
                        P2=Ekv1(Zvl,v,0,1)[1]
                        Ze2=Ekv1(Zvl,v,0,1)[0]
                    for i in range(r[0]):
                            for j in range(r[1]):
                                    if i==j and i==3 and r[0]==4:
                                            W[k+i,k+j]=Ze1
                                    elif i==j and i==4 and r[0]==5:
                                            W[k+i,k+j]=Ze2
                                    elif i != j and (i==3 or j==3):
                                            W[k+i,k+j]=P1*M[i,j]
                                    elif i != j and (i==4 or j==4):
                                            W[k+i,k+j]=P1*M[i,j]
            k +=Prolet[v]*5+5
            if Zvl[v][3] !=0:
                    r=np.shape(Prisv(Zvl,v,1))
                    M = np.zeros(r,dtype=np.float64)
                    if (4,4)==r:
                            M[:,3]=1
                            M[3,:]=1
                    else:
                            M[:,3]=1
                            M[3,:]=1
                            M[:,4]=1
                            M[4,:]=1
                    M=M*Prisv(Zvl,v,1)
                    P1=Ekv1(Zvl,v,1,0)[1]
                    Ze1=Ekv1(Zvl,v,1,0)[0]
                    if Zvl[v][8] !=0:
                        P2=Ekv1(Zvl,v,1,1)[1]
                        Ze2=Ekv1(Zvl,v,1,1)[0]
                    for i in range(r[0]):
                            for j in range(r[1]):
                                    if i==j and i==3 and r[0]==4:
                                            W[k+i,k+j]=Ze1
                                    elif i==j and i==4 and r[0]==5:
                                            W[k+i,k+j]=Ze2
                                    elif i != j and (i==3 or j==3):
                                            W[k+i,k+j]=P1*M[i,j]
                                    elif i != j and (i==4 or j==4):
                                            W[k+i,k+j]=P1*M[i,j]
            k +=5
    
        sparse.save_npz('Hesh_files/W'+str(File_hesh[0])+'.npz', W.tocsc())
    else:
        W = sparse.load_npz('Hesh_files/W'+str(File_hesh[0])+'.npz').tolil()
                                

                    
        


    # Начинаем матричные преобразования для составления системы уравнений
    dZ=sparse.lil_matrix((vetvi0, vetvi0),  dtype=np.complex128)
    W+=Wn
    # Создаём диаганальную матрицу собственных сопротивлений
    dZ = sparse.diags(W.diagonal(), format='lil')
    W -= dZ
    # Получаем матрицу соботвенных проводимостей
    dZ = dZ.tocsc()
    #print(dZ)
    dY=linalg.inv(dZ)
    
    #Получаем матрицу узловых проводимостей
    A = A.tocsc()
    At = A.transpose()
    Yy=(A.dot(dY)).dot(At)
    # Добавляем учет емкости и заземления
    Yy = Yy.tolil()
    Yy+=Yu+Yuz
    
    # Составдяем вектор столбец для решения системы уравнений
    E = E.tocsc()
    B=(A.dot(dY)).dot(E)

    JE=sparse.vstack([B, E],format='csr')
    # Высчитываем матрицу коэффициентов по току
    KI=A.dot(((W.tocsc()).dot(dY)).transpose())

    # Составляем саму систему уравнений
    #W+=dZ

    
    YZ=sparse.vstack([sparse.hstack([Yy, KI]), sparse.hstack([At, W+dZ])],format='csr')
    #del Yy,KI,At,W
    
    tic = time()
    # Запускаем функцию решения системы линейных алгебраических уравнений
    H=linalg.spsolve(YZ,JE)

    if True:
        k=-1
        while k<20:
            e=0
            k+=1
            for i in range(np.shape(vvleds)[0]):
                S1=vvleds[i][3]+H[yzli+vvlnagr[i][0]]*np.conjugate(H[yzli0+vetvi+vvlnagr[i][1]])
                S2=vvleds[i][4]+H[yzli+vvlnagr[i][0]+1]*np.conjugate(H[yzli0+vetvi+vvlnagr[i][1]+1])
                S3=vvleds[i][5]+H[yzli+vvlnagr[i][0]+2]*np.conjugate(H[yzli0+vetvi+vvlnagr[i][1]+2])
                e = max(e,abs(S1),abs(S2),abs(S3))

                zn1 = (abs(H[yzli+vvlnagr[i][0]])**2)/np.conjugate(vvleds[i][3]) if abs(vvleds[i][3])!=0 else vvleds[i][6]
                zn2 = (abs(H[yzli+vvlnagr[i][0]+1])**2)/np.conjugate(vvleds[i][4]) if abs(vvleds[i][4])!=0 else vvleds[i][7]
                zn3 = (abs(H[yzli+vvlnagr[i][0]+2])**2)/np.conjugate(vvleds[i][5]) if abs(vvleds[i][5])!=0 else vvleds[i][8]

                dZ[vetvi+vvlnagr[i][1],vetvi+vvlnagr[i][1]]+=zn1-vvleds[i][6]
                dZ[vetvi+vvlnagr[i][1]+1,vetvi+vvlnagr[i][1]+1]+=zn2-vvleds[i][7]
                dZ[vetvi+vvlnagr[i][1]+2,vetvi+vvlnagr[i][1]+2]+=zn3-vvleds[i][8]

                dY[vetvi+vvlnagr[i][1],vetvi+vvlnagr[i][1]]+=1/zn1-1/vvleds[i][6]
                dY[vetvi+vvlnagr[i][1]+1,vetvi+vvlnagr[i][1]+1]+=1/zn2-1/vvleds[i][7]
                dY[vetvi+vvlnagr[i][1]+2,vetvi+vvlnagr[i][1]+2]+=1/zn3-1/vvleds[i][8]

                az = np.array([[1/vvleds[i][6],0,0,0,0,-1/vvleds[i][6]],\
                                [0,1/vvleds[i][7],0,0,0,-1/vvleds[i][7]],\
                                [0,0,1/vvleds[i][8],0,0,-1/vvleds[i][8]],\
                                [0,0,0,0,0,0],\
                                [0,0,0,0,0,0],\
                                [-1/vvleds[i][6],-1/vvleds[i][7],-1/vvleds[i][8],0,0,\
                                    1/vvleds[i][6]+1/vvleds[i][7]+1/vvleds[i][8]]], dtype=np.complex128)

                nz = np.array([[1/zn1,0,0,0,0,-1/zn1],\
                                [0,1/zn2,0,0,0,-1/zn2],\
                                [0,0,1/zn3,0,0,-1/zn3],\
                                [0,0,0,0,0,0],\
                                [0,0,0,0,0,0],\
                                [-1/zn1,-1/zn2,-1/zn3,0,0,\
                                    1/zn1+1/zn2+1/zn3]], dtype=np.complex128)


                Yy[yzli+vvlnagr[i][0]:yzli+vvlnagr[i][0]+6,yzli+vvlnagr[i][0]:yzli+vvlnagr[i][0]+6]+=nz-az

                vvleds[i][6]=zn1
                vvleds[i][7]=zn2
                vvleds[i][8]=zn3

            print("итерация:",k)
            if e<100000: break
            KI=A.dot(((W.tocsc()).dot(dY)).transpose())
            YZ=sparse.vstack([sparse.hstack([Yy, KI]), sparse.hstack([At, W+dZ])],format='csr')
            H=linalg.spsolve(YZ,JE)

    del Yy,KI,At,W



    #vetvi0 = vetvi + vetvi1
    #yzli0 = yzli + yzli1

    tuc = time()
    print("Решение СЛАУ", tuc-tic)
    # Опребеляем модули значений в векторе столбце
    Hm=np.abs(H)

    """ dPtros = 0
    dPwires = 0
    for i in range(yzli0+vetvi+5,np.shape(H)[0]-5,5):
        #print(Hm[i+3])
        dPtros+=(Hm[i+3]**2)*np.real(YZ[i+3,i+3])
        dPwires+=(Hm[i]**2)*np.real(YZ[i,i])
        dPwires+=(Hm[i+1]**2)*np.real(YZ[i+1,i+1])
        dPwires+=(Hm[i+2]**2)*np.real(YZ[i+2,i+2])
    print("dPtros=",dPtros)
    print("dPwires=",dPwires) """

    # Создаём функцию для формирования матриц потенциалов в фазах и тросах
    def faz1(Hm,h):
        rr={}
        k=1
        Fi=np.zeros(np.shape(opm), dtype=np.float64)
        for i in range(np.shape(opm)[0]):
                for j in range(np.shape(opm)[1]):
                    if opm[i][j] not in rr and opm[i][j] != 0:
                        Fi[i][j]=Hm[k*5-h-1]
                        rr[opm[i][j]] = Hm[k*5-h-1]
                        k +=1
                    elif opm[i][j] in rr and opm[i][j] != 0:
                        Fi[i][j] = rr.get(opm[i][j])
                               
        return Fi
    # Инвертируем словарь номеров опор
    key = list(per.keys())
    inv_per = {}
    for i in key:      
    	a = {v:k for k, v in per[i].items()}
    	inv_per[i]=a
    
    # Создаём функцию индексов опор
    
    def pp(opm):
        a=[]
        for i in range(np.shape(opm)[0]):
                b=[]
                for j in range(np.shape(opm)[1]):
                        if opm[i][j] in inv_per[key[i]]:
                                b.append(inv_per[key[i]][opm[i][j]])
                a.append(b)
                                

        return a
    

# Формируем списки значений для построения графиков
    FiA = faz1(Hm,4)
    FiB = faz1(Hm,3)
    FiC = faz1(Hm,2)
    FiT1 = faz1(Hm,1)
    l = np.shape(FiA)
    point = pp(opm)
    cols = np.array([len(i) for i in point], dtype=np.int64)
    p = np.array([point[i]+[0]*(l[1]-cols[i]) for i in range(l[0])], dtype=np.int64)

    if trig_r:
        np.save('Hesh_files/FiA'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy',FiA)
        np.save('Hesh_files/FiB'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy',FiB)
        np.save('Hesh_files/FiC'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy',FiC)
        np.save('Hesh_files/FiT1'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy',FiT1)
        np.save('Hesh_files/p'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy',p)
        np.save('Hesh_files/cols'+str(File_hesh[0])+'_'+str(File_hesh[1])+'_'+str(File_hesh[2])+'.npy',cols)

    return [list(FiA[i,:cols[i]]*koef_graf) for i in range(l[0])],\
            [list(FiB[i,:cols[i]]*koef_graf) for i in range(l[0])],\
            [list(FiC[i,:cols[i]]*koef_graf) for i in range(l[0])],\
            [list(FiT1[i,:cols[i]]*koef_graf) for i in range(l[0])],\
            [list(p[i,:cols[i]]) for i in range(l[0])]
    



                                
                                
                                

                
            
                    



               
                

                                                        



                





    
        
        
        
        
