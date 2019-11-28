# pylint: disable=E0611
# pylint: disable=E1101
from PyQt5.QtWidgets import QSplashScreen, QApplication
from PyQt5.QtGui import QPixmap, QFont
from PyQt5.QtCore import Qt
import sys


# Выводим заствку
app=QApplication(sys.argv)
splash=QSplashScreen(QPixmap('Pictures/splash.jpg'))
splash.show()

import socket
from time import time, sleep
try:
    #socket.gethostbyaddr('local.besp.by')
    1
except socket.gaierror:
    font = splash.font()
    font.setPixelSize(16)
    font.setWeight(QFont.Bold)
    splash.setFont(font)
    splash.showMessage("У вас недостаточно прав для использования", \
    Qt.AlignHCenter | Qt.AlignBottom, Qt.black)
    sleep(4)
    exit()




from PyQt5.QtWidgets import QWidget, QTableWidgetItem, QTabWidget\
     ,QComboBox, QTextEdit, QAction, QFileDialog, QMainWindow, QErrorMessage\
     ,QMessageBox, QVBoxLayout,QPushButton, QProgressDialog, QLabel, QHBoxLayout\
     ,QInputDialog,  QTreeWidget, QAbstractItemView, QFrame, QScrollArea,QSizePolicy\
     , QShortcut, QCompleter,QTableWidgetSelectionRange, QCheckBox

from PyQt5.QtCore import QCoreApplication,  QSize, QPersistentModelIndex, QThread,QSortFilterProxyModel,QEvent, QTimer
from PyQt5.QtGui import QPixmap, QPalette, QBrush, QImage, QIcon, QStandardItemModel, QStandardItem\
     ,QPainter, QPaintEvent, QColor, QPen,QTextOption, QKeySequence, QTransform

from PyQt5 import QtCore 

from functools import partial

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import matplotlib.ticker
import matplotlib
from matplotlib.ticker import FuncFormatter
import matplotlib.font_manager as fm
arial_font = fm.FontProperties(fname = "Fonts/arial.ttf")

import numpy as np
from scipy.signal import argrelextrema

from PyQt5 import uic

from openpyxl import Workbook, load_workbook

import copy

import os
import pickle
import itertools
from docx import Document
import shutil


import Kat_lists as Kl
import rnn
import Sz
from Kop import KOp
import OtchetW
import SxT
import OpRis
import BD
import Gr_sblizh
import Oid
import ShMod
import FindErrors as FE


class ExtendedComboBox(QComboBox):
    """ Переопределяем поведение QComboBox, чтобы можно было в его вписать свою строку """
    def __init__(self,trig, parent=None):
        super(ExtendedComboBox, self).__init__(parent)
        self.Trig = trig

        self.setFocusPolicy(Qt.StrongFocus)
        self.setEditable(True)

        # add a filter model to filter matching items
        self.pFilterModel = QSortFilterProxyModel(self)
        self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
        self.pFilterModel.setSourceModel(self.model())

        # add a completer, which uses the filter model
        self.completer = QCompleter(self.pFilterModel, self)
        # always show all (filtered) completions
        self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
        self.setCompleter(self.completer)

        # connect signals
        self.lineEdit().textEdited[str].connect(self.pFilterModel.setFilterFixedString)
        self.completer.activated.connect(self.on_completer_activated)


    # on selection of an item from the completer, select the corresponding item from combobox 
    def on_completer_activated(self, text):
        if text:
            index = self.findText(text)
            self.setCurrentIndex(index)
            self.activated[str].emit(self.itemText(index))


    # on model change, update the models of the filter and completer as well 
    def setModel(self, model):
        super(ExtendedComboBox, self).setModel(model)
        self.pFilterModel.setSourceModel(model)
        self.completer.setModel(self.pFilterModel)


    # on model column change, update the model column of the filter and completer as well
    def setModelColumn(self, column):
        self.completer.setCompletionColumn(column)
        self.pFilterModel.setFilterKeyColumn(column)
        super(ExtendedComboBox, self).setModelColumn(column)

    def wheelEvent(self, *args, **kwargs):
        if self.Trig["trig"]:
            self.Trig["timer"].stop()
            self.Trig["timer"].start(500)
            return QComboBox.wheelEvent(self, *args, **kwargs)

class UserComboBox(QComboBox):
    def __init__(self,trig, parent=None):
        super(UserComboBox, self).__init__(parent)
        self.Trig = trig

    def wheelEvent(self, *args, **kwargs):
        if self.Trig["trig"]:
            self.Trig["timer"].stop()
            self.Trig["timer"].start(500)
            return QComboBox.wheelEvent(self, *args, **kwargs)


# Поток одиночного расчёта
class OneRasch(QThread):
    mysignal = QtCore.pyqtSignal(str)
    def __init__(self,b, parent=None):
        QtCore.QThread.__init__(self, parent)
        self.b=b
    def run(self):
        try:
            ivl, self.b.nm_ivl, self.b.km, self.b.per, yz, self.b.ivl_gr, yr_v, self.b.KL_sp = Kl.Sort_IVL(self.b.sp_ivl)
            vvl = Kl.Sort_VVL(self.b.sp_vvl,self.b.per,yz)
            zy = Kl.Sort_Zy(self.b.sp_zy,self.b.per,yz)
            zytr = Kl.Sort_Zytr(self.b.sp_zytr,self.b.per,yz)
            zvl = Kl.Sort_Zvl(self.b.sp_zvl)

            # Хэшируем запускаемый расчёт
            k_ivl = [tuple(i) for i in ivl]
            k_vvl1 = []
            k_vvl2 = []
            for i in vvl:
                if i[0]>0: 
                    k_vvl1.append(tuple(i))
                else: 
                    k_vvl2.append(tuple(i))
            k_zy = [tuple(i) for i in zy]
            k_zytr = [tuple(i) for i in zytr]
            k_zvl1 = [i[0] for i in ivl]
            k_zvl2 = [tuple(i[1:9]) for i in zvl]
            
            Hesh_now = [tuple(k_ivl+k_vvl2+k_zvl2+[self.b.pz])]
            k_1 = self.b.Hesh[0].get(Hesh_now[0],len(self.b.Hesh[0]))
            Hesh_now.append(tuple(k_zy+k_zytr+k_zvl1+[k_1]))
            Hesh_now.append(tuple(k_vvl1+[k_1]))
            Kesh = [self.b.kesh_1,self.b.kesh_2,self.b.kesh_3]
            Bool = [Hesh_now[i] in self.b.Hesh[i] for i in range(3)]
            Bool_hesh = [Bool[i] and Kesh[i] for i in range(3)]
            File_hesh = [self.b.Hesh[i].setdefault(Hesh_now[i],len(self.b.Hesh[i])) for i in range(3)]
            print(Bool_hesh)
            print(File_hesh)
            
            
            # Запускаем расчет, получаем результаты
            tic = time()
            try:
                self.b.FiA, self.b.FiB, self.b.FiC, self.b.FiT1, self.b.p\
                = rnn.Vvod_inf(ivl, vvl, zy, zytr, zvl, self.b.per, self.b.k_yvgr,self.b.pz,Bool_hesh,File_hesh,self.b.kesh_4)
            except Exception as ex:
                if not Bool[0]:
                    del self.b.Hesh[0][Hesh_now[0]]
                if not Bool[1]:
                    del self.b.Hesh[1][Hesh_now[1]]
                if not Bool[2]:
                    del self.b.Hesh[2][Hesh_now[2]]
                raise Exception(str(ex))

            tac = time()
            print(tac-tic)
            
            if not self.b.FTRCh:
                for i in range(len(self.b.FiT1)):
                    for j in range(len(self.b.FiT1[i])):
                        self.b.FiT1[i][j]=0
            
            #Проверяем обьеденяем ли ветви
            self.b.marker=True
            for i in yr_v:
                if i =='Нет':
                    self.b.marker=False
                    break
           
            # Если заданы уровни обьеденяем ветви
            self.b.yr_v_d = False
            if self.b.marker:
                self.b.yr_v_d=Kl.SpSoed(yr_v)
                
                self.b.FiA=Kl.SoedV(self.b.yr_v_d,self.b.FiA)
                self.b.FiB=Kl.SoedV(self.b.yr_v_d,self.b.FiB)
                self.b.FiC=Kl.SoedV(self.b.yr_v_d,self.b.FiC)
                self.b.FiT1=Kl.SoedV(self.b.yr_v_d,self.b.FiT1)
                self.b.p=Kl.SoedV(self.b.yr_v_d,self.b.p)
                self.b.km=Kl.SoedV(self.b.yr_v_d,self.b.km)
                
                self.b.nm_ivl=Kl.ReName(self.b.yr_v_d,self.b.nm_ivl)

            
            
            # Подготавливаем исходники для схем заземления и рисуем их
            self.per_name = Kl.Yz_p(self.b.ivl_gr,self.b.yr_v_d)
            sp_pit = Kl.PodpSz(self.b.ivl_gr, self.b.yr_v_d)
            if self.b.marker:
                sp_b, per=Kl.RePer(self.b.yr_v_d,self.b.per)
                m, n = Kl.Otpaiki(sp_b,self.b.ivl_gr,self.b.nm_ivl)

                inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per),self.b.nm_ivl,Kl.ReZy(sp_b,self.b.sp_zy))

                f, f1, f2 = Kl.Zapret(self.b.FiA, self.b.FiB, self.b.FiC,self.b.FiT1,self.b.p,self.b.sr_vel,self.b.FACh,self.b.FBCh,self.b.FCCh,self.b.FTRCh)#,lbsz,inf

                kl_list = Kl.KL_metki(self.b.KL_sp,self.b.yr_v_d)
                self.per_name = Kl.UserReOp(self.b.KL_sp,self.per_name,self.b.yr_v_d,self.b.per)
                re_op_key = list(self.per_name.keys())

                for i in range(len(self.b.nm_ivl)):
                    Sz.Sxeme(inf[i],f1[i],lbsz[i],m[i],n[i],self.b.grop,self.b.zpps,self.b.dz,self.per_name[re_op_key[i]],sp_pit[i],kl_list[i])

 
            else:
                inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(self.b.per),self.b.nm_ivl,self.b.sp_zy)
                f, f1, f2 = Kl.Zapret(self.b.FiA, self.b.FiB, self.b.FiC,self.b.FiT1,self.b.p, self.b.sr_vel,self.b.FACh,self.b.FBCh,self.b.FCCh,self.b.FTRCh)
                
                kl_list = Kl.KL_metki(self.b.KL_sp)
                self.per_name = Kl.UserReOp(self.b.KL_sp,self.per_name) #,self.b.per
                re_op_key = list(self.per_name.keys())

                for i in range(len(self.b.nm_ivl)):
                    Sz.Sxeme(inf[i],f1[i],lbsz[i],[],[],self.b.grop,self.b.zpps,self.b.dz,self.per_name[re_op_key[i]],sp_pit[i],kl_list[i])

            
            self.mysignal.emit('Good')
            
        except Exception as ex:
            self.mysignal.emit(str(ex))

    def SchitInf(self):
        
        return (self.b.nm_ivl, self.b.km, self.b.per, self.b.ivl_gr, self.b.FiA, self.b.FiB, self.b.FiC, self.b.FiT1, self.b.p, self.b.yr_v_d,self.b.Hesh,self.b.KL_sp,self.per_name)
            
    # остановка потока
    def stop( self ):
        print('stop')
        self.terminate()
        self.wait()
        return 'stop'

class MyFrame(QWidget):
    def __init__(self,Trig,parent=None):
        super().__init__(parent)
        self.scaleFactor = 1.0
        self.imageLabel = QLabel()
        self.imageLabel.setBackgroundRole(QPalette.Base)
        self.imageLabel.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.imageLabel.setScaledContents(True)
        self.scrollArea = QScrollArea()
        self.scrollArea.setBackgroundRole(QPalette.Dark)
        self.scrollArea.setWidget(self.imageLabel)
        self.scrollArea.setVisible(False)
        self.layout = QVBoxLayout(self)
        self.layout.addWidget(self.scrollArea)
        self.Trig = Trig
        self.zoom = 1
        self.angle = 0
        self.open()

    def zoomIn(self):
        try:
            if not self.Trig:
                self.zoom *=1.25
                self.scaleImage(self.zoom)
        except Exception:
            None
 
    def zoomOut(self):
        try:
            if not self.Trig:
                self.zoom *=0.8
                self.scaleImage(self.zoom)
        except Exception:
            None

    def Rotate(self, direction):
        try:
            self.angle += 90 * direction
            t = QTransform().rotate(self.angle)
            self.imageLabel.setPixmap(QPixmap.fromImage(self.image).transformed(t))
            self.scrollArea.setVisible(True)
            self.imageLabel.adjustSize()
            self.scaleImage(self.zoom)
            if self.Trig:
                self.scrollArea.setWidgetResizable(True)
        except Exception:
            None
 
    def normalSize(self):
        try:
            if not self.Trig:
                self.imageLabel.adjustSize()
                self.scaleFactor = min(self.x*0.77/1397,self.y*0.77/883)
                self.zoom =1
                self.scaleImage(self.zoom)
        except Exception:
            None
            
    def fitToWindow(self):
        try:
            self.Trig = not self.Trig
            self.scrollArea.setWidgetResizable(self.Trig)
            if not self.Trig:
                self.normalSize()
        except Exception:
            None

    def scaleImage(self, factor):
        try:
            self.imageLabel.resize(self.scaleFactor * factor * self.imageLabel.pixmap().size())
            self.adjustScrollBar(self.scrollArea.horizontalScrollBar(), factor)
            self.adjustScrollBar(self.scrollArea.verticalScrollBar(), factor)
        except Exception:
            None
 
    def adjustScrollBar(self, scrollBar, factor):
        try:
            scrollBar.setValue(int(factor * scrollBar.value()
                                + ((factor - 1) * scrollBar.pageStep() / 2)))
        except Exception:
            None
  
    def Update(self, size): 
        try:
            self.x = size.width()
            self.y = size.height()
            self.resize(self.x,self.y)
            self.scaleFactor = min(self.x*0.77/1397,self.y*0.77/883)
        except Exception:
            None

    def open(self):
        try:
            self.image = QImage("Pictures/sh1.jpg")
            t = QTransform().rotate(self.angle)
            self.imageLabel.setPixmap(QPixmap.fromImage(self.image).transformed(t))
            self.scrollArea.setVisible(True)
            self.imageLabel.adjustSize()
            self.scaleImage(self.zoom)
            if self.Trig:
                self.scrollArea.setWidgetResizable(True)
        except Exception:
            None



class MyWindow(QWidget):
    def __init__(self,parent=None):
        super(MyWindow,self).__init__(parent)
        uic.loadUi('Vive_window.ui',self)
        try:
            path_home = os.path.expanduser("~\\Desktop\\")
        except Exception:
            path_home = ""

        for curren_dir in ["Hesh_files","gr_sb","images_grafik","Op_Fig","result_schemes"]:
            if os.path.exists(curren_dir):
                if os.path.isdir(curren_dir):
                    print(curren_dir+" is here")
                else:
                    try:
                        os.mkdir(curren_dir)
                    except OSError:
                        print ("Error generate dir "+curren_dir)
            else:
                try:
                    os.mkdir(curren_dir)
                except OSError:
                    print ("Error generate dir "+curren_dir)
        
        self.pyt = path_home # Меню открытия и закрытия файлов открывает последнюю посещаемую директорию
        self.pytEx = path_home # Путь сохранения Excel файлов
        self.pytWIsx = path_home # Путь сохранения исходных данных
        self.pytWRez = path_home # Путь сохранения результатов
            
        # Создаём списки для выпадающих списков
        self.FF=Kl.Faza() # Фазировка
        self.TO=Kl.TipOp() # Тип опоры
        self.MR=Kl.Marka() # Марка провода
        self.MR_d={self.MR[x]: x for x in range(len(self.MR))}
        self.Katalog=rnn.ExDict()
        self.XHkat={ 'ABC': 1, 'ACB': 2, 'BCA': 3, 'BAC': 4, 'CAB': 5, 'CBA': 6,
                      '-ABC': -6, '-ACB': -3, '-BCA': -2, '-BAC': -5, '-CAB': -4, '-CBA': -1}
        self.d_to=Kl.Sl_TipOp()
        self.Katal = load_workbook(filename = 'Katal.xlsx')
        self.Kat = self.Katal['Kat']

        self.shr_gr = 14 # Шрифт надписей на графике
        self.Vive = ' - Vive 1.0' # Имя программы в титульнике
        self.Razd_simv = '-' # Символ разделения названий в именах участков
        self.zy_om.setText('30') # значения поля ЗУ БЗ Ом
        
# Пустые списки для построения графиков
        self.Tabs = [] # Создаём пустой список для закладок
        self.Figures = [] # Создаём пустой список фигур для графиков
        self.Canvas = [] # Создаём пустой список для контейнеров содержащие фигуры
        self.GrLayout = [] # Создаём пустой список для заполнения графиком экрана
        self.vkl = -1 # Индикатор вкладок

#Подгоняем размера окна
        self.setWindowTitle(self.Vive[3:])
        desktop = QApplication.desktop()
        self.resize(1000,500)
        x = (desktop.width() - self.frameSize().width())//2
        y = (desktop.height() - self.frameSize().height())//2
        self.move(x, y)
        self.showMaximized()

# Словарь Хэша вычислений
        try:
            self.Hesh  = pickle.load( open( "Hesh_files/dicts.pkl", "rb" ) )
        except Exception:
            self.Hesh = [{},{},{}]

        
        
# Назначаем кнопки     
       
        
        # Назначаем кнопку сохранения
        self.btn_save_file.clicked.connect(self.Save_tables)

        # Назначаем кнопку Открытия файла
        self.btn_open_file.clicked.connect(self.Open_tables)

        # Назначаем кнопку для выполнения расчета
        self.btn_raschet.clicked.connect(self.Raschet)

        # Назначаем кнопку для выполнения расчета
        self.btn_save_Excel.clicked.connect(self.Save_Excel)

        # Кнопка о программе
        self.prog_inf.clicked.connect(self.Prog_Inf)

        # Загружаем из базы данных настройки
        self.CheckBD()

        # Включение поиска ошибок
        self.ch_miss.stateChanged.connect(self.Check_Find_Errors)

        # Включение, обрезания граничных опор
        self.GrOp.stateChanged.connect(self.Check_GrOp)

        # Включение, запрета работы на разземлённый ПС
        self.ZpPS.stateChanged.connect(self.Check_ZpPS)
        
        # Включение отображения схемы заземления на графиках
        self.Ch_sxzgr.stateChanged.connect(self.Check_sxzgr)

        # Включение ДЗ
        self.DZ.stateChanged.connect(self.Check_DZ)

        self.Ch_A.stateChanged.connect(self.Check_FA)
        self.Ch_B.stateChanged.connect(self.Check_FB)
        self.Ch_C.stateChanged.connect(self.Check_FC)
        self.Ch_Tr.stateChanged.connect(self.Check_FTR)

        # Изменение коэффициента графика и допустимого напряжения
        self.Dop_V.editingFinished.connect(self.DopVolt)
        self.grunt.editingFinished.connect(self.Grunt)
        self.Koef_gr.editingFinished.connect(self.KefGr)
        self.Ln_A.editingFinished.connect(self.LnA)
        self.Ln_B.editingFinished.connect(self.LnB)
        self.Ln_C.editingFinished.connect(self.LnC)
        self.Ln_Tr.editingFinished.connect(self.LnTr)

        # Событие выбора вкладки
        self.lists.currentChanged.connect(self.onChange) 

        # Горячая клавиша добавить ветвь
        self.new_v = QShortcut(QKeySequence("Ctrl+W"), self)
        self.new_v.activated.connect(self.New_V) 

        # Горячая клавиша добавить участок
        self.new_y = QShortcut(QKeySequence("Ctrl+Q"), self)
        self.new_y.activated.connect(self.New_Y) 

        # Горячая клавиша удалить строку
        self.delete = QShortcut(QKeySequence("Ctrl+D"), self)
        self.delete.activated.connect(self.Delete) 

        # Горячая клавиша очистить таблицу
        self.clearT = QShortcut(QKeySequence("Ctrl+Shift+D"), self)
        self.clearT.activated.connect(self.ClearT) 

        # Горячая клавиша запустить расчёт
        self.TrRasch = QShortcut(QKeySequence("Ctrl+R"), self)
        self.TrRasch.activated.connect(self.btn_raschet.click) 

        # Горячая клавиша натуральный размер картинки
        self.PictF = QShortcut(QKeySequence("Ctrl+F"), self)
        self.PictF.activated.connect(self.natsize.click) 

        # Горячая клавиша растянуть картинку на весь экран
        self.PictS = QShortcut(QKeySequence("Ctrl+G"), self)
        self.PictS.activated.connect(self.foolwindow.click) 

        # Горячая клавиша увеличить картинку
        self.PictPlus = QShortcut(QKeySequence("Ctrl++"), self)
        self.PictPlus.activated.connect(self.CPlus)

        # Горячая клавиша уменьшить картинку
        self.PictPlus = QShortcut(QKeySequence("Ctrl+-"), self)
        self.PictPlus.activated.connect(self.CMinus)

        # Горячая клавиша сохранить в текуший файл
        self.SaveTr = QShortcut(QKeySequence("Ctrl+S"), self)
        self.SaveTr.activated.connect(self.Save1)

        # Горячая клавиша сохранить в текуший файл
        self.SaveNewTr = QShortcut(QKeySequence("Ctrl+Shift+S"), self)
        self.SaveNewTr.activated.connect(self.Save2)

        # Горячая клавиша сохранить в текуший файл
        self.OpTr = QShortcut(QKeySequence("Ctrl+O"), self)
        self.OpTr.activated.connect(self.btn_open_bd.click)

        # Горячая клавиша сохранить в текуший файл
        self.NewPTr = QShortcut(QKeySequence("Ctrl+P"), self)
        self.NewPTr.activated.connect(self.btn_newpapk_bd.click)

        # Горячая клавиша сохранить в текуший файл
        self.NewNTr = QShortcut(QKeySequence("Ctrl+N"), self)
        self.NewNTr.activated.connect(self.btn_rename_bd.click)

        # Горячая клавиша сохранить в текуший файл
        self.RotatePictRight = QShortcut(QKeySequence("Ctrl+Shift++"), self)
        self.RotatePictRight.activated.connect(lambda:self.PaintForm.Rotate(1))

        self.RotatePictLeft = QShortcut(QKeySequence("Ctrl+Shift+-"), self)
        self.RotatePictLeft.activated.connect(lambda:self.PaintForm.Rotate(-1))

        self.CoppyRow = QShortcut(QKeySequence("Ctrl+Shift+C"), self)
        self.CoppyRow.activated.connect(self.CoppyRowCels)

        self.SetCoppy = QShortcut(QKeySequence("Ctrl+Shift+V"), self)
        self.SetCoppy.activated.connect(self.SetCoppyRow)
     

        # Горячая клавиша сохранить в текуший файл
        self.Page = []
        self.Page.append(QShortcut(QKeySequence("Ctrl+1"), self))
        self.Page[0].activated.connect(lambda:self.PageSelection(1))
        self.Page.append(QShortcut(QKeySequence("Ctrl+2"), self))
        self.Page[1].activated.connect(lambda:self.PageSelection(2))
        self.Page.append(QShortcut(QKeySequence("Ctrl+3"), self))
        self.Page[2].activated.connect(lambda:self.PageSelection(3))
        self.Page.append(QShortcut(QKeySequence("Ctrl+4"), self))
        self.Page[3].activated.connect(lambda:self.PageSelection(4))
        self.Page.append(QShortcut(QKeySequence("Ctrl+5"), self))
        self.Page[4].activated.connect(lambda:self.PageSelection(5))
        self.Page.append(QShortcut(QKeySequence("Ctrl+6"), self))
        self.Page[5].activated.connect(lambda:self.PageSelection(6))
        self.Page.append(QShortcut(QKeySequence("Ctrl+7"), self))
        self.Page[6].activated.connect(lambda:self.PageSelection(7))
        self.Page.append(QShortcut(QKeySequence("Ctrl+8"), self))
        self.Page[7].activated.connect(lambda:self.PageSelection(8))
        self.Page.append(QShortcut(QKeySequence("Ctrl+9"), self))
        self.Page[8].activated.connect(lambda:self.PageSelection(9))
          
        # Назначаем кнопики вкладки ИВЛ
        self.btn_1_new_v.clicked.connect(self.new_1_v)
        self.btn_1_new_y.clicked.connect(self.new_1_y)
        self.btn_1_del.clicked.connect(self.del_1)
        self.combo_1 = Kl.PystoiSpisok() # Создаём пустой список для экземпляров QComboBox
        self.ivl_row = 0 # Первоначальное значение строк в таблице "ИВЛ"
        self.btn_1_clear.clicked.connect(lambda:self.Clearing(1))
        self.table_IVL.cellClicked[int,int].connect(partial(self.OpenFig,1))
        
        # Назначаем кнопики вкладки ВВЛ
        self.btn_2_new_v.clicked.connect(self.new_2_v)
        self.btn_2_new_y.clicked.connect(self.new_2_y)
        self.btn_2_del.clicked.connect(self.del_2)
        self.combo_2 = Kl.PystoiSpisok() # Создаём пустой список для экземпляров QComboBox
        self.vvl_row = 0 # Первоначальное значение строк в таблице "ВВЛ"
        self.num_vvl = 0 # Количество влияющих вл
        self.vvl_pos = {} # Позициия каждой ВВЛ в таблице
        self.btn_2_clear.clicked.connect(lambda:self.Clearing(2))
        self.table_VVL.cellClicked[int,int].connect(partial(self.OpenFig,2))
        
        # Назначаем кнопики вкладки "Заземление в точках"
        self.btn_3_new.clicked.connect(self.new_3)
        self.btn_3_del.clicked.connect(self.del_3)
        self.combo_3 = Kl.PystoiSpisok() # Создаём пустой список для экземпляров QComboBox
        self.zy_row = 0 # Первоначальное значение строк в таблице "Заземление в точках"
        self.btn_3_clear.clicked.connect(lambda:self.Clearing(3))

        # Назначаем кнопики вкладки "Заземление троссов опор ИВЛ"
        self.btn_4_new.clicked.connect(self.new_4)
        self.btn_4_del.clicked.connect(self.del_4)
        self.zytr_row = 0 # Первоначальное значение строк в таблице "Заземление троссов опор ИВЛ"
        self.btn_4_clear.clicked.connect(lambda:self.Clearing(4))

        # Назначаем кнопики вкладки "Заземление троссов опор ВВЛ"
        self.btn_5_new.clicked.connect(self.new_5)
        self.btn_5_del.clicked.connect(self.del_5)
        self.combo_5 = Kl.PystoiSpisok() # Создаём пустой список для экземпляров QComboBox
        self.zvl_row = 0 # Первоначальное значение строк в таблице "Заземление троссов опор ВВЛ"
        self.btn_5_clear.clicked.connect(lambda:self.Clearing(5))

        # Назначаем кнопки вкладки "Отчёт"
        self.btn_Oid.clicked.connect(self.Oid) # Кнопка исходных данных
        self.btn_S_one.clicked.connect(self.Vib_sx) # Кнопка исходных данных

        # Установка настроек пролзунков вкладки "Схема сближения"


        
# Задаём ширину столбцов таблиц        
        # Задаём ширину столбцов таблицы ИВЛ
        self.table_IVL.setHorizontalHeaderLabels(["(1)\nНазвание ветви\nИзображение опоры",
                                                    "(2)\nНаправление ветви\nДлина участка, км",
                                                    "(3)\nУзел начала\nПервая опора",
                                                    "(4)\nУзел конца\nПоследняя опора",
                                                    "(5)\nУровень\nФазировка",
                                                    "(6)\nТип опоры",
                                                    "(7)\nМарка провода",
                                                    "(8)\nМарка троса №1",
                                                    "(9)\nМарка троса №2"])

        self.ch_1_5.setCheckState(3)
        self.ch_1_6.setCheckState(3)
        self.ch_1_7.setCheckState(3)
        self.ch_1_8.setCheckState(3)
        self.ch_1_9.setCheckState(3)
               

        self.table_IVL.setColumnWidth(0,250) # Название ветви
        self.table_IVL.setColumnWidth(1,130) # Название ветви
        self.table_IVL.setColumnWidth(3,110) # Последняя опора
        self.table_IVL.setColumnWidth(7,110) # Марка тросса №1
        self.table_IVL.setColumnWidth(8,110) # Марка тросса №2
        

        # Задаём ширину столбцов таблицы ВВЛ
        self.table_VVL.setHorizontalHeaderLabels(["(1)\nНазвание ВЛ\nИзображение опоры",
                                                    "(2)\nУзел начала\nПервая опора",
                                                    "(3)\nУзел конца\nПоследняя опора",
                                                    "(4)\nU, кВ\nРастояние сближения, м",
                                                    "(5)\nР, МВт\nФазировка",
                                                    "(6)\nQ, Мвар\nТип опоры",
                                                    "(7)\nНаправл. Р\nМарка провода",
                                                    "(8)\nМарка троса №1",
                                                    "(9)\nМарка троса №2"])

        self.ch_2_5.setCheckState(3)
        self.ch_2_6.setCheckState(3)
        self.ch_2_7.setCheckState(3)
        self.ch_2_8.setCheckState(3)
        self.ch_2_9.setCheckState(3)

        self.table_VVL.setColumnWidth(0,250) # Название ветви
        self.table_VVL.setColumnWidth(2,110) # Последняя опора
        self.table_VVL.setColumnWidth(3,150) # U, кВ/Растояние сближения, м
        self.table_VVL.setColumnWidth(4,90) # P, МВт/Фазировка
        self.table_VVL.setColumnWidth(5,100) # Q,Мвар/Тип опоры
        self.table_VVL.setColumnWidth(6,110) # Направление Р/Марка провода
        self.table_VVL.setColumnWidth(7,110) # Марка тросса №1
        self.table_VVL.setColumnWidth(8,110) # Марка тросса №2

        # Задаём ширину столбцов таблицы "Заземление в точках"
        self.table_Zy.setHorizontalHeaderLabels(["(1)\nУзел начала",
                                                    "(2)\nУзел конца",
                                                    "(3)\nОпора №",
                                                    "(4)\nСопротивление заземления, Ом",
                                                    "(5)\nЗаземлено фаз, Nф",
                                                    "(6)\nФаза №1",
                                                    "(7)\nФаза №2",
                                                    "(8)\nТип заземления"])
        self.ch_3_1.setCheckState(3)
        self.ch_3_2.setCheckState(3)
        self.ch_3_3.setCheckState(3)
        self.ch_3_4.setCheckState(3)
        self.ch_3_5.setCheckState(3)
        self.ch_3_6.setCheckState(3)
        self.ch_3_7.setCheckState(3)
        self.ch_3_8.setCheckState(3)

        self.table_Zy.setColumnWidth(3,200) # Сопротивление заземлителя
        self.table_Zy.setColumnWidth(4,130) # Тип заземления

        # Задаём ширину столбцов таблицы "Заземление троссов опор ИВЛ"
        self.table_Zytr.setHorizontalHeaderLabels(["(1)\nУзел начала",
                                                    "(2)\nУзел конца",
                                                    "(3)\nПервая опора",
                                                    "(4)\nПоследняя опора",
                                                    "(5)\nСопротивление заземления, Ом"])

        self.ch_4_1.setCheckState(3)
        self.ch_4_2.setCheckState(3)
        self.ch_4_3.setCheckState(3)
        self.ch_4_4.setCheckState(3)
        self.ch_4_5.setCheckState(3)

        self.table_Zytr.setColumnWidth(3,110) # Последняя опора
        self.table_Zytr.setColumnWidth(4,200) # Сопротивление заземлителя

        self.table_Zvl.setHorizontalHeaderLabels(["(1)\nСопротивление\nзаземления, Ом",
                                                    "(2)\nДлина\nпредыдущего участка",
                                                    "(3)\nКол. опор\nпредыдущего уч.",
                                                    "(4)\nДлина\nпоследующего участка",
                                                    "(5)\nКол. опор\nпоследующего уч.",
                                                    "(6)\nФазировка",
                                                    "(7)\nТип опор",
                                                    "(8)\nМарка троса №1",
                                                    "(9)\nМарка троса №2"])

        self.ch_5_1.setCheckState(3)
        self.ch_5_2.setCheckState(3)
        self.ch_5_3.setCheckState(3)
        self.ch_5_4.setCheckState(3)
        self.ch_5_5.setCheckState(3)
        self.ch_5_6.setCheckState(3)
        self.ch_5_7.setCheckState(3)
        self.ch_5_8.setCheckState(3)
        self.ch_5_9.setCheckState(3)

        # Задаём ширину столбцов таблицы "Заземление троссов опор ВВЛ"
        self.table_Zvl.setColumnWidth(0,110) # Сопротивление заземления
        self.table_Zvl.setColumnWidth(1,130) # Длинна предыдущего участка
        self.table_Zvl.setColumnWidth(2,120) # Кол. опор предыдущего участка
        self.table_Zvl.setColumnWidth(3,140) # Длинна последующего участка
        self.table_Zvl.setColumnWidth(4,120) # Кол опор последующего участка
        self.table_Zvl.setColumnWidth(7,110) # Марка тросса №1
        self.table_Zvl.setColumnWidth(8,110) # Марка тросса №2

# Переменные для работы с каталогом
        self.kat_bd.pressed.connect(self.presseditem)


        # Кнопка открыть или двойной счелчок
        self.btn_open_bd.clicked.connect(lambda:self.OpenF(self.adres))
        self.kat_bd.doubleClicked.connect(self.OpenF)

        # Кнопка сохранить
        self.btn_save_bd.clicked.connect(self.SaveP)

        # Кнопка сохранить как
        self.btn_savek_bd.clicked.connect(self.SaveK)

        # Кнопка переименовать
        self.btn_rename_bd.clicked.connect(self.ReNam)

        # Кнопка добавления новой папки
        self.btn_newpapk_bd.clicked.connect(self.NewPapk)

        # Кнопка удаления файлов и папок
        self.btn_del_bd.clicked.connect(self.DelFP)
        
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(['Каталог ВЛ'])
        self.itemroot = self.model.invisibleRootItem()
        self.kat_list=[]
        self.kat_dict={}
        self.kat_dict_t={}
        self.kat_adr=[]
        self.kat_ind=[]

        # Переменная для хранения адреса нажатого элемента
        self.adres=None
        # Переменная для хранения адреса последнего открытого файла
        self.CurentFile=None
        
        self.KatZagr('Koren',None)
        self.kat_bd.setModel(self.model)
        self.kat_bd.setEditTriggers(QAbstractItemView.NoEditTriggers) # Запрещаем редактирование двойным счелчком

        # Подключение класса для рисования
        self.PaintForm = MyFrame(self.FoolWind,parent=self.PaintArea)
        self.ivl_ris = []
        self.vvl_ris = []
        self.zoommin.clicked.connect(self.PaintForm.zoomOut)
        self.zoommax.clicked.connect(self.PaintForm.zoomIn)
        self.foolwindow.clicked.connect(self.PaintForm.fitToWindow)
        self.natsize.clicked.connect(self.PaintForm.normalSize)
        self.title_f_s.valueChanged.connect(lambda:self.onChange(1))
        self.f_s.valueChanged.connect(lambda:self.onChange(1))

        # Загружаем размер кэша
        self.Kesh_label()
        self.btn_rl.clicked.connect(self.Kesh_label)
        self.btn_clear_kesh.clicked.connect(self.Clearing_Kesh)
        self.Kesh_VL.stateChanged.connect(self.Check_Kesh1)
        self.Kesh_ZY.stateChanged.connect(self.Check_Kesh2)
        self.Kesh_P.stateChanged.connect(self.Check_Kesh3)
        self.Kesh_R.stateChanged.connect(self.Check_Kesh4)

        # Генерация набора отчётной документации
        self.btn_all_sch.clicked.connect(self.AllSchemsStart)
        self.Ch_one_sch4.stateChanged.connect(self.Check_one_sch4)

        # Запуск генерации мест установки СЗ
        self.btn_SZ_generate.clicked.connect(self.SZGenerator)
        self.Ch_too_sz.stateChanged.connect(self.Check_too_sz)

        # Собирать отчёт в один файл
        self.Ch_one_file.stateChanged.connect(self.Check_one_file)
        self.Ch_dbl_vl.stateChanged.connect(self.Check_dbl_vl)
        self.Ch_use_templ.stateChanged.connect(self.Check_use_templ)
        self.btn_sof.clicked.connect(self.Save_docx_onefile)
        self.btn_cl_onfile.clicked.connect(self.Clear_sp_file)
        self.btn_op_temp.clicked.connect(self.Open_template)
        self.btn_ch_list.clicked.connect(self.SP_file_docx)
        self.Docx_new()
        self.show()

        # Тригер для комбобоксов
        self.cntr_pr = {"trig":False,"timer":QTimer()}
        self.cntr_pr["timer"].timeout.connect(self.on_timeout)
        self.cntr_pr["timer"].setSingleShot(True)

        # Массивы для копирования
        self.coppy_param = ["ABC","Не выбран","Не выбран","Не выбран","Не выбран"]
        self.coppy_zy = ["","","","",'3ф с тр.',"A","A",'ПЗ']
        self.coppy_zy_tr = ["","","","",""]
        self.coppy_zvl = ["30","0","0","0","0","ABC","Не выбран","Не выбран","Не выбран"]
        self.coppy_ych1 =["","",""]
        self.coppy_ych2 =["","",""]

        # Коректировка схемы заземления
        self.btn_ych_open.clicked.connect(lambda:self.CorrectionSxZ(5))
        self.btn_ych_close.clicked.connect(lambda:self.CorrectionSxZ(6))
        self.btn_pz1.clicked.connect(lambda:self.CorrectionSxZ(1))
        self.btn_pz2.clicked.connect(lambda:self.CorrectionSxZ(2))
        self.btn_pz1_zapr.clicked.connect(lambda:self.CorrectionSxZ(3))
        self.btn_pz2_zapr.clicked.connect(lambda:self.CorrectionSxZ(4))
        self.correct = [False,False,False,False,False,False]


    """ def event(self,e):
        return QWidget.event(self,e) """

    def on_timeout(self):
        self.cntr_pr["trig"] = False

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Control:
            self.cntr_pr["trig"] = True
            self.cntr_pr["timer"].start(500)
            
    def keyReleaseEvent(self,e):
        if e.key() == Qt.Key_Control:
            self.cntr_pr["trig"] = False

          
       
    def Clear_sp_file(self):
        Message = QMessageBox(QMessageBox.Question,  'Очистка списка ВЛ',
                "Вы дейстивлеьно хотите очистить список ВЛ ?", parent=self)
        Message.addButton('Да', QMessageBox.YesRole)
        Message.addButton('Нет', QMessageBox.NoRole)
        reply = Message.exec()       
        if reply == 0:
            self.Docx_new()

    def SP_file_docx(self):
        self.SP_docx=KOp()
        self.SP_docx.Spisok(self.l_vl_onedocx)
        self.SP_docx.show()

    def Docx_new(self):
        if self.use_templ:
            self.Docx_file = Document('template_word/default.docx')
        else:
            self.Docx_file = Document()
        self.d_vl_onedocx = set()
        self.l_vl_onedocx = []

    def Open_template(self):
        try:
            fname,a = QFileDialog.getOpenFileName(self, 'Загрузка шаблона', self.pyt,'*.docx') # Обрати внимание на последний элемент
            if fname != '':
                self.pyt = Kl.adres(fname)
                if a == "*.docx":
                    shutil.copy2(fname,'template_word/default.docx')
                #elif a == "*.doc":
                    
        except Exception as ex:
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Не получилось загрузить файл шаблона'+
                                'Проверьте данный файл.'+str(ex))
        else:
            self.Docx_new()
            mes = QMessageBox.information(self, 'Загрузка шаблона','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)
        


    def Save_docx_onefile(self):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить инструкцию', self.pytWRez+'Инструкция','*.docx')[0]
            self.pytWRez = Kl.adres(fname)
            
            self.Docx_file.save(fname)
        except Exception as ex:
            if str(ex) != "string index out of range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Не получилось сгенерировать инструкцию. '+
                                'В процесе генерации инструкии возникла ошибка. '+str(ex))
        else:
            mes = QMessageBox.information(self, 'Генерация инструкции','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)

    def New_V(self):
        """ Метод гоячей клавиши новой ветви Cntr+W """
        i = self.lists.currentIndex()
        if i==1: self.btn_1_new_v.click() # Искуственно нажимаем на кнопку
        elif i==2: self.btn_2_new_v.click()
        elif i==3: self.btn_3_new.click()
        elif i==4: self.btn_4_new.click()
        elif i==5: self.btn_5_new.click()
    
    def New_Y(self):
        """ Метод гоячей клавиши нового участка Cntr+Q """
        i = self.lists.currentIndex()
        if i==1: self.btn_1_new_y.click() # Искуственно нажимаем на кнопку
        elif i==2: self.btn_2_new_y.click()
    
    def Delete(self):
        """ Метод гоячей клавиши удаления строки Cntr+D """
        i = self.lists.currentIndex()
        if i==0: self.btn_del_bd.click() 
        elif i==1: self.btn_1_del.click() # Искуственно нажимаем на кнопку
        elif i==2: self.btn_2_del.click()
        elif i==3: self.btn_3_del.click()
        elif i==4: self.btn_4_del.click()
        elif i==5: self.btn_5_del.click()
    
    def ClearT(self):
        """ Метод гоячей клавиши очистки таблицы Cntr+Shift+D """
        i = self.lists.currentIndex()
        if i==1: self.btn_1_clear.click() # Искуственно нажимаем на кнопку
        elif i==2: self.btn_2_clear.click()
        elif i==3: self.btn_3_clear.click()
        elif i==4: self.btn_4_clear.click()
        elif i==5: self.btn_5_clear.click()

    def Save1(self):
        i = self.lists.currentIndex()
        if i == 8: self.btn_Oid.click()
        else: self.SaveP(self.CurentFile)

    def Save2(self):
        i = self.lists.currentIndex()
        if i == 0: self.btn_savek_bd.click()
        elif i == 8: self.btn_S_one.click()

    def PageSelection(self,i):
        try:
            self.lists.setCurrentIndex(i-1) # Делаем активной созданую закладку
        except Exception:
            1
    def CPlus(self):
        i = self.lists.currentIndex()
        if i==6: 
            try:
                j = self.TWGr.currentIndex()
                self.TWGr.setCurrentIndex(j+1)
            except Exception:
                1
        elif i==7:
            self.zoommax.click()
    def CMinus(self):
        i = self.lists.currentIndex()
        if i==6:
            try:
                j = self.TWGr.currentIndex()
                self.TWGr.setCurrentIndex(j-1)
            except Exception:
                1
        elif i==7:
            self.zoommin.click()

        
    def onChange(self,i):
        if self.lists.currentIndex() == 7 or i =="print":
            self.RisDataGenerate()
            try:
                ShMod.Ris_Sh_Sb(self.ivl_ris,self.vvl_ris, self.per_ris, self.km_ris,self.yr_op, self.f_s.value(), self.title_f_s.value(), self.name_vl.text(),self.reper_ris)
            except Exception as ex:
                print(ex)

            self.Msbl()
            
    def resizeEvent(self, event):
        if self.lists.currentIndex() == 7:
            self.Msbl()
    
    def Msbl(self):        
        try:

            self.PaintForm.Update(self.PaintArea.size())
            self.PaintForm.open()
            self.PaintArea.update()

        except Exception as ex:
            print(ex)


        
    def RisDataGenerate(self):
        try:
            self.Zapis()
            ivl, nm_ivl, km, per, yz, ivl_gr, yr_v, KL_sp = Kl.Sort_IVL(self.sp_ivl)
            VVL_dict, nm_d = Kl.Graph(self.sp_vvl, ivl_gr)

            #Проверяем обьеденяем ли ветви
            self.marker=True
            self.reper_ris=Kl.Yz_p(ivl_gr, False)
            self.reper_ris = Kl.UserReOp(KL_sp,self.reper_ris)
            for i in yr_v:
                if i =='Нет':
                    self.marker=False
                    break
            yr_v_d = False
            if self.marker:
                yr_v_d=Kl.SpSoed(yr_v)
                per_t = Kl.Yz_p(ivl_gr, yr_v_d)
            else:
                per_t = Kl.Yz_p(ivl_gr, False)
            
            k=0
            inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per),nm_ivl,self.sp_zy)
            
            for j in range(len(ivl_gr)):
                ivl_gr[j].append(0)
                ivl_gr[j].append('')
                ivl_gr[j].append(0)
                ivl_gr[j].append('')
            sp_pzy =[]
            nps = 0
            for i in range(len(self.sp_zy)):
                if self.sp_zy[i][7] == 'ПЗ':
                    nps+=1
                    for j in range(len(ivl_gr)):
                        if ivl_gr[j][0] == int(self.sp_zy[i][0]) and ivl_gr[j][1] == int(self.sp_zy[i][1]):
                            
                            s1,s2 =ivl_gr[j][6].split(self.Razd_simv,1)
                            s1=s1.strip()
                            s2=s2.strip()
                            
                            if ivl_gr[j][4] == int(self.sp_zy[i][2]):
                                ivl_gr[j][7]=1
                                ivl_gr[j][8]='ПС-'+str(nps)
                                sp_pzy.append(['ПС-'+str(nps),s1,self.sp_zy[i][3]])
                            if ivl_gr[j][5] == int(self.sp_zy[i][2]):
                                ivl_gr[j][9]=2
                                ivl_gr[j][10]='ПС-'+str(nps)
                                sp_pzy.append(['ПС-'+str(nps),s2,self.sp_zy[i][3]])

            for j in range(len(ivl_gr)):
                n1, n2, a, b, spX = Sz.DelOp(inf[j][0], inf[j][2], inf[j][3], [], self.grop)
                ivl_gr[j][4] = n1
                ivl_gr[j][5] = n2
                """
                if ivl_gr[j][7] != 0:
                    ivl_gr[j][8] += '\n'+str(n1)
                if ivl_gr[j][9] != 0:
                    ivl_gr[j][10] += '\n'+str(n2)
                """

            self.ivl_ris = ivl_gr

            self.vvl_ris = VVL_dict

            self.sp_pzy = sp_pzy

            self.nm_d = nm_d

            self.per_ris = [list(per[key]) for key in per]

            self.km_ris = km

            self.yr_op = yr_v_d
            
        except Exception as ex:
            print(ex)
     
    
    # Метод определения адреса на элемента иерархического списка
    def presseditem(self, modelindex):
        modelindex=QPersistentModelIndex(modelindex)
        if modelindex == self.adres:
            if modelindex in self.kat_dict:
                sp=self.kat_adr[self.kat_dict[modelindex]]
                if len(sp)==1:
                    if sp[0][:5]=='papk_':
                        self.kat_bd.clearSelection()
                        self.adres=None
                    else:
                        self.adres=modelindex
                else:
                    self.adres=modelindex
            else:
                self.adres=modelindex            
        else:   
            self.adres=modelindex
        
        #print(self.kat_dict[self.adres])
        

    # Функция добавления новой папки
    def NewPapk(self):
        if self.lists.currentIndex() == 0:
            Name, ok = QInputDialog().getText(self, 'Новая папка',
                    'Введите имя:',text='Новая папка')
            if ok:
                if self.adres == None:
                    Table='Koren'
                    b=[]
                else:
                    sp=self.kat_adr[self.kat_dict[QPersistentModelIndex(self.adres)]]
                    a=len(sp)
                    if a != 0:
                        if sp[a-1][:5]=='file_':
                            if a==1:
                                b=[]
                                Table='Koren'
                            else:
                                Table=sp[a-2]
                                b=copy.deepcopy(sp)
                                del b[a-1]
                        else:
                            Table=sp[a-1]
                            b=copy.deepcopy(sp)

                sql_name=BD.NewZap(Table,Name,'papk_')
                if sql_name != 'Имя занято':
                    self.kat_list.append(QStandardItem(QIcon('Pictures/papk.png'),Name))
                    a=len(self.kat_list)-1
                    if b==[]:
                        self.itemroot.appendRow(self.kat_list[a])
                    else:
                        self.kat_list[self.kat_dict_t[tuple(b)]].appendRow(self.kat_list[a])

                    b.append(sql_name)
                    self.kat_adr.append(b)
                    self.kat_dict_t[tuple(b)]=a
                    self.kat_dict[QPersistentModelIndex(self.kat_list[a].index())]=a
                    self.kat_bd.clearSelection()
                    self.adres=None


    # Функция Сохранить как
    def SaveK(self):
        self.Zapis() # читываем данные с таблиц
        Name, ok = QInputDialog().getText(self, 'Новый файл',
                'Введите имя:',text='Новый файл')
        if ok:
            if self.adres == None:
                Table='Koren'
                b=[]
            else:
                sp=self.kat_adr[self.kat_dict[QPersistentModelIndex(self.adres)]]
                a=len(sp)
                if a != 0:
                    if sp[a-1][:5]=='file_':
                        if a==1:
                            b=[]
                            Table='Koren'
                        else:
                            Table=sp[a-2]
                            b=copy.deepcopy(sp)
                            del b[a-1]
                    else:
                       Table=sp[a-1]
                       b=copy.deepcopy(sp)
                       
            sql_name=BD.NewZap(Table,Name,'file_')
            if sql_name != 'Имя занято':
                self.kat_list.append(QStandardItem(QIcon('Pictures/file.png'),Name))
                a=len(self.kat_list)-1
                if b==[]:
                    self.itemroot.appendRow(self.kat_list[a])
                else:
                    self.kat_list[self.kat_dict_t[tuple(b)]].appendRow(self.kat_list[a])

                b.append(sql_name)
                self.kat_adr.append(b)
                self.kat_dict_t[tuple(b)]=a
                self.kat_dict[QPersistentModelIndex(self.kat_list[a].index())]=a

                
                BD.ZapFile(sql_name, self.sp_ivl, self.sp_vvl, self.sp_zy, self.sp_zytr, self.sp_zvl)
                self.kat_bd.clearSelection()
                self.name_vl.setText(Name) # Заполняем стрку назания линии именем файла
                self.setWindowTitle(Name+self.Vive)
                self.CurentFile = QPersistentModelIndex(self.kat_list[a].index())
                self.adres=None
                
        
    # Функция удаления
    def DelFP(self):
        if self.adres != None:
            nam=self.kat_list[self.kat_dict[self.adres]].text()
            Message = QMessageBox(QMessageBox.Question,  'Удаление',
                    "Вы дейстивлеьно хотите удалить "+nam+"?", parent=self)
            Message.addButton('Да', QMessageBox.YesRole)
            Message.addButton('Нет', QMessageBox.NoRole)
            reply = Message.exec()       
            if reply == 0:
                sp=self.kat_adr[self.kat_dict[QPersistentModelIndex(self.adres)]]               
                c=len(sp)
                Name=sp[c-1]
                if Name[:5]=='file_':
                    self.dfile(sp,c,Name,nam,QPersistentModelIndex(self.adres))
                elif Name[:5]=='papk_':
                    self.dpapk(Name,sp,nam)


                                
    # Функция переименования
    def ReNam(self):
        if self.adres != None and self.lists.currentIndex() == 0:
            Name=self.kat_list[self.kat_dict[QPersistentModelIndex(self.adres)]].text()
            NewName, ok = QInputDialog().getText(self, 'Переименовать',
                'Введите новое имя:',text=Name)
            if ok:
                sp=self.kat_adr[self.kat_dict[QPersistentModelIndex(self.adres)]]
                if len(sp)==1:
                    Table='Koren'
                elif len(sp)>1:
                    Table=sp[len(sp)-2]
                    
                prow=BD.ReName(Table,Name ,NewName)
                if prow !='Имя занято':
                    self.kat_list[self.kat_dict[QPersistentModelIndex(self.adres)]].setText(NewName)
                    self.kat_bd.clearSelection()
                    self.adres=None

    # Метод для открития файла
    def OpenF(self,a):
        #self.kat_list[self.kat_dict[a]].setText('bbbb')
        
        if a != None and self.lists.currentIndex() == 0:
            try:
                sp=self.kat_adr[self.kat_dict[a]]
                text = self.kat_list[self.kat_dict[a]].text()
            except Exception:
                sp=self.kat_adr[self.kat_dict[QPersistentModelIndex(a)]]
                text = self.kat_list[self.kat_dict[QPersistentModelIndex(a)]].text()
            Name=sp[len(sp)-1]
            if Name[:5]=='file_':
                self.name_vl.setText(text) # Заполняем стрку назания линии именем файла
                self.setWindowTitle(text+self.Vive) # Меняем назание титульника на имя открытого файла
                
                self.Zapis() # читываем данные с таблиц
                sp_ivl0 = self.sp_ivl
                sp_vvl0 = self.sp_vvl
                sp_zy0 = self.sp_zy
                sp_zytr0 = self.sp_zytr
                sp_zvl0 = self.sp_zvl
                try:
                    
                    sp_ivl,sp_vvl,sp_zy,sp_zytr,sp_zvl=BD.ChitFile(Name)
                    self.Vvod_open_file(sp_ivl,sp_vvl,sp_zy,sp_zytr,sp_zvl)
                    self.lists.setCurrentIndex(1) # Делаем активной созданую закладку
                except Exception:
                    self.Vvod_open_file(sp_ivl0,sp_vvl0,sp_zy0,sp_zytr0,sp_zvl0)
                    ems = QErrorMessage(self)
                    ems.setWindowTitle('Возникла ошибка')
                    ems.showMessage('Файл поврежён.\n Его открытие не представляется возможным.')
                
                self.kat_bd.clearSelection()
                self.CurentFile = self.adres # Запоминаем текущий открытый файл
                self.adres=None

    # Метод для сохранить как
    def SaveP(self,Trig=None):  
        try:          
            if self.adres != None and self.lists.currentIndex() == 0:
                Trig = True
            elif Trig !=None: 
                self.adres = self.CurentFile
                Trig = False

            self.Zapis() # читываем данные с таблиц
            if self.adres != None:
                sp=self.kat_adr[self.kat_dict[QPersistentModelIndex(self.adres)]]
                Name=sp[len(sp)-1]
                if Name[:5]=='file_':
                    nam=self.kat_list[self.kat_dict[self.adres]].text()
                    Message = QMessageBox(QMessageBox.Question,  'Сохранить',
                        "Вы дейстивлеьно хотите заменить "+nam+"?", parent=self)
                    Message.addButton('Да', QMessageBox.YesRole)
                    Message.addButton('Нет', QMessageBox.NoRole)
                    reply = Message.exec()
                    if reply == 0:
                        BD.DelFile(Name)
                        BD.File(Name)
                        BD.ZapFile(Name, self.sp_ivl, self.sp_vvl, self.sp_zy, self.sp_zytr, self.sp_zvl)
                        self.kat_bd.clearSelection()
                        if Trig: self.CurentFile = self.adres
                        self.adres=None
        except Exception:
            self.CurentFile = None
                    
                    
                    

    # Функция которая будет вызыватся при удалении файла
    def dfile(self,sp,c,Name,nam,f):
        if c == 1:
            self.model.removeRow(f.row())
            # Удаляем упоминания в словарях и списках
            del self.kat_dict_t[tuple(sp)]
            del self.kat_dict[f]
            Table='Koren'
                        
        elif c > 1:
            b=copy.deepcopy(sp)
            del b[c-1]
            a=self.kat_list[self.kat_dict_t[tuple(b)]].index()
            self.model.removeRow(f.row(),parent=a)
            # Удаляем упоинания в словорях и списках
            del self.kat_dict_t[tuple(sp)]
            del self.kat_dict[f]
            Table=sp[c-2]
        BD.DelZap(Table,nam)
        self.kat_bd.clearSelection()
        self.adres=None

    #
    def dpapk(self,Name,pyt,nam):
        sp=BD.SdPapk(Name)
        for i in range(len(sp)):
            if sp[i][2][:5]=='papk_':
                b=copy.deepcopy(pyt)
                b.append(sp[i][2])
                self.dpapk(sp[i][2],b,sp[i][1])
            else: 
                b=copy.deepcopy(pyt)
                b.append(sp[i][2])
                c=len(b)
                f=QPersistentModelIndex(self.kat_list[self.kat_dict_t[tuple(b)]].index())
                self.dfile(b,c,sp[i][2],sp[i][1],f)
                
        f=QPersistentModelIndex(self.kat_list[self.kat_dict_t[tuple(pyt)]].index())
        self.dfile(pyt,len(pyt),Name,nam,f)
                
        #print(sp)

   
    # Зарузка каталога из базы данных    
    def KatZagr(self,Name,ots):
        sp=BD.SdPapk(Name)
        for i in range(len(sp)):
            if sp[i][2][:5]=='papk_':
                self.kat_list.append(QStandardItem(QIcon('Pictures/papk.png'),sp[i][1])) # Создаём вкладку списка
                a=len(self.kat_list)-1
                if ots !=None:
                    self.kat_list[ots].appendRow(self.kat_list[a]) # Формируем иерархию списка
                    # Создаём адреса соотвествующеие вкладкам списка
                    b=copy.deepcopy(self.kat_adr[ots])
                    b.append(sp[i][2])
                    self.kat_adr.append(b)

                    self.kat_dict_t[tuple(b)]=a
                    self.kat_dict[QPersistentModelIndex(self.kat_list[a].index())]=a
                    
                elif ots ==None:
                    self.itemroot.appendRow(self.kat_list[a])
                    
                    self.kat_adr.append([sp[i][2]])

                    self.kat_dict_t[tuple([sp[i][2]])]=a
                    self.kat_dict[QPersistentModelIndex(self.kat_list[a].index())]=a
                     
                self.KatZagr(sp[i][2],a) # Метод вызывает сам себя
            else:
                self.kat_list.append(QStandardItem(QIcon('Pictures/file.png'),sp[i][1])) # Создаём вкладку списка
                a=len(self.kat_list)-1
                if ots !=None:
                    self.kat_list[ots].appendRow(self.kat_list[a]) # Формируем иерархию списка
                    # Создаём адреса соотвествующеие вкладкам списка
                    b=copy.deepcopy(self.kat_adr[ots])
                    b.append(sp[i][2])
                    self.kat_adr.append(b)

                    self.kat_dict_t[tuple(b)]=a
                    self.kat_dict[QPersistentModelIndex(self.kat_list[a].index())]=a
       
                elif ots ==None:
                    self.itemroot.appendRow(self.kat_list[a])
                    
                    self.kat_adr.append([sp[i][2]])
                    self.kat_dict_t[tuple([sp[i][2]])]=a
                    self.kat_dict[QPersistentModelIndex(self.kat_list[a].index())]=a
        #print(self.kat_dict_t)

        

# Метод для выставления флажков и базы данных
    def CheckBD(self):
        sp = BD.Chit_nasrt()
        #print(len(sp))

        self.VkPdop = bool(sp[0][1])
        self.ch_miss.setCheckState(3 if self.VkPdop else 0)
        self.grop = bool(sp[1][1])
        self.GrOp.setCheckState(3 if self.grop else 0)
        self.zpps = bool(sp[2][1])
        self.ZpPS.setCheckState(3 if self.zpps else 0)
        self.grafsx = bool(sp[3][1])
        self.Ch_sxzgr.setCheckState(3 if self.grafsx else 0)
        self.sr_vel = sp[4][2]
        self.Dop_V.setText(str(self.sr_vel)) # Предельно допустимое напряжение
        self.k_yvgr = sp[5][2]
        self.Koef_gr.setText(str(self.k_yvgr)) # Коэффициент увеличения значений графика
        self.dz = bool(sp[6][1])
        self.DZ.setCheckState(3 if self.dz else 0)

        self.FACh = bool(sp[7][1])
        self.Ch_A.setCheckState(3 if self.FACh else 0)
        self.FAName = sp[7][3]
        self.Ln_A.setText(self.FAName)

        self.FBCh = bool(sp[8][1])
        self.Ch_B.setCheckState(3 if self.FBCh else 0)
        self.FBName = sp[8][3]
        self.Ln_B.setText(self.FBName)

        self.FCCh = bool(sp[9][1])
        self.Ch_C.setCheckState(3 if self.FCCh else 0)
        self.FCName = sp[9][3]
        self.Ln_C.setText(self.FCName)

        self.FTRCh = bool(sp[10][1])
        self.Ch_Tr.setCheckState(3 if self.FTRCh else 0)
        self.FTRName = sp[10][3]
        self.Ln_Tr.setText(self.FTRName)

        self.title_f_s.setValue(sp[11][1])
        self.f_s.setValue(sp[12][1])
        self.FoolWind=bool(sp[13][1])

        self.pz = sp[14][2]
        self.grunt.setText(str(self.pz)) #Сопротивление грунта

        self.kesh_1 = bool(sp[15][1])
        self.Kesh_VL.setCheckState(3 if self.kesh_1 else 0)

        self.kesh_2 = bool(sp[16][1])
        self.Kesh_ZY.setCheckState(3 if self.kesh_2 else 0)

        self.kesh_3 = bool(sp[17][1])
        self.Kesh_P.setCheckState(3 if self.kesh_3 else 0)

        self.kesh_4 = bool(sp[18][1])
        self.Kesh_R.setCheckState(3 if self.kesh_4 else 0)

        self.ValPZ1.setValue(sp[19][1])
        self.ValPZ2.setValue(sp[20][1])

        self.one_sch4 = bool(sp[21][1])
        self.Ch_one_sch4.setCheckState(3 if self.one_sch4 else 0)

        # Данные для выбора СЗ
        self.ValSZ.setValue(sp[22][1])
        self.ValMaxSZ.setValue(sp[23][2])
        self.ValMinSZ.setValue(sp[24][2])
        self.ValEpsSZ.setValue(sp[25][2])
        self.ValTooSZ.setValue(sp[26][1])

        self.too_sz = bool(sp[27][1])
        self.Ch_too_sz.setCheckState(3 if self.too_sz else 0)

        # Генерация набора инструкций
        self.one_docx = bool(sp[28][1])
        self.Ch_one_file.setCheckState(3 if self.one_docx else 0)
        self.dbl_vl = bool(sp[29][1])
        self.Ch_dbl_vl.setCheckState(3 if self.dbl_vl else 0)
        self.use_templ = bool(sp[30][1])
        self.Ch_use_templ.setCheckState(3 if self.use_templ else 0)
            
               
# Выбираем метод для различных схем        
    def Vib_sx(self):
        if self.N_sxem.currentText()!='№ 3':
            self.OtW()
        else:
            self.SxT()
# Метод ввода значения для настройки допустимого напляжения          
    def DopVolt(self):
        s=Kl.Tochka(self.Dop_V.text())
        try:
            a=float(s)
        except:
            self.Dop_V.setText(str(self.sr_vel))
        else:
            self.sr_vel=a
# Метод ввода сопротивления грунта
    def Grunt(self):
        s=Kl.Tochka(self.grunt.text())
        try:
            a=float(s)
        except:
            self.grunt.setText(str(self.pz))
        else:
            self.pz=a
# Метод ввода коэфициента маштабирования
    def KefGr(self):
        s=Kl.Tochka(self.Koef_gr.text())
        try:
            a=float(s)
        except:
            self.Koef_gr.setText(str(self.k_yvgr))
        else:
            self.k_yvgr=a

    
    def LnA (self):
        self.FAName=Kl.Tochka(self.Ln_A.text())
    def LnB (self):
        self.FBName=Kl.Tochka(self.Ln_B.text())
    def LnC (self):
        self.FCName=Kl.Tochka(self.Ln_C.text())
    def LnTr (self):
        self.FTRName=Kl.Tochka(self.Ln_Tr.text())
                
        
            
# Функция для расчёта и записи Pдоп
    def onActivated(self,i,text):     
        1 
        """ ind=Kl.Idop(self.MR_d[text])
        if ind !=-1 and ind !=None and self.VkPdop:
            for j in range(-i,1):
                if self.table_VVL.item(-j,0).text() != 'Конфигурация опоры':
                    if self.table_VVL.item(-j,3).text() !='':
                        try:
                            u=float(Kl.Tochka(self.table_VVL.item(-j,3).text()))
                        except Exception:
                            break
                    else:
                        break
                    if self.table_VVL.item(-j,4).text() !='':
                        try:
                            a=float(Kl.Tochka(self.table_VVL.item(-j,4).text()))
                        except Exception:
                            a=10**6
                    else:
                        a=10**6
                    b=3**(0.5)*u*ind*10**(-3)
                    if a > b:
                        self.table_VVL.setItem(-j,4, QTableWidgetItem(str(round(b,2))))
                        self.table_VVL.item(-j,4).setBackground(QColor(162,205,90))
                    break """

# функция для построения изображения выбранной опоры и её фазировки

    def OpenFig(self,k,i,j):
        if j==0:
            if k==1:
                text=self.table_IVL.item(i,0).text()
            elif k==2:
                text=self.table_VVL.item(i,0).text()
            if j==0 and text == 'Конфигурация опоры':
                if k==1:
                    Fz=self.combo_1[i][0].currentText()
                    To=self.combo_1[i][1].currentText()
                elif k==2:
                    Fz=self.combo_2[i][0].currentText()
                    To=self.combo_2[i][1].currentText()
                if To == "Не выбран":
                    return
                w=rnn.VF(self.d_to[To],self.XHkat[Fz],'OpFig',self.Katalog,self.Kat)
                OpRis.RisOpor(w,i,k,To)
                
                self.Mod=KOp()
                self.Mod.KonfOp(k,i,To)
                self.Mod.show()
        elif j==8 and k==2:
            if self.table_VVL.item(i,0).text() != 'Конфигурация опоры' and\
                self.table_VVL.item(i,8).text() == 'Расч. Pdop':
                try:
                    u=float(Kl.Tochka(self.table_VVL.item(i,3).text()))
                except Exception:
                    return
                a = []
                for n in range(i+1,self.vvl_row):
                    if self.table_VVL.item(n,0).text() == 'Конфигурация опоры':
                        a.append(Kl.Idop(self.MR_d[self.combo_2[n][2].currentText()]))
                    else:
                        break
                b = []
                for m in range(len(a)):
                    b.append(3**(0.5)*u*a[m]*10**(-3))
                if self.table_VVL.item(i,0).text() != 'Конфигурация опоры' and\
                    self.table_VVL.item(i,8).text() == 'Расч. Pdop' and len(b) != 0:
                    self.table_VVL.setItem(i,4, QTableWidgetItem(str(round(min(b),2))))
                    self.table_VVL.item(i,4).setBackground(QColor(162,205,90))
                #onActivated

                
        
# Создаём метод для выпадающих списков
    def Spiski(self,i,j):
        
        # заполним пустой список выпадающими списками
        combo = ['','','','','','','','','','','']
        #Создаём выпадающий список для "Фазировка"
        combo[0] = UserComboBox(self.cntr_pr)
        for item in self.FF:
            combo[0].addItem(item)
        #Создаём выпадающий список для "Тип опор"
        combo[1] = UserComboBox(self.cntr_pr)
        combo[1].addItem("Не выбран")
        for item in self.TO:
            combo[1].addItem(item)
        #Создаём выпадающий список для "Марка провода"
        combo[2] = UserComboBox(self.cntr_pr)
        for item in self.MR:
            combo[2].addItem(item)
        #Создаём выпадающий список для "Марка тросса №1"
        combo[3] = UserComboBox(self.cntr_pr)
        for item in self.MR:
            combo[3].addItem(item)
        #Создаём выпадающий список для "Марка тросса №2"
        combo[4] = UserComboBox(self.cntr_pr)
        for item in self.MR:
            combo[4].addItem(item)
        #Всоздаём выпадающий список для "Тип заземления"
        combo[5] = UserComboBox(self.cntr_pr)
        for item in ['3ф с тр.', '3ф без тр.', '2ф с тр', '2ф без тр', '1ф с тр', '1ф без тр', 'Разземлен']:
            combo[5].addItem(item)
        #Всоздаём выпадающий список для "Фаза №1"
        combo[6] = UserComboBox(self.cntr_pr)
        for item in 'ABC':
            combo[6].addItem(item)
        #Всоздаём выпадающий список для "Фаза №2"
        combo[7] = UserComboBox(self.cntr_pr)
        for item in 'ABC':
            combo[7].addItem(item)
        # Выпадающий список прорисовки схем
        combo[8] = UserComboBox(self.cntr_pr)
        for item in ['В','С','З','Ю']:#['В','С-В','С','С-З','З','Ю-З','Ю','Ю-В']
            combo[8].addItem(item)
        # Выпадающий список типов заземления
        combo[9] = ExtendedComboBox(self.cntr_pr)#QComboBox()
        for item in ['ПЗ','БЗ','СЗ','ЛЗ']:
            combo[9].addItem(item)
        # Выпадающий список уровня линии
        combo[10] = UserComboBox(self.cntr_pr)
        d=[]
        d.append('Нет')
        for n in range(1,21):
            d.append('Уровень '+str(n))
        for item in d:
            combo[10].addItem(item)

        # Присваеваем их нужной таблице
        
        if j == 1:
            self.combo_1[i][0]=combo[0]
            self.combo_1[i][1]=combo[1]
            self.combo_1[i][2]=combo[2]
            self.combo_1[i][3]=combo[3]
            self.combo_1[i][4]=combo[4]
            self.combo_1[i][5]=combo[8]
            self.combo_1[i][6]=combo[10]
        elif j == 2:
            self.combo_2[i][0]=combo[0]
            self.combo_2[i][1]=combo[1]
            self.combo_2[i][2]=combo[2]
            self.combo_2[i][3]=combo[3]
            self.combo_2[i][4]=combo[4]
        elif j == 3:
            self.combo_3[i][0]=combo[5]
            self.combo_3[i][1]=combo[6]
            self.combo_3[i][2]=combo[7]
            self.combo_3[i][3]=combo[9]

        elif j == 5:
            self.combo_5[i][0]=combo[0]
            self.combo_5[i][1]=combo[1]
            self.combo_5[i][2]=combo[2]
            self.combo_5[i][3]=combo[3]

    # Метод для полной очистки вкладки
    def Clearing(self,okn):
        if okn==1:
            self.Del_obl(0,self.ivl_row,1)
            self.ivl_row=0
        elif okn==2:
            for i in range(0,self.vvl_row):
                self.combo_2[i][2].activated[str].disconnect()
            self.Del_obl(0,self.vvl_row,2)
            self.vvl_row=0
            self.num_vvl =0
        elif okn==3:
            self.Del_obl(0,self.zy_row,3)
            self.zy_row=0
        elif okn==4:
            self.Del_obl(0,self.zytr_row,4)
            self.zytr_row=0
        elif okn==5:
            self.Del_obl(0,self.zvl_row,5)
            self.zvl_row=0


    # Проверить выделеные элементы строки
    def Videl_str(self, sp):
        if len(sp) != 0:
            ind=sp[0].row()
            for i in range(1,len(sp)):
                if ind != sp[i].row():
                    ind=-1
                    break
        else:
            ind=-1
        return ind
    
    def Videl_str_s(self, sp):
        l =set()
        for i in range(len(sp)):
            l.add(sp[i].row())
        return list(l)
    
    def ivl_coppy(self,ind,trig=False):
        a=[]
        b=[]
        for j in range(1,4):
            if trig: b.append(self.table_IVL.item(ind,j).text() if trig[j-1] else self.coppy_ych1[j-1])
            else:b.append(self.table_IVL.item(ind,j).text()) 
        for j in range(4,9):
            if trig:a.append(self.combo_1[ind][j-4].currentText() if trig[j-1] else self.coppy_param[j-4])
            else: a.append(self.combo_1[ind][j-4].currentText())
        return a,b
    
    def vvl_coppy(self,ind,trig=False):
        a=[]
        b=[]
        for j in range(1,4):
            if trig: b.append(self.table_VVL.item(ind,j).text() if trig[j-1] else self.coppy_ych2[j-1])
            else:b.append(self.table_VVL.item(ind,j).text())
        for j in range(4,9):
            if trig:a.append(self.combo_2[ind][j-4].currentText() if trig[j-1] else self.coppy_param[j-4])
            else: a.append(self.combo_2[ind][j-4].currentText())    
        return a,b

    def set_coppy_ivl(self,ind,a,b,trig=False):
        for j in range(1,4):
            if trig:
                if trig[j-1]:self.table_IVL.item(ind,j).setText(b[j-1])
        for j in range(4,9):
            if trig:
                if trig[j-1]: self.combo_1[ind][j-4].setCurrentText(a[j-4])
            else: self.combo_1[ind][j-4].setCurrentText(a[j-4])

    def set_coppy_vvl(self,ind,a,b,trig=False):
        for j in range(1,4):
            if trig:
                if trig[j-1]:self.table_VVL.item(ind,j).setText(b[j-1])
        for j in range(4,9):
            if trig:
                if trig[j-1]: self.combo_2[ind][j-4].setCurrentText(a[j-4])
            else: self.combo_2[ind][j-4].setCurrentText(a[j-4])

    def previous_coppy_ivl(self,ind):
        for i in range(ind-1,-1,-1):
            if self.table_IVL.item(i,0).text() == 'Конфигурация опоры':
                a,b = self.ivl_coppy(i)
                self.set_coppy_ivl(ind,a,b)
                break
    def previous_coppy_vvl(self,ind):
        for i in range(ind-1,-1,-1):
            if self.table_VVL.item(i,0).text() == 'Конфигурация опоры':
                a,b = self.vvl_coppy(i)
                self.set_coppy_vvl(ind,a,b)
                break

    def CheckStatus(self,page):
        p = [
            [self.ch_1_2,self.ch_1_3,self.ch_1_4,self.ch_1_5,self.ch_1_6,self.ch_1_7,self.ch_1_8,self.ch_1_9],\
            [self.ch_2_2,self.ch_2_3,self.ch_2_4,self.ch_2_5,self.ch_2_6,self.ch_2_7,self.ch_2_8,self.ch_2_9],\
            [self.ch_3_1,self.ch_3_2,self.ch_3_3,self.ch_3_4,self.ch_3_5,self.ch_3_6,self.ch_3_7,self.ch_3_8],\
            [self.ch_4_1,self.ch_4_2,self.ch_4_3,self.ch_4_4,self.ch_4_5],\
            [self.ch_5_1,self.ch_5_2,self.ch_5_3,self.ch_5_4,self.ch_5_5,self.ch_5_6,self.ch_5_7,self.ch_5_8,self.ch_5_9]]
        tr =[]
        for i in p[page-1]:
            tr.append(True if i.checkState()==Qt.Checked else False)
        return tr


    def CoppyRowCels(self):
        i = self.lists.currentIndex()
        trig = self.CheckStatus(i)
        if i == 1:
            ind=self.Videl_str(self.table_IVL.selectedIndexes())
            if ind != -1:
                if self.table_IVL.item(ind,0).text() == 'Конфигурация опоры':
                    self.coppy_param,self.coppy_ych1 = self.ivl_coppy(ind,trig=trig)
                            
        elif i==2:
            ind=self.Videl_str(self.table_VVL.selectedIndexes())
            if ind != -1:
                if self.table_VVL.item(ind,0).text() == 'Конфигурация опоры':
                    self.coppy_param, self.coppy_ych2 = self.vvl_coppy(ind)

        elif i==3:
            ind=self.Videl_str(self.table_Zy.selectedIndexes())
            if ind != -1:
                for j in range(4):
                    if trig[j]: self.coppy_zy[j]=self.table_Zy.item(ind,j).text()
                for j in range(4,8):
                    if trig[j]: self.coppy_zy[j]=self.combo_3[ind][j-4].currentText()

        elif i==4:
            ind=self.Videl_str(self.table_Zytr.selectedIndexes())
            if ind != -1:
                for j in range(5):
                    if trig[j]: self.coppy_zy_tr[j]=self.table_Zytr.item(ind,j).text()

        elif i==5:
            ind=self.Videl_str(self.table_Zvl.selectedIndexes())
            if ind != -1:
                for j in range(5):
                    if trig[j]: self.coppy_zvl[j]=self.table_Zvl.item(ind,j).text()
                for j in range(5,9):
                    if trig[j]: self.coppy_zvl[j]=self.combo_5[ind][j-5].currentText()


    
    def SetCoppyRow(self):
        i = self.lists.currentIndex()
        trig = self.CheckStatus(i)
        if i == 1:
            l=self.Videl_str_s(self.table_IVL.selectedIndexes())
            for ind in l:
                if ind != -1:
                    if self.table_IVL.item(ind,0).text() == 'Конфигурация опоры':
                        self.set_coppy_ivl(ind,self.coppy_param,self.coppy_ych1,trig=trig)
        
        elif i==2:
            l=self.Videl_str_s(self.table_VVL.selectedIndexes())
            for ind in l:
                if ind != -1:
                    if self.table_VVL.item(ind,0).text() == 'Конфигурация опоры':
                        self.set_coppy_vvl(ind,self.coppy_param,self.coppy_ych2,trig=trig)
                    

        elif i==3:
            l=self.Videl_str_s(self.table_Zy.selectedIndexes())
            for ind in l:
                if ind != -1:
                    for j in range(4):
                        if trig[j]: self.table_Zy.item(ind,j).setText(self.coppy_zy[j])
                    for j in range(4,8):
                        if trig[j]: self.combo_3[ind][j-4].setCurrentText(self.coppy_zy[j])

        elif i==4:
            l=self.Videl_str_s(self.table_Zytr.selectedIndexes())
            for ind in l:
                if ind != -1:
                    for j in range(5):
                        if trig[j]: self.table_Zytr.item(ind,j).setText(self.coppy_zy_tr[j])

        elif i==5:
            l=self.Videl_str_s(self.table_Zvl.selectedIndexes())
            for ind in l:
                if ind != -1:
                    for j in range(5):
                        if trig[j]: self.table_Zvl.item(ind,j).setText(self.coppy_zvl[j])
                    for j in range(5,9):
                        if trig[j]: self.combo_5[ind][j-5].setCurrentText(self.coppy_zvl[j])


    # Функция удаления заданной области таблицы
    def Del_obl(self,n,k,i):
        for j in range(-k,-n+1):
            if i==1:
                self.table_IVL.setRowCount(-j)
            elif i==2:
                self.table_VVL.setRowCount(-j)
                if -j in self.vvl_pos and j!=-n:
                    self.num_vvl -= 1
                    del self.vvl_pos[-j]
            elif i==3:
                self.table_Zy.setRowCount(-j)
            elif i==4:
                self.table_Zytr.setRowCount(-j)
            elif i==5:
                self.table_Zvl.setRowCount(-j)

    # Заполнение конца таблицы
    def Zap_k(self,n,k,i,y):
        for j in range(n,k):
            self.Spiski(abs(j),i)
            if i==1:
                self.Vvod_ivl(self.sp_ivl[j-y],j)
            elif i==2:
                self.Vvod_vvl(self.sp_vvl[j-y],j)
            elif i==3:
                self.Vvod_zy(self.sp_zy[j-y],j)
            elif i==4:
                self.Vvod_zytr(self.sp_zytr[j-y],j)
            elif i==5:
                self.Vvod_zvl(self.sp_zvl[j-y],j)
            
        
    def Pyst_1_v(self,k):
        self.Spiski(k,1)
        self.table_IVL.setRowCount(k+1)
        self.table_IVL.setItem(k,0, QTableWidgetItem(''))
        self.table_IVL.setCellWidget(k,1, self.combo_1[k][5])
        self.table_IVL.setCellWidget(k,4, self.combo_1[k][6])
        self.table_IVL.item(k,0).setBackground(QColor(162,205,90))
        for i in range(2,4):
            self.table_IVL.setItem(k,i, QTableWidgetItem(''))
            self.table_IVL.item(k,i).setBackground(QColor(162,205,90))
        for i in range(5,9):
            self.table_IVL.setItem(k,i, QTableWidgetItem('*****'))
            self.table_IVL.item(k,i).setBackground(QColor(162,205,90))
    
    # Медоды для создания и удаления строк таблицы ИВЛ    
    def new_1_v(self):
        ind=self.Videl_str(self.table_IVL.selectedIndexes()) # Получаем индекс выделенную строку
        if ind != -1 and ind+1 != self.ivl_row: 
            self.Zapis_ivl()
            self.ivl_row += 1
            self.Del_obl(ind+1,self.ivl_row,1)
            self.Pyst_1_v(ind+1)
            self.Zap_k(ind+2,self.ivl_row,1,1)
        else:
            self.ivl_row += 1
            self.Pyst_1_v(self.ivl_row-1)

    def Pyst_1_y(self,k):
        self.Spiski(k,1)
        self.table_IVL.setRowCount(k+1)
        self.table_IVL.setItem(k,0, QTableWidgetItem('Конфигурация опоры'))
        for i in range(1,4):
            self.table_IVL.setItem(k,i, QTableWidgetItem(''))
        # Помещаем выпадающие списки в ячейки таблицы
        for i in range(4,9):
            self.table_IVL.setCellWidget(k,i, self.combo_1[k][i-4])
            
    def new_1_y(self): 
        ind=self.Videl_str(self.table_IVL.selectedIndexes()) # Получаем индекс выделенную строку
        if ind != -1 and ind+1 != self.ivl_row:
            self.Zapis_ivl()
            self.ivl_row += 1
            self.Del_obl(ind+1,self.ivl_row,1)
            self.Pyst_1_y(ind+1)
            self.Zap_k(ind+2,self.ivl_row,1,1)
            self.previous_coppy_ivl(ind+1)
        else:
            self.ivl_row += 1
            self.Pyst_1_y(self.ivl_row-1)
            self.previous_coppy_ivl(self.ivl_row-1)
        
    def del_1(self):
        if self.ivl_row ==0: 1
        else:
            ind=self.Videl_str(self.table_IVL.selectedIndexes()) # Получаем индекс выделенную строку
            if ind != -1 and ind+1 != self.ivl_row:
                self.Zapis_ivl()
                self.Del_obl(ind,self.ivl_row,1)
                self.ivl_row -= 1
                self.Zap_k(ind,self.ivl_row,1,-1)
            else:    
                self.Spiski(self.ivl_row-1,1)
                self.table_IVL.setRowCount(self.ivl_row-1)
                self.ivl_row -=1

    # Медоды для создания и удаления строк таблицы ВВЛ
    def Pyst_2_v(self,k):
        self.Spiski(k,2)
        self.table_VVL.setRowCount(k+1)
        for i in range(5):
            self.table_VVL.setItem(k,i, QTableWidgetItem(''))
            self.table_VVL.item(k,i).setBackground(QColor(162,205,90))
        self.table_VVL.setItem(k,5, QTableWidgetItem('0'))
        self.table_VVL.setItem(k,6, QTableWidgetItem('1'))
        self.table_VVL.setItem(k,7, QTableWidgetItem('L'+str(self.num_vvl)))
        self.table_VVL.item(k,5).setBackground(QColor(162,205,90))
        self.table_VVL.item(k,6).setBackground(QColor(162,205,90))
        self.table_VVL.item(k,7).setBackground(QColor(162,205,90))
        self.table_VVL.setItem(k,8, QTableWidgetItem('Расч. Pdop'))
        self.table_VVL.item(k,8).setBackground(QColor(162,205,90))
        self.combo_2[k][2].activated[str].connect(partial(self.onActivated,k)) # Отправляем текст выпадающего списка марок проводников и его расположение

            
    def new_2_v(self):
        ind=self.Videl_str(self.table_VVL.selectedIndexes()) # Получаем индекс выделенную строку
        if ind != -1 and ind+1 != self.vvl_row: 
            self.Zapis_vvl()
            for i in range(ind+1,self.vvl_row):
                self.combo_2[i][2].activated[str].disconnect()
            self.vvl_row += 1
            self.Del_obl(ind+1,self.vvl_row,2)
            self.num_vvl += 1
            self.vvl_pos[ind+2] = self.num_vvl
            self.Pyst_2_v(ind+1)
            self.Zap_k(ind+2,self.vvl_row,2,1)

            for i in range(ind+1,self.vvl_row):
                self.combo_2[i][2].activated[str].connect(partial(self.onActivated,i))

        else:
            self.vvl_row += 1
            self.num_vvl += 1
            self.vvl_pos[self.vvl_row] = self.num_vvl
            self.Pyst_2_v(self.vvl_row-1)


    def Pyst_2_y(self,k):
        self.Spiski(k,2)
        self.table_VVL.setRowCount(k+1)
        self.table_VVL.setItem(k,0, QTableWidgetItem('Конфигурация опоры'))
        for i in range(1,4):
            self.table_VVL.setItem(k,i, QTableWidgetItem(''))
        # Помещаем выпадающие списки в ячейки таблицы
        for i in range(4,9):
            self.table_VVL.setCellWidget(k,i, self.combo_2[k][i-4])
        self.combo_2[k][2].activated[str].connect(partial(self.onActivated,k)) # Отправляем текст выпадающего текста и его расположение

        
    def new_2_y(self):
        ind=self.Videl_str(self.table_VVL.selectedIndexes()) # Получаем индекс выделенную строку
        if ind != -1 and ind+1 != self.vvl_row: 
            self.Zapis_vvl()
            for i in range(ind+1,self.vvl_row):
                self.combo_2[i][2].activated[str].disconnect()
            self.vvl_row += 1
            self.Del_obl(ind+1,self.vvl_row,2)
            self.Pyst_2_y(ind+1)
            self.Zap_k(ind+2,self.vvl_row,2,1)
            self.previous_coppy_vvl(ind+1)
            for i in range(ind+2,self.vvl_row):
                self.combo_2[i][2].activated[str].connect(partial(self.onActivated,i))

        else:
            self.vvl_row += 1
            self.Pyst_2_y(self.vvl_row-1)
            self.previous_coppy_vvl(self.vvl_row-1)
        
        
    def del_2(self):
        if self.vvl_row ==0: 1
        else:

            ind=self.Videl_str(self.table_VVL.selectedIndexes()) # Получаем индекс выделенную строку
            if ind != -1 and ind+1 != self.vvl_row:
                self.Zapis_vvl()
                for i in range(ind,self.vvl_row):
                    self.combo_2[i][2].activated[str].disconnect()
                self.Del_obl(ind,self.vvl_row,2)
                self.vvl_row -= 1
                self.Zap_k(ind,self.vvl_row,2,-1)
                for i in range(ind,self.vvl_row):
                    self.combo_2[i][2].activated[str].connect(partial(self.onActivated,i))
            else:
                self.combo_2[self.vvl_row-1][2].activated[str].disconnect()
                self.Spiski(self.vvl_row-1,2)
                self.table_VVL.setRowCount(self.vvl_row-1)
                if self.vvl_row in self.vvl_pos:
                    del self.vvl_pos[self.vvl_row]
                    self.num_vvl-=1
                self.vvl_row -=1

    # Медоды для создания и удаления строк таблицы "Заземление в точках"
    def Pyst_3(self,k):
        self.Spiski(k,3)
        self.table_Zy.setRowCount(k+1)
        for i in range(4):
            self.table_Zy.setItem(k,i, QTableWidgetItem(''))
        # Помещаем выпадающие списки в ячейки таблицы
        for i in range(4,8):
            self.table_Zy.setCellWidget(k,i, self.combo_3[k][i-4])
    
    def new_3(self): 
        ind=self.Videl_str(self.table_Zy.selectedIndexes()) # Получаем индекс выделенную строку
        if ind != -1 and ind+1 != self.zy_row: 
            self.Zapis_zy()
            self.zy_row += 1
            self.Del_obl(ind+1,self.zy_row,3)
            self.Pyst_3(ind+1)
            self.Zap_k(ind+2,self.zy_row,3,1)
        else:
            self.zy_row += 1
            self.Pyst_3(self.zy_row-1)
        
        
    def del_3(self):
        if self.zy_row ==0: 1
        else:
            ind=self.Videl_str(self.table_Zy.selectedIndexes()) # Получаем индекс выделенную строку
            if ind != -1 and ind+1 != self.zy_row:
                self.Zapis_zy()
                self.Del_obl(ind,self.zy_row,3)
                self.zy_row -= 1
                self.Zap_k(ind,self.zy_row,3,-1)
            else:    
                self.Spiski(self.zy_row-1,3)
                self.table_Zy.setRowCount(self.zy_row-1)
                self.zy_row -=1

    # Медоды для создания и удаления строк таблицы "Заземление троссов опор ИВЛ"
    def Pyst_4(self,k):
        self.table_Zytr.setRowCount(k+1)
        for i in range(4):
            self.table_Zytr.setItem(k,i, QTableWidgetItem(''))
        self.table_Zytr.setItem(k,4, QTableWidgetItem('30'))
    
    def new_4(self):
        ind=self.Videl_str(self.table_Zytr.selectedIndexes()) # Получаем индекс выделенную строку
        if ind != -1 and ind+1 != self.zytr_row: 
            self.Zapis_zytr()
            self.zytr_row += 1
            self.Del_obl(ind+1,self.zytr_row,4)
            self.Pyst_4(ind+1)
            self.Zap_k(ind+2,self.zytr_row,4,1)
        else:
            self.zytr_row += 1
            self.Pyst_4(self.zytr_row-1)
        
        
        
    def del_4(self):
        if self.zytr_row ==0: 1
        else:
            ind=self.Videl_str(self.table_Zytr.selectedIndexes()) # Получаем индекс выделенную строку
            if ind != -1 and ind+1 != self.zytr_row:
                self.Zapis_zytr()
                self.Del_obl(ind,self.zytr_row,4)
                self.zytr_row -= 1
                self.Zap_k(ind,self.zytr_row,4,-1)
            else:    
                self.Spiski(self.zytr_row-1,4)
                self.table_Zytr.setRowCount(self.zytr_row-1)
                self.zytr_row -=1

    # Медоды для создания и удаления строк таблицы "Заземление троссов опор ВВЛ"
    def Pyst_5(self,k):
        self.Spiski(k,5)
        self.table_Zvl.setRowCount(k+1)
        for i in range(1):
            self.table_Zvl.setItem(k,i, QTableWidgetItem('30'))
        for i in range(1,5):
            self.table_Zvl.setItem(k,i, QTableWidgetItem('0'))
        # Помещаем выпадающие списки в ячейки таблицы
        for i in range(5,9):
            self.table_Zvl.setCellWidget(k,i, self.combo_5[k][i-5])
        
    def new_5(self):
        ind=self.Videl_str(self.table_Zvl.selectedIndexes()) # Получаем индекс выделенную строку
        if ind != -1 and ind+1 != self.zvl_row: 
            self.Zapis_zvl()
            self.zvl_row += 1
            self.Del_obl(ind+1,self.zvl_row,5)
            self.Pyst_5(ind+1)
            self.Zap_k(ind+2,self.zvl_row,5,1)
        else:
            self.zvl_row += 1
            self.Pyst_5(self.zvl_row-1)
        
        
    def del_5(self):
        if self.zvl_row ==0: 1
        else:
            ind=self.Videl_str(self.table_Zvl.selectedIndexes()) # Получаем индекс выделенную строку
            if ind != -1 and ind+1 != self.zvl_row:
                self.Zapis_zvl()
                self.Del_obl(ind,self.zvl_row,5)
                self.zvl_row -= 1
                self.Zap_k(ind,self.zvl_row,5,-1)
            else:    
                self.Spiski(self.zvl_row-1,5)
                self.table_Zvl.setRowCount(self.zvl_row-1)
                self.zvl_row -=1

    def Zapis(self):
        self.Zapis_ivl()
        self.Zapis_vvl()
        self.Zapis_zy()
        self.Zapis_zytr()
        self.Zapis_zvl()
        
    def Zapis_ivl(self):
        # Считываем таблицу ИВЛ
        self.sp_ivl=[]
        for i in range(self.ivl_row):
            a=[]
            if self.table_IVL.item(i,0).text() != 'Конфигурация опоры':
                a.append(self.table_IVL.item(i,0).text())
                a.append(self.combo_1[i][5].currentText())
                for j in range(2,4):
                    a.append(self.table_IVL.item(i,j).text())
                a.append(self.combo_1[i][6].currentText())
                for j in range(5,9):
                    a.append(self.table_IVL.item(i,j).text())
            else:
                for j in range(4):
                    a.append(self.table_IVL.item(i,j).text())
                for j in range(4,9):
                    a.append(self.combo_1[i][j-4].currentText())
            self.sp_ivl.append(a)
    def Zapis_vvl(self):
        # Считываем таблицу ВВЛ
        self.sp_vvl=[]
        for i in range(self.vvl_row):
            a=[]
            if self.table_VVL.item(i,0).text() != 'Конфигурация опоры':
                for j in range(9):
                    a.append(self.table_VVL.item(i,j).text())
            else:
                for j in range(4):
                    a.append(self.table_VVL.item(i,j).text())
                for j in range(4,9):
                    a.append(self.combo_2[i][j-4].currentText())
            self.sp_vvl.append(a)
    def Zapis_zy(self):
        # Считываем таблицу Заземление в точке
        self.sp_zy=[]
        for i in range(self.zy_row):
            a=[]
            for j in range(4):
                a.append(self.table_Zy.item(i,j).text())
            for j in range(4,8):
                a.append(self.combo_3[i][j-4].currentText())
            self.sp_zy.append(a)
    def Zapis_zytr(self):
        # Считываем таблицу Заземление в точке
        self.sp_zytr=[]
        for i in range(self.zytr_row):
            a=[]
            for j in range(5):
                a.append(self.table_Zytr.item(i,j).text())
                
            self.sp_zytr.append(a)
    def Zapis_zvl(self):
        # Считываем таблицу Заземление в точке
        self.sp_zvl=[]
        for i in range(self.zvl_row):
            a=[]
            for j in range(5):
                a.append(self.table_Zvl.item(i,j).text())
            for j in range(5,9):
                a.append(self.combo_5[i][j-5].currentText())
            self.sp_zvl.append(a)

    # Блок для сахранения файла
    def Save_tables(self):
        self.Zapis() # читываем данные с таблиц
        
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.pyt+self.name_vl.text()+" ИД",'*.txt')[0] # Обрати внимание на последний элемент
            self.pyt = Kl.adres(fname)
            f = open(fname, 'w')
            f.write(str(self.sp_ivl)+'\n'+str(self.sp_vvl)+'\n'+str(self.sp_zy)\
                    +'\n'+str(self.sp_zytr)+'\n'+str(self.sp_zvl))
            f.close()
        except Exception:
            1


    def ErrorsShow(self,data,signal):
        print(data)
        if data != None:
            if data[0]==1: table,ind = self.table_IVL, 1
            elif data[0]==2: table,ind = self.table_VVL, 2
            elif data[0]==3: table,ind = self.table_Zy, 3
            elif data[0]==4: table,ind = self.table_Zytr, 4
            elif data[0]==5: table,ind = self.table_Zvl, 5
            
            s=""
            if data[1] != -1 and data[2]==9:
                table.selectRow(data[1])
                s=". Строка: "+str(data[1]+1)+". "
            elif data[1] != -1:
                table.setRangeSelected(QTableWidgetSelectionRange(data[1], data[2][0], data[1], data[2][1]), True)
                s=". Строка: "+str(data[1]+1)+". "+"Столбцы: "+str(data[2][0]+1)+"-"+str(data[2][1]+1)+"."
            self.lists.setCurrentIndex(ind)
            signal.mysignal.emit(data[3]+s)
                
        
        
    # Блок для выполнения расчета
    def Raschet(self):
        self.FiA=[]
        self.FiB=[]
        self.FiC=[]
        self.FiT1=[]
        self.p=[]
        self.km=[]
        self.per=[]
        self.ivl_gr=[]
        self.Zapis()
        self.correct = [False,False,False,False,False,False]
        data = FE.ScanId(self.sp_ivl,self.sp_vvl,self.sp_zy,self.sp_zytr,self.sp_zvl,self.Razd_simv, self.VkPdop)
        
        
        self.Ind = QProgressDialog('Производится расчёт','Отмена', 0, 0, self)
        self.Ind.setWindowTitle('Расчет ')
        self.Ind.setMinimumDuration(0)
        self.Ind.setWindowModality(Qt.WindowModal)
        self.Ind.canceled.connect(self.ClosePotok)
        self.Ind.show()
               
        self.potokRasch = OneRasch(self,parent=self)
        self.potokRasch.mysignal.connect(self.on_change, Qt.QueuedConnection)
        #self.potokRasch.mysignal.emit(str("Вася"))
        self.ErrorsShow(data,self.potokRasch)
        
        if not self.potokRasch.isRunning() and data==None:
            self.potokRasch.start()
        
        
            
    def ClosePotok(self):

        self.potokRasch.stop()
        self.potokRasch.mysignal.disconnect()
        del self.potokRasch
    
    def on_change(self, s):
        try:
      
            if s=='Good':
                (self.nm_ivl, self.km, self.per, self.ivl_gr, self.FiA, self.FiB, self.FiC, self.FiT1, self.p, self.yr_v_d,self.Hesh,self.KL_sp,self.per_name) = self.potokRasch.SchitInf()
                self.Ind.close()
                # Создаём графики в отдельных вкладках
                if self.vkl != -1:
                    for i in range(-self.vkl,1):
                        self.TWGr.removeTab(-i) # Удаляем вкладку
                        plt.close(self.Figures[-i]) # Закрываем старые графики (Потому что жрут память)
                        del self.Tabs[-i], self.Figures[-i]\
                            ,self.Canvas[-i], self.GrLayout[-i] # Удаляем все элементы её характеризовавшие
                    self.vkl = -1
                    l=[]
                self.Tiks = []
                if self.marker:
                    self.per_name = Kl.UserReOp(self.KL_sp,self.per_name,self.yr_v_d,self.per)
                else:
                    self.per_name = Kl.UserReOp(self.KL_sp,self.per_name)

                self.gr_list, self.axi_list, self.otp_list, self.point_list =\
                    Kl.PointGen(self.ivl_gr,self.yr_v_d,self.sp_vvl,self.marker,self.per,self.nm_ivl,self.sp_zy,self.km,self.per_name) # Вызываем медод для прорисовки характерных точек на графике
                self.lbl=[]
                
                for i in range(len(self.FiA)):
                    
                    self.vkl +=1
                    self.Tabs.append(QWidget()) # Создаём виджед закладки
                    self.TWGr.addTab(self.Tabs[self.vkl],self.nm_ivl[i]) # Добавляем закладку в QTabWidget
                    self.Figures.append(plt.figure(dpi=75)) # Создаём фигуру графика 
                    self.Canvas.append(FigureCanvas(self.Figures[self.vkl])) # Помещаем фигуру в контейнер
                    self.GrLayout.append(QVBoxLayout()) # Добавляем бокы для заполнение по вертикали
                    # Вставляем схему заземления в контейнер
                    pixmap = QPixmap("result_schemes/"+str(i)+".jpg")
                    self.lbl.append(QLabel(self))
                    pixmap= pixmap.scaled(QSize(665,187), Qt.KeepAspectRatioByExpanding, transformMode = Qt.SmoothTransformation)
                    self.lbl[self.vkl].setPixmap(pixmap)
                    self.lbl[self.vkl].setAlignment(Qt.AlignHCenter)
                
                
                    self.GrLayout[self.vkl].addWidget(self.Canvas[self.vkl],stretch=20) # Помещаем в этот бокс контейнер

                    if self.grafsx:
                        self.GrLayout[self.vkl].addWidget(self.lbl[self.vkl],stretch=0)
                
                    self.Tabs[self.vkl].setLayout(self.GrLayout[self.vkl]) # Распологаем бокс на созданном виджете
                    ax = self.Figures[self.vkl].add_subplot(111) #
                    # Добавляем на график кривую, и задаём цвет и тип линии
                
                    if self.FACh: ax.plot(self.km[i],self.FiA[i],'y',label=self.FAName) 
                    if self.FBCh: ax.plot(self.km[i],self.FiB[i],'g--',label=self.FBName) 
                    if self.FCCh: ax.plot(self.km[i],self.FiC[i],'r-.',label=self.FCName)
                    if self.FTRCh: ax.plot(self.km[i],self.FiT1[i],'b:',label=self.FTRName)

                    ax.plot([self.km[i][0],self.km[i][(len(self.km[i]))-1]],[self.sr_vel,self.sr_vel],'k')
                    
                    mx=1.1*max(max(self.FiA[i])*int(self.FACh),max(self.FiB[i])*int(self.FBCh),\
                            max(self.FiC[i])*int(self.FCCh),max(self.FiT1[i])*int(self.FTRCh))
                    
                    ax.axis([self.km[i][0],self.km[i][(len(self.km[i]))-1],0,mx]) # Задаём область графика

                    for j in self.point_list[i]:
                        ax.text (j[0], mx, j[1], horizontalalignment='center', verticalalignment='bottom')
                        ax.plot([j[0],j[0]],[0,mx],'k:', linewidth=1)
                        
                    
                    ax.grid(True) # Включаем сетку
                    ax.set_xlabel(u'Длина, км',fontsize=self.shr_gr) # Подпись оси х
                    ax.set_ylabel(u'U, В',fontsize=self.shr_gr) # Подпись оси у
                    ax.legend(frameon=False,fontsize=self.shr_gr) # Выводим легенду графика
                    ax.tick_params(labelsize=self.shr_gr) # Размер подписей шкалы графика
                    locator = matplotlib.ticker.MaxNLocator()
                    ax.xaxis.set_major_locator(locator)

                    # Делаем отображение граничных значений на эпюрах
                    T = ax.get_xticks()
                    TT = []
                    t_len = abs(T[0]-T[1])
                    if self.km[i][0] < self.km[i][(len(self.km[i]))-1]:
                        TT.append(self.km[i][0])
                        for h in range((np.shape(T))[0]):
                            if self.km[i][0] < T[h] < self.km[i][(len(self.km[i]))-1] and\
                                self.km[i][(len(self.km[i]))-1]-T[h] > t_len/3 and T[h]-self.km[i][0]>t_len/2:
                                TT.append(T[h])
                        TT.append(self.km[i][(len(self.km[i]))-1])
                    elif self.km[i][0] > self.km[i][(len(self.km[i]))-1]:
                        TT.append(self.km[i][0])
                        for h in range((np.shape(T))[0]):
                            if self.km[i][0] > T[h] > self.km[i][(len(self.km[i]))-1] and\
                                self.km[i][0]-T[h]>t_len/3 and T[h]-self.km[i][(len(self.km[i]))-1]>t_len/2:
                                TT.append(T[h])
                        TT.append(self.km[i][(len(self.km[i]))-1])
                    T = np.array(TT)
                    self.Tiks.append(T)

                    ax.set_xticks(T)
                    ax.set_title("№ опоры\n",fontsize=self.shr_gr)
                    
                    """ 
                    if self.VkOp:
                        ax1=ax.twiny() # Добавляем дополнительную ось Х, twinx() добавляет дополнительную ось Y
                        ax1.axis([self.p[i][0],self.p[i][(len(self.km[i]))-1],0,mx]) # Задаём область графика
                        locator1 = matplotlib.ticker.MaxNLocator (integer=True)# Создаем экземпляр класса, который будет отвечать за расположение меток
                        ax1.xaxis.set_major_locator (locator1)# Установим локатор для главных меток (только целые числа)
                        ax1.set_xlabel('№ опоры',fontsize=self.shr_gr)
                        #ax1.set_xlabel(self.nm_ivl[i]+'\n№ опоры',fontsize=self.shr_gr)
                        ax1.tick_params(labelsize=self.shr_gr) # Размер подписей шкалы графика
                    #else:
                        #ax.set_title(self.nm_ivl[i],fontsize=self.shr_gr)
                     """
                    for obj in self.Figures[self.vkl].findobj(matplotlib.text.Text):
                        obj.set_fontproperties(arial_font)
                        obj.set_fontsize(self.shr_gr)
                    
                    self.Canvas[self.vkl].draw() # Выводим график в виджет
                    self.Figures[self.vkl].set_size_inches(17, 5,forward=True) # Изменяем размер сохраняемого графика
                    self.Figures[self.vkl].savefig('images_grafik/'+str(i)+'.jpg', format='jpg', dpi=100) # Cохраняем графики
            
                self.lists.setCurrentIndex(6) # Делаем активной созданую закладку
            else:
                self.Ind.close()
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('В исходных данных допущена ошибка. Проверте исходные данные. ('+s+')')
        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('В исходных данных допущена ошибка. Проверте исходные данные. ('+str(ex)+')')
        
    # Метод для открытия файлов
    def Open_tables(self):
        # Блок открытия файла (также обрабытываем ошибку при отмене открытия файла)
        try:
            self.Zapis() # читываем данные с таблиц
            sp_ivl0 = self.sp_ivl
            sp_vvl0 = self.sp_vvl
            sp_zy0 = self.sp_zy
            sp_zytr0 = self.sp_zytr
            sp_zvl0 = self.sp_zvl
            
            fname = QFileDialog.getOpenFileName(self, 'Открыть файл', self.pyt,'*.txt*')[0] # Обрати внимание на последний элемент
            if fname != '':
                self.pyt = Kl.adres(fname)
                i = 0
            
                with open(fname, 'r') as f:
                    for line in f:
                        i += 1
                        if i==1:
                            sp_ivl=eval(eval(repr(line)))
                        elif i==2:
                            sp_vvl=eval(eval(repr(line)))
                        elif i==3:
                            sp_zy=eval(eval(repr(line)))
                        elif i==4:
                            sp_zytr=eval(eval(repr(line)))
                        elif i==5:
                            sp_zvl=eval(eval(repr(line)))
                self.Vvod_open_file(sp_ivl,sp_vvl,sp_zy,sp_zytr,sp_zvl)
                
                if fname[-1:-5:-1] == 'txt.':
                    i = -5
                    while True:
                        if fname[i] == '/':
                            break
                        else:
                            i -= 1
                    self.setWindowTitle((fname[-5:i:-1])[::-1]+self.Vive) # Меняем назание титульника на имя открытого файла
                    self.name_vl.setText((fname[-5:i:-1])[::-1]) # Заполняем стрку назания линии именем файла
                self.lists.setCurrentIndex(1) # Делаем активной созданую закладку
        except Exception:
            self.Vvod_open_file(sp_ivl0,sp_vvl0,sp_zy0,sp_zytr0,sp_zvl0)
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Файл поврежён.\n Его открытие не представляется возможным.')
            
    def Vvod_open_file(self,sp_ivl,sp_vvl,sp_zy,sp_zytr,sp_zvl):
        self.correct = [False,False,False,False,False,False]
        # Создадим выпадающие списки для таблиц
        for i in range(len(sp_ivl)):
            self.Spiski(i,1)
        for i in range(len(sp_vvl)):
            self.Spiski(i,2)
        for i in range(len(sp_zy)):
            self.Spiski(i,3)
        for i in range(len(sp_zvl)):
            self.Spiski(i,5)
        # Записываем значения в таблицу ИВЛ
        self.table_IVL.setRowCount(0)
        if len(sp_ivl) != 0:
            for i in range(len(sp_ivl)):
                self.Vvod_ivl(sp_ivl[i],i)
            self.ivl_row = i+1
        else:
            self.ivl_row = 0
        # Записываем значения в таблицу ВВЛ
        self.table_VVL.setRowCount(0)
        self.num_vvl = 0
        self.vvl_pos = {}
        if len(sp_vvl) != 0:
            for i in range(len(sp_vvl)):
                self.Vvod_vvl(sp_vvl[i],i)
                self.combo_2[i][2].activated[str].connect(partial(self.onActivated,i))
            self.vvl_row = i+1
        else:
            self.vvl_row = 0

        # Записываем значения в таблицу Заземление в точке
        self.table_Zy.setRowCount(0)
        if len(sp_zy) != 0:
            for i in range(len(sp_zy)):
                self.Vvod_zy(sp_zy[i],i)
            self.zy_row = i+1
        else:
            self.zy_row = 0
        # Записываем значения в таблицу Заземление троссов ИВЛ
        self.table_Zytr.setRowCount(0)
        if len(sp_zytr) != 0:
            for i in range(len(sp_zytr)):
                self.Vvod_zytr(sp_zytr[i],i)
            self.zytr_row = i+1
        else:
            self.zytr_row = 0
        # Записываем значения в таблицу Заземление троссов ВВЛ
        self.table_Zvl.setRowCount(0)
        if len(sp_zvl) != 0:
            for i in range(len(sp_zvl)):
                self.Vvod_zvl(sp_zvl[i],i)
            self.zvl_row = i+1
        else:
            self.zvl_row = 0
    
    # Ввод строки ИВЛ
    def Vvod_ivl(self,sp_ivl,i):
                self.table_IVL.setRowCount(i+1)
                if sp_ivl[0] != 'Конфигурация опоры' and sp_ivl[0] != '*****':
                    self.table_IVL.setItem(i,0, QTableWidgetItem(sp_ivl[0]))
                    self.table_IVL.item(i,0).setBackground(QColor(162,205,90))
                    self.combo_1[i][5].setCurrentText(sp_ivl[1]) # Делаем активной вкладку списка, которая которая совпадает с заданой в скобках
                    self.table_IVL.setCellWidget(i,1, self.combo_1[i][5])
                    for j in range(2,4):
                        self.table_IVL.setItem(i,j, QTableWidgetItem(sp_ivl[j]))
                        self.table_IVL.item(i,j).setBackground(QColor(162,205,90))
                    self.combo_1[i][6].setCurrentText(sp_ivl[4]) # Делаем активной вкладку списка, которая которая совпадает с заданой в скобках
                    self.table_IVL.setCellWidget(i,4, self.combo_1[i][6])
                    for j in range(5,9):
                        self.table_IVL.setItem(i,j, QTableWidgetItem(sp_ivl[j]))
                        self.table_IVL.item(i,j).setBackground(QColor(162,205,90))
                elif sp_ivl[0] == 'Конфигурация опоры' or sp_ivl[0] == '*****':
                    for j in range(4):
                        self.table_IVL.setItem(i,j, QTableWidgetItem(sp_ivl[j]))
                    self.table_IVL.setItem(i,0, QTableWidgetItem('Конфигурация опоры'))
                    for j in range(4,9):
                        self.combo_1[i][j-4].setCurrentText(sp_ivl[j]) # Делаем активной вкладку списка, которая которая совпадает с заданой в скобках
                        self.table_IVL.setCellWidget(i,j, self.combo_1[i][j-4])
                    

    # Ввод строки ВВЛ
    def Vvod_vvl(self,sp_vvl,i):
                self.table_VVL.setRowCount(i+1)
                if sp_vvl[0] != 'Конфигурация опоры' and sp_vvl[0] != '*****':
                    self.num_vvl+=1
                    self.vvl_pos[i+1]=self.num_vvl
                    for j in range(7):
                        self.table_VVL.setItem(i,j, QTableWidgetItem(sp_vvl[j]))
                        self.table_VVL.item(i,j).setBackground(QColor(162,205,90))
                    self.table_VVL.setItem(i,7, QTableWidgetItem('L'+str(self.num_vvl)))
                    self.table_VVL.setItem(i,8, QTableWidgetItem("Расч. Pdop"))#
                    self.table_VVL.item(i,7).setBackground(QColor(162,205,90))
                    self.table_VVL.item(i,8).setBackground(QColor(162,205,90))
                elif sp_vvl[0] == 'Конфигурация опоры' or sp_vvl[0] == '*****':
                    for j in range(4):
                        self.table_VVL.setItem(i,j, QTableWidgetItem(sp_vvl[j]))
                    self.table_VVL.setItem(i,0, QTableWidgetItem('Конфигурация опоры'))
                    for j in range(4,9):
                        self.combo_2[i][j-4].setCurrentText(sp_vvl[j]) # Делаем активной вкладку списка, которая которая совпадает с заданой в скобках
                        self.table_VVL.setCellWidget(i,j, self.combo_2[i][j-4])
                    

    # Ввод строки заземления
    def Vvod_zy(self,sp_zy,i):
                self.table_Zy.setRowCount(i+1)
                for j in range(4):
                    self.table_Zy.setItem(i,j, QTableWidgetItem(sp_zy[j]))
                for j in range(4,8):
                    self.combo_3[i][j-4].setCurrentText(sp_zy[j]) # Делаем активной вкладку списка, которая которая совпадает с заданой в скобках
                    self.table_Zy.setCellWidget(i,j, self.combo_3[i][j-4])
    # Ввод строки заземления тросов
    def Vvod_zytr(self,sp_zytr,i):
                self.table_Zytr.setRowCount(i+1)
                for j in range(5):
                    self.table_Zytr.setItem(i,j, QTableWidgetItem(sp_zytr[j]))

    # Ввод строки заземления тросов ВВЛ
    def Vvod_zvl(self,sp_zvl,i):
                self.table_Zvl.setRowCount(i+1)
                for j in range(5):
                    self.table_Zvl.setItem(i,j, QTableWidgetItem(sp_zvl[j]))
                for j in range(5,9):
                    self.combo_5[i][j-5].setCurrentText(sp_zvl[j]) # Делаем активной вкладку списка, которая которая совпадает с заданой в скобках
                    self.table_Zvl.setCellWidget(i,j, self.combo_5[i][j-5])

    # Проверка желания закрыть программу
    def closeEvent(self, event):
        
        Message = QMessageBox(QMessageBox.Question,  'Выход из программы',
            "Вы дейстивлеьно хотите выйти?", parent=self)
        Message.addButton('Да', QMessageBox.YesRole)
        Message.addButton('Нет', QMessageBox.NoRole)
        #Message.addButton('Сохранить', QMessageBox.ActionRole)
        reply = Message.exec()
        if reply == 0:
            try:
                self.potokRasch.stop()
            except Exception:
                1
            try:
                self.stree.stop()
            except Exception:
                1
            sp = [(1,int(self.VkPdop),0,''),\
                (2,int(self.grop),0,''),\
                (3,int(self.zpps),0,''),\
                (4,int(self.grafsx),0,''),\
                (5,0,self.sr_vel,''),\
                (6,0,self.k_yvgr,''),\
                (7,int(self.dz),0,''),\
                (8,int(self.FACh ),0,self.FAName),\
                (9,int(self.FBCh),0,self.FBName),\
                (10,int(self.FCCh),0,self.FCName),\
                (11,int(self.FTRCh),0,self.FTRName),
                (12,self.title_f_s.value(),0,""),
                (13,self.f_s.value(),0,""),
                (14,int(self.PaintForm.Trig),0,""),
                (15,0,self.pz,""),\
                (16,int(self.kesh_1 ),0,''),\
                (17,int(self.kesh_2 ),0,''),\
                (18,int(self.kesh_3 ),0,''),\
                (19,int(self.kesh_4 ),0,''),\
                (20,self.ValPZ1.value(),0,''),\
                (21,self.ValPZ2.value(),0,''),\
                (22,int(self.one_sch4 ),0,''),
                (23,self.ValSZ.value(),0,''),
                (24,0,self.ValMaxSZ.value(),''),
                (25,0,self.ValMinSZ.value(),''),
                (26,0,self.ValEpsSZ.value(),''),
                (27,self.ValTooSZ.value(),0,''),
                (28,int(self.too_sz),0,''),
                (29,int(self.one_docx),0,''),
                (30,int(self.dbl_vl),0,''),
                (31,int(self.use_templ),0,'')] 
            BD.Vvod_nasrt(sp)
            BD.CloseBD() # Закрываем базу данных
            pickle.dump( self.Hesh, open( "Hesh_files/dicts.pkl", "wb" ) )
            event.accept()
        elif reply == 1:
            event.ignore()
        """
        elif reply == 2:
            self.Save_tables()
            BD.Vvod_nasrt(self.FTRCh,self.VkOp,self.VkPdop,self.grop,self.zpps,self.grafsx,self.sr_vel,self.k_yvgr)
            BD.CloseBD() # закрываем базу данных
            event.accept()
        """

    def Check_Find_Errors(self,state):
        if state == Qt.Checked:
            self.VkPdop=True
        else:
            self.VkPdop=False

    # Включение выключение граничных опор
    def Check_GrOp(self,state):
        if state == Qt.Checked:
            self.grop=True
        else:
            self.grop=False


    # Включение выключение запрета работы на ПС
    def Check_ZpPS(self,state):
        if state == Qt.Checked:
            self.zpps=True
        else:
            self.zpps=False

    # Включение отображения схемы заземления на графиках
    def Check_sxzgr(self,state):
        if state == Qt.Checked:
            self.grafsx=True
        else:
            self.grafsx=False
    # Включение отображения схемы заземления на графиках
    def Check_DZ(self,state):
        if state == Qt.Checked:
            self.dz=True
        else:
            self.dz=False

    def Check_FA(self,state):
        if state == Qt.Checked:
            self.FACh=True
        else:
            self.FACh=False

    def Check_FB(self,state):
        if state == Qt.Checked:
            self.FBCh=True
        else:
            self.FBCh=False
    
    def Check_FC(self,state):
        if state == Qt.Checked:
            self.FCCh=True
        else:
            self.FCCh=False

    def Check_FTR(self,state):
        if state == Qt.Checked:
            self.FTRCh=True
        else:
            self.FTRCh=False

    def Check_Kesh1(self,state):
        if state == Qt.Checked:
            self.kesh_1=True
        else:
            self.kesh_1=False

    def Check_Kesh2(self,state):
        if state == Qt.Checked:
            self.kesh_2=True
        else:
            self.kesh_2=False
    
    def Check_Kesh3(self,state):
        if state == Qt.Checked:
            self.kesh_3=True
        else:
            self.kesh_3=False
    
    def Check_Kesh4(self,state):
        if state == Qt.Checked:
            self.kesh_4=True
        else:
            self.kesh_4=False

    def Check_one_sch4(self,state):
        if state == Qt.Checked:
            self.one_sch4=True
        else:
            self.one_sch4=False

    
    def Check_too_sz(self,state):
        if state == Qt.Checked:
            self.too_sz=True
        else:
            self.too_sz=False

    def Check_one_file(self,state):
        if state == Qt.Checked:
            self.one_docx=True
            self.Docx_new() 
        else:
            self.one_docx=False
            self.Docx_new()

    def Check_dbl_vl(self,state):
        if state == Qt.Checked:
            self.dbl_vl=True
        else:
            self.dbl_vl=False

    def Check_use_templ(self,state):
        if state == Qt.Checked:
            self.use_templ=True
            self.Docx_new()
        else:
            self.use_templ=False
            self.Docx_new()


    # Кнопка о программе
    def Prog_Inf(self):
        self.Mod=KOp()
        self.Mod.Program_info()
        self.Mod.show()
        

            
    # Сохраняем результаты в Excel
    def Save_Excel(self):     
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.pytEx+self.name_vl.text(),'*.xlsx;;*.xls')[0] # Обрати внимание на последний элемент
            self.pytEx = Kl.adres(fname)
            wb = Workbook()
            # К существующуму листу будем обращаться через ws1
            ws=[]
            for i in range(len(self.nm_ivl)):
                
                if i == 0:
                    ws.append(wb.active)
                    try:
                        ws[i].title = self.nm_ivl[i]
                    except Exception:
                        ws[i].title = str(i+1)
                    ws[i].append(['№','Км','A','B','C','Трос'])
                else:
                    ws.append(wb.create_sheet())
                    try:
                        ws[i].title = self.nm_ivl[i]
                    except Exception:
                        ws[i].title = str(i+1)
                    ws[i].append(['№','Км','A','B','C','Трос'])
                for j in range(len(self.FiA[i])):
                    ws[i].append([self.p[i][j],self.km[i][j],self.FiA[i][j],\
                                  self.FiB[i][j],self.FiC[i][j],self.FiT1[i][j]])
            # Сохранение книги с путём
            wb.save(fname)
        except Exception as ex:
            if str(ex) != "string index out of range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Не получилось сгенерировать массив точек. '+
                                'Проверьте произведён ли перед этим расчёт схемы.'+str(ex))
        else:
            mes = QMessageBox.information(self, 'Генерация массива точек','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)
    # Формируем и сохраняем файл основных исходных данных
    def Oid(self):
        try:
            fname = ""
            name_ins = self.name_vl.text()+" ИД"
            if not self.one_docx:
                fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.pytWIsx+name_ins,'*.docx')[0] # Обрати внимание на последний элемент
                self.pytWIsx = Kl.adres(fname)
            else:
                if self.dbl_vl:
                    if name_ins not in self.d_vl_onedocx:
                        self.d_vl_onedocx.add(name_ins)
                    self.l_vl_onedocx.append(name_ins)
                else:
                    if name_ins not in self.d_vl_onedocx:
                        self.d_vl_onedocx.add(name_ins)
                        self.l_vl_onedocx.append(name_ins)
                    else:
                        raise Exception("Данный файл уже добавлен")
            self.onChange("print")       
            
            docx_copy = self.Docx_file 
            docx_copy = Oid.WordOid(self.sp_vvl, self.sp_pzy, self.nm_d, self.ivl_ris,fname,  self.name_vl.text(),self.grop,docx_copy,self.one_docx)
            if self.one_docx and docx_copy!=None:
                self.Docx_file = docx_copy
            
        except Exception as ex:
            if str(ex) != "string index out of range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                if str(ex) == "Данный файл уже добавлен":
                    ems.showMessage(str(ex))
                else:
                    ems.showMessage('Не получилось сгенерировать отчёт. '+
                                    'Возможно произошла попытка сохранить в файл, который сейчас открыт. '+str(ex))
        else:
            mes = QMessageBox.information(self, 'Генерация отчёта','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)


    def CorrectionSxZ(self,btn):
        try:
            per_name = Kl.Yz_p(self.ivl_gr,self.yr_v_d)
            per_name = Kl.UserReOp(self.KL_sp,per_name,self.yr_v_d if self.marker else None,self.per if self.marker else None)
            sp_pit = Kl.PodpSz(self.ivl_gr, self.yr_v_d)
            if self.marker:
                sp_b, per=Kl.RePer(self.yr_v_d,self.per)
                m, n = Kl.Otpaiki(sp_b,self.ivl_gr,self.nm_ivl)           
                inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per),self.nm_ivl,Kl.ReZy(sp_b,self.sp_zy))
                f, f1, f2 = Kl.Zapret(self.FiA, self.FiB, self.FiC,self.FiT1,self.p, self.sr_vel,self.FACh,self.FBCh,self.FCCh,self.FTRCh)
                kl_list = Kl.KL_metki(self.KL_sp,self.yr_v_d)
                re_op_key = list(per_name.keys())
                ind = self.TWGr.currentIndex()
                f3 = [[False,False,False,False] for i in range(len(f))]

                if self.correct[0]:
                    f,f1,f2,inf,f3 = self.correct[1:]

                if btn==5:
                    rez=Kl.ReZaprYch(self.tx_op1.value(),self.tx_op2.value(),f,f1,f2,ind,False)
                    if rez!=None:
                        f,f1,f2=rez
                elif btn ==6:
                    rez=Kl.ReZaprYch(self.tx_op1.value(),self.tx_op2.value(),f,f1,f2,ind,True)
                    if rez!=None:
                        f,f1,f2=rez
                elif 1<=btn<=4:
                    inf,f2 = Kl.PZandZpar(inf,f2,ind,btn)
                    if btn==3:
                        f3[ind][0]=True
                        f3[ind][1]=f2[ind][inf[ind][2]]
                    if btn==4:
                        f3[ind][2]=True
                        f3[ind][3]=f2[ind][inf[ind][3]]
                    
                self.correct=[True,f,f1,f2,inf,f3]
                
                Sz.Sxeme(inf[ind],f1[ind],lbsz[ind],m[ind],n[ind],self.grop,self.zpps,self.dz,per_name[re_op_key[ind]],sp_pit[ind],kl_list[ind],rzapr=f3[ind])

                pixmap = QPixmap("result_schemes/"+str(ind)+".jpg")
                pixmap= pixmap.scaled(QSize(665,187), Qt.KeepAspectRatioByExpanding, transformMode = Qt.SmoothTransformation)
                self.lbl[ind].setPixmap(pixmap)
                self.lbl[ind].setAlignment(Qt.AlignHCenter)


            else:
                inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(self.per),self.nm_ivl,self.sp_zy)
                f, f1, f2 = Kl.Zapret(self.FiA, self.FiB, self.FiC,self.FiT1,self.p, self.sr_vel,self.FACh,self.FBCh,self.FCCh,self.FTRCh)
                kl_list = Kl.KL_metki(self.KL_sp)
                re_op_key = list(per_name.keys())
                ind = self.TWGr.currentIndex()
                f3 = [[False,False,False,False] for i in range(len(f))]

                if self.correct[0]:
                    f,f1,f2,inf,f3 = self.correct[1:]
                
                if btn==5:
                    rez=Kl.ReZaprYch(self.tx_op1.value(),self.tx_op2.value(),f,f1,f2,ind,False)
                    if rez!=None:
                        f,f1,f2=rez
                elif btn ==6:
                    rez=Kl.ReZaprYch(self.tx_op1.value(),self.tx_op2.value(),f,f1,f2,ind,True)
                    if rez!=None:
                        f,f1,f2=rez
                elif 1<=btn<=4:
                    inf,f2 = Kl.PZandZpar(inf,f2,ind,btn)
                    if btn==3:
                        f3[ind][0]=True
                        f3[ind][1]=f2[ind][inf[ind][2]]
                    if btn==4:
                        f3[ind][2]=True
                        f3[ind][3]=f2[ind][inf[ind][3]]

                self.correct=[True,f,f1,f2,inf,f3]
                
                Sz.Sxeme(inf[ind],f1[ind],lbsz[ind],[],[],self.grop,self.zpps,self.dz,per_name[re_op_key[ind]],sp_pit[ind],kl_list[ind],rzapr=f3[ind])

                pixmap = QPixmap("result_schemes/"+str(ind)+".jpg")
                pixmap= pixmap.scaled(QSize(665,187), Qt.KeepAspectRatioByExpanding, transformMode = Qt.SmoothTransformation)
                self.lbl[ind].setPixmap(pixmap)
                self.lbl[ind].setAlignment(Qt.AlignHCenter)
        except Exception:
            1



    # Отчет по схемам 1, 2, 4
    def OtW(self):
        tab_ep =self.tablepo.currentText()
        if tab_ep =='Выводить таблицы и эпюры':
            te=''
        elif tab_ep =='Выводить таблицы':
            te=' таблицы'
        elif tab_ep =='Выводить эпюры':
            te=' эпюры'    
        try:
            fname = ""
            name_ins = self.name_vl.text()+" Cx. "+self.N_sxem.currentText()+"."+self.S_one_var.text()+te
            if not self.one_docx:
                fname = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.pytWRez+name_ins,'*.docx')[0] # Обрати внимание на последний элемент
                self.pytWRez = Kl.adres(fname)
            else:
                if self.dbl_vl:
                    if name_ins not in self.d_vl_onedocx:
                        self.d_vl_onedocx.add(name_ins)
                    self.l_vl_onedocx.append(name_ins)
                else:
                    if name_ins not in self.d_vl_onedocx:
                        self.d_vl_onedocx.add(name_ins)
                        self.l_vl_onedocx.append(name_ins)
                    else:
                        raise Exception("Данный файл уже добавлен")

            per_name = Kl.Yz_p(self.ivl_gr,self.yr_v_d)
            
            per_name = Kl.UserReOp(self.KL_sp,per_name,self.yr_v_d if self.marker else None,self.per if self.marker else None)
            

            sp_pit = Kl.PodpSz(self.ivl_gr, self.yr_v_d)
            VVL_dict, nm_d = Kl.Graph(self.sp_vvl, self.ivl_gr)
            self.gr_list, self.axi_list, self.otp_list, self.point_list =\
                 Kl.PointGen(self.ivl_gr,self.yr_v_d,self.sp_vvl,self.marker,self.per,self.nm_ivl,self.sp_zy,self.km,per_name)
 
            
            if self.marker:

                sp_b, per=Kl.RePer(self.yr_v_d,self.per)
                m, n = Kl.Otpaiki(sp_b,self.ivl_gr,self.nm_ivl)
                
                n_o = [list(per[key].keys()) for key in per]
                yr_list = list(self.yr_v_d.keys())
                
                inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per),self.nm_ivl,Kl.ReZy(sp_b,self.sp_zy))
                f, f1, f2 = Kl.Zapret(self.FiA, self.FiB, self.FiC,self.FiT1,self.p, self.sr_vel,self.FACh,self.FBCh,self.FCCh,self.FTRCh)

                kl_list = Kl.KL_metki(self.KL_sp,self.yr_v_d)
                
                re_op_key = list(per_name.keys())

                f3 = [[False,False,False,False] for i in range(len(f))]
                if self.correct[0]:
                    f,f1,f2,inf,f3 = self.correct[1:]
                
                for i in range(len(self.nm_ivl)):
                    Sz.Sxeme(inf[i],f1[i],lbsz[i],m[i],n[i],self.grop,self.zpps,self.dz,per_name[re_op_key[i]],sp_pit[i],kl_list[i],rzapr=f3[i])

                    Gr_sblizh.Sx_sb(i,self.gr_list[i],self.axi_list[i],self.shr_gr, self.Tiks[i], self.otp_list[i])

                    
 
            else:
                inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(self.per),self.nm_ivl,self.sp_zy)
                f, f1, f2 = Kl.Zapret(self.FiA, self.FiB, self.FiC,self.FiT1,self.p, self.sr_vel,self.FACh,self.FBCh,self.FCCh,self.FTRCh)
                n_o = [ list(self.per[key].keys()) for key in self.per]
                kl_list = Kl.KL_metki(self.KL_sp)
                
                re_op_key = list(per_name.keys())

                f3 = [[False,False,False,False] for i in range(len(f))]
                if self.correct[0]:
                    f,f1,f2,inf,f3 = self.correct[1:]

                for i in range(len(self.nm_ivl)):
                    Sz.Sxeme(inf[i],f1[i],lbsz[i],[],[],self.grop,self.zpps,self.dz,per_name[re_op_key[i]],sp_pit[i],kl_list[i],rzapr=f3[i])
                    
                    Gr_sblizh.Sx_sb(i,self.gr_list[i],self.axi_list[i],self.shr_gr, self.Tiks[i], self.otp_list[i])

            docx_copy = self.Docx_file
            docx_copy = OtchetW.Word(inf, lbsz, f, f2, fname, self.name_vl.text(),\
                self.N_sxem.currentText(),self.S_one_var.text(),self.grop,self.zpps,per_name, docx_copy,self.one_docx,tab_ep)    
            if self.one_docx and docx_copy!=None:
                self.Docx_file = docx_copy 
                           
        except Exception as ex:
            if str(ex) != "string index out of range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                if str(ex) == "Данный файл уже добавлен":
                    ems.showMessage(str(ex))
                else:
                    ems.showMessage('Не получилось сгенерировать отчёт. '+
                                    'Возможно произошла попытка сохранить в файл, который сейчас открыт. '+str(ex))
        else:
            mes = QMessageBox.information(self, 'Генерация отчёта','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)
    # Отчет по схеме №3
    def SxT(self):
        try:
            try:
               self.stree.isRunning()
               print(1)
            except Exception:
                self.fname3 = ""
                name_ins = self.name_vl.text()+" Cx. "+self.N_sxem.currentText()
                if not self.one_docx:
                    self.fname3 = QFileDialog.getSaveFileName(self, 'Сохранить файл', self.pytWRez+name_ins,'*.docx')[0] # Обрати внимание на последний элемент
                    self.pytWRez = Kl.adres(self.fname3)
                else:
                    if self.dbl_vl:
                        if name_ins not in self.d_vl_onedocx:
                            self.d_vl_onedocx.add(name_ins)
                        self.l_vl_onedocx.append(name_ins)
                    else:
                        if name_ins not in self.d_vl_onedocx:
                            self.d_vl_onedocx.add(name_ins)
                            self.l_vl_onedocx.append(name_ins)
                        else:
                            raise Exception("Данный файл уже добавлен")
                # Диалоговое окно индикации процесса
                self.Ind2 = QProgressDialog('Обработка данных','Отмена', 0, 0, self)
                self.Ind2.setWindowTitle('Сxемы № 3')
                self.Ind2.setWindowModality(Qt.NonModal)
                #self.Ind2.setWindowModality(Qt.WindowModal)
                self.Ind2.canceled.connect(lambda:self.CloseTreeSx(1))
                self.Ind2.show()
                self.Zapis() # читываем данные с таблиц
            
                self.stree = SxemThree(self,parent=self)
                self.stree.mysignal.connect(partial(self.TreeSEv,1), Qt.QueuedConnection)
            
                if not self.stree.isRunning():
                    self.stree.start()
            
        except Exception as ex:
            if str(ex) != "string index out of range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                if str(ex) == "Данный файл уже добавлен":
                    ems.showMessage(str(ex))
                else:
                    ems.showMessage('Не получилось сгенерировать отчёт. '+
                                'Проверьте введёно ли сопротивление БЗ. '+str(ex))
                


    def TreeSEv(self,a,s):
        if a==1:      
            if s[:8]=='mistake:':
                self.Ind2.close()
                del self.Ind2
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('В исходных данных допущена ошибка. Проверте исходные данные ('+s[8:]+')')
            elif s == 'Good':
                self.Ind2.close()
                del self.Ind2
                mes = QMessageBox.information(self, 'Генерация отчёта','Операция прошла успешно.',
                                            buttons=QMessageBox.Ok,
                                            defaultButton=QMessageBox.Ok)
            elif s[:8] != 'mistake:' and s != 'Good':
                self.Ind2.setLabelText(s)
        elif a==2:
            if s[:8]=='mistake:':
                self.Ind3.close()
                del self.Ind3
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('В исходных данных допущена ошибка. Проверте исходные данные ('+s[8:]+')')
            elif s == 'Good':
                self.Ind3.close()
                del self.Ind3
                mes = QMessageBox.information(self, 'Генерация отчётов','Операция прошла успешно.',
                                            buttons=QMessageBox.Ok,
                                            defaultButton=QMessageBox.Ok)
            elif s[:8] != 'mistake:' and s != 'Good':
                self.Ind3.setLabelText(s)
        elif a==3:
            if s[:8]=='mistake:':
                self.Ind4.close()
                del self.Ind4
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('В исходных данных допущена ошибка. Проверте исходные данные ('+s[8:]+')')
            elif s == 'Good':
                try:
                    sp_zy = self.stree2.SchitInf()
                    self.Zapis()
                    self.Vvod_open_file(self.sp_ivl,self.sp_vvl,sp_zy,self.sp_zytr,self.sp_zvl)
                    self.lists.setCurrentIndex(3)
                except Exception as ex:
                    print(ex)
                self.Ind4.close()
                del self.Ind4
                mes = QMessageBox.information(self, 'Генерация СЗ','Операция прошла успешно.',
                                            buttons=QMessageBox.Ok,
                                            defaultButton=QMessageBox.Ok)
                
            elif s[:8] != 'mistake:' and s != 'Good':
                self.Ind4.setLabelText(s)
            

    def CloseTreeSx(self,a):
        if a == 1:
            self.stree.stop()
            self.stree.mysignal.disconnect()
            del self.stree
        elif a==2:
            self.stree1.stop()
            self.stree1.mysignal.disconnect()
            del self.stree1
        elif a==3:
            self.stree2.stop()
            self.stree2.mysignal.disconnect()
            del self.stree2
            

    def Kesh_label(self):
        """ Рачитаем размер кэша """
        try:
            dirs = ['Hesh_files/','gr_sb/','images_grafik/','Op_Fig/','result_schemes/']
            files_path = [os.listdir(i) for i in dirs]
            files = sum([len(i) for i in files_path])
            size = 0
            for i in range(len(dirs)):
                size += sum([os.path.getsize(dirs[i]+f) if os.path.isfile(dirs[i]+f) else 0  for f in files_path[i]])
            self.kol_files.setText('Файлов, шт: '+str(files))
            self.size_files.setText('Размер Кэша, МБ: '+str(round(size/1000000,2)))
        except Exception:
            1

    def Clearing_Kesh(self):
        """ Метод очистки кеша """
        Message = QMessageBox(QMessageBox.Question,  'Очистка кэша',\
        """ Вы дейстивлеьно хотите удалить файлы кэша? Все раннее расчитанные данные будут утеряны. """,\
             parent=self)
        Message.addButton('Да', QMessageBox.YesRole)
        Message.addButton('Нет', QMessageBox.NoRole)
        reply = Message.exec()
        if reply == 0:
            try:
                dirs = ['Hesh_files/','gr_sb/','images_grafik/','Op_Fig/','result_schemes/']
                for d in dirs:
                    for root, dirs, files in os.walk(d):
                        for f in files:
                            os.unlink(os.path.join(root, f))
                self.Hesh = [{},{},{}]
            except Exception as ex:
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('При очистке возникла ошибка ('+str(ex)+').')
            else:
                self.Kesh_label()
                mes = QMessageBox.information(self, 'Очистка кэша','Операция прошла успешно.',
                                                buttons=QMessageBox.Ok,
                                                defaultButton=QMessageBox.Ok)
        

    def AllSchemsStart(self):
        """ Метод для запуска расчёта всех схем """
        
        try:
            try:
               self.stree1.isRunning()
            except Exception:
                self.fname4=""
                if not self.one_docx:
                    self.fname4 = QFileDialog.getExistingDirectory(self, 'Сохранить файлы', self.pytWRez) # Обрати внимание на последний элемент
                    if self.fname4 =="": raise Exception("string index out of range")
                    self.pytWRez = self.fname4+'//'
                #+self.name_vl.text()+" Cx. "+self.N_sxem.currentText()
                # Диалоговое окно индикации процесса
                self.Ind3 = QProgressDialog('Обработка данных','Отмена', 0, 0, self)
                self.Ind3.setWindowTitle('Все схемы')
                self.Ind3.setWindowModality(Qt.NonModal)
                #self.Ind2.setWindowModality(Qt.WindowModal)
                self.Ind3.canceled.connect(lambda:self.CloseTreeSx(2))
                self.Ind3.show()
                self.Zapis() # читываем данные с таблиц
            
                self.stree1 = AllSchemsCalculate(self,parent=self)
                self.stree1.mysignal.connect(partial(self.TreeSEv,2), Qt.QueuedConnection)
            
                if not self.stree1.isRunning():
                    self.stree1.start()
            
        except Exception as ex:
            if str(ex) != "string index out of range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Не получилось сгенерировать отчёт. '+
                                'Проверьте введёно сопротивление БЗ. '+str(ex))

    def SZGenerator(self):
        """ Метод подбора мест для установки СЗ """
        try:
            try:
               self.stree2.isRunning()
            except Exception:
                # Диалоговое окно индикации процесса
                self.Ind4 = QProgressDialog('Процес подбора','Отмена', 0, 0, self)
                self.Ind4.setWindowTitle('Подбор СЗ')
                self.Ind4.setWindowModality(Qt.NonModal)
                #self.Ind2.setWindowModality(Qt.WindowModal)
                self.Ind4.canceled.connect(lambda:self.CloseTreeSx(3))
                self.Ind4.show()
                self.Zapis() # читываем данные с таблиц
            
                self.stree2 = SZGeneratorCalculate(self,parent=self)
                self.stree2.mysignal.connect(partial(self.TreeSEv,3), Qt.QueuedConnection)
            
                if not self.stree2.isRunning():
                    self.stree2.start()

                
            
        except Exception as ex:
            if str(ex) != "string index out of range":
                ems = QErrorMessage(self)
                ems.setWindowTitle('Возникла ошибка')
                ems.showMessage('Не получилось сгенерировать отчёт. '+
                                'Проверьте введёно сопротивление БЗ. '+str(ex))

class SZGeneratorCalculate(QThread):
    mysignal = QtCore.pyqtSignal(str)
    def __init__(self,b, parent=None):
        QtCore.QThread.__init__(self, parent)
        self.b=b
    def run(self):
        try:
            self.spzy = copy.deepcopy(self.b.sp_zy)
            for i in range(len(self.spzy)):
                if self.spzy[i][7] == 'ПЗ':
                    self.spzy[i][4] = '3ф с тр.'
                else:
                    self.spzy[i][4] = 'Разземлен'

            self.ivl, nm_ivl, km, self.per, yz, ivl_gr, yr_v,KL_sp = Kl.Sort_IVL(self.b.sp_ivl)
            self.vvl = Kl.Sort_VVL(self.b.sp_vvl,self.per,yz)
            self.zy = Kl.Sort_Zy(self.spzy,self.per,yz)
            self.zytr = Kl.Sort_Zytr(self.b.sp_zytr,self.per,yz)
            self.zvl = Kl.Sort_Zvl(self.b.sp_zvl)


            self.per_list = list(self.per.keys())
            # Хэшируем запускаемый расчёт
            k_ivl = [tuple(i) for i in self.ivl]
            k_vvl1 = []
            k_vvl2 = []
            for i in self.vvl:
                if i[0]>0: 
                    k_vvl1.append(tuple(i))
                else: 
                    k_vvl2.append(tuple(i))
            k_zy = [tuple(i) for i in self.zy]
            k_zytr = [tuple(i) for i in self.zytr]
            k_zvl1 = [i[0] for i in self.ivl]
            k_zvl2 = [tuple(i[1:9]) for i in self.zvl]

            self.Hesh_now = [tuple(k_ivl+k_vvl2+k_zvl2+[self.b.pz])]
            k_1 = self.b.Hesh[0].get(self.Hesh_now[0],len(self.b.Hesh[0]))
            self.Hesh_now.append(tuple(k_zy+k_zytr+k_zvl1+[k_1]))
            self.Hesh_now.append(tuple(k_vvl1+[k_1]))
            self.Kesh = [self.b.kesh_1,self.b.kesh_2,self.b.kesh_3]

            e = self.b.ValEpsSZ.value()
            left_R = self.b.ValMinSZ.value()
            right_R = self.b.ValMaxSZ.value()
            
            self.spsz_now = []
            d_sz = {}

            R_now = []
            r_r = []
            l_r = []
            ab = []
            
            for i in range(self.b.ValSZ.value()):
                Rez, Trig = self.OneRasch()
                if not Trig: break
                k=101
                for j in range(len(Rez)):
                    if tuple(Rez[j][1:4]) not in d_sz:
                        d_sz[tuple(Rez[j][1:4])] = i
                        R_now.append(right_R)
                        self.spsz_now.append([str(Rez[j][1]),str(Rez[j][2]),str(Rez[j][3])]+[str(R_now[i]),'3ф с тр.','A','A','СЗ'])
                        l_r.append(left_R) 
                        r_r.append(right_R)
                        ab.append(Rez[j][4:])
                        k = 0
                        break
                
                while k<=100:
                    k+=1
                    #print(R_now[i], k, r_r[i]-l_r[i], r_r[i], l_r[i], self.MX[ab[i][0]][ab[i][1]])
                    self.spsz_now[i][3] = str(R_now[i])
                    self.zy = Kl.Sort_Zy(self.spzy+self.spsz_now,self.per,yz) 
                    Rez, Trig = self.OneRasch()    
                    if r_r[i]-l_r[i]<e: # and  (self.b.sr_vel >= self.MX[ab[i][0]][ab[i][1]] or R_now[i]+e >= right_R or R_now[i]-e <= left_R or r_r[i]-l_r[i]==0)
                        break 
                    if self.b.sr_vel >= self.MX[ab[i][0]][ab[i][1]]:
                        dR = round((r_r[i]-R_now[i])/2,3)
                        l_r[i] = R_now[i]
                        R_now[i] = round(R_now[i]+dR,3)
                        
                    else:
                        dR = round((R_now[i]-l_r[i])/2,3)
                        r_r[i] = R_now[i]
                        R_now[i] = round(R_now[i]-dR,3)
            k2 = 0    
            
            Trig_list = [0]*len(self.spsz_now)        
            if (Trig or self.b.too_sz) and len(self.spsz_now)>1:
                print("Этап 2")  
                while k2<=20 and sum(Trig_list) !=len(self.spsz_now):
                    k2+=1
                    if self.b.too_sz:
                        if not Trig and k2>self.b.ValTooSZ.value(): break
                        if k2<=self.b.ValTooSZ.value():
                            l_r = [left_R for i in range(len(self.spsz_now))]
                            r_r = [right_R for i in range(len(self.spsz_now))]
                    else:
                        if not Trig: break
                    #print(Trig_list)
                    for i in range(len(self.spsz_now)):
                        k=0
                        if self.b.sr_vel < self.MX[ab[i][0]][ab[i][1]]:
                            l_r[i] = left_R
                        while k<=100:
                            k+=1
                            #print(R_now[i], k, r_r[i]-l_r[i], r_r[i], l_r[i], self.MX[ab[i][0]][ab[i][1]])
                            self.spsz_now[i][3] = str(R_now[i])
                            self.zy = Kl.Sort_Zy(self.spzy+self.spsz_now,self.per,yz) 
                            Rez, Trig = self.OneRasch()    
                            if r_r[i]-l_r[i]<e : #and  (self.b.sr_vel >= self.MX[ab[i][0]][ab[i][1]] or R_now[i]+e >= right_R or R_now[i]-e <= left_R or r_r[i]-l_r[i]==0)
                                if k==1: 
                                    Trig_list[i]=1
                                else:
                                    Trig_list[i]=0
                                break 
                            if self.b.sr_vel >= self.MX[ab[i][0]][ab[i][1]]:
                                dR = round((r_r[i]-R_now[i])/2,3)
                                l_r[i] = R_now[i]
                                R_now[i] = round(R_now[i]+dR,3)
                                
                            else:
                                dR = round((R_now[i]-l_r[i])/2,3)
                                r_r[i] = R_now[i]
                                R_now[i] = round(R_now[i]-dR,3)
            k2 = 0 
            koef = [1.0 for i in range(len(self.spsz_now))]
            
            if Trig and len(self.spsz_now)>0:
                print("Этап 3")
                while k2<=20 and Trig:
                    k2+=1
                    for i in range(len(self.spsz_now)):
                        koef[i]-=0.05
                        k=0
                        l_r[i] = left_R
                        R_now[i] = round(abs((R_now[i]-left_R)*koef[i])+left_R,3)
                        r_r[i] = R_now[i]
                        while k<=100:
                            k+=1
                            #print(R_now[i], k, r_r[i]-l_r[i], r_r[i], l_r[i], self.MX[ab[i][0]][ab[i][1]])
                            self.spsz_now[i][3] = str(R_now[i])
                            self.zy = Kl.Sort_Zy(self.spzy+self.spsz_now,self.per,yz) 
                            Rez, Trig = self.OneRasch()    
                            if r_r[i]-l_r[i]<e : break
                            if self.b.sr_vel >= self.MX[ab[i][0]][ab[i][1]]:
                                dR = round((r_r[i]-R_now[i])/2,3)
                                l_r[i] = R_now[i]
                                R_now[i] = round(R_now[i]+dR,3)
                                
                            else:
                                dR = round((R_now[i]-l_r[i])/2,3)
                                r_r[i] = R_now[i]
                                R_now[i] = round(R_now[i]-dR,3)


            #print(self.spsz_now)
            #print(Trig)
            #print(Rez)
                
            self.mysignal.emit('Good')
        except Exception as ex:
            self.mysignal.emit('mistake:'+str(ex))

    def SchitInf(self):
        return self.spzy+self.spsz_now

    def OneRasch(self):
        Bool = [self.Hesh_now[0] in self.b.Hesh[0], False, self.Hesh_now[2] in self.b.Hesh[2]]
        Bool_hesh = [Bool[i] and self.Kesh[i] for i in range(3)]
        File_hesh = [self.b.Hesh[0].setdefault(self.Hesh_now[0],len(self.b.Hesh[0])),\
            len(self.b.Hesh[1]),self.b.Hesh[2].setdefault(self.Hesh_now[2],len(self.b.Hesh[2]))]
        #print(Bool_hesh)
        #print(File_hesh)

        # Запускаем расчет, получаем результаты
        try:
            FiA, FiB, FiC, FiT1, p\
            = rnn.Vvod_inf(self.ivl, self.vvl, self.zy, self.zytr, self.zvl, self.per, self.b.k_yvgr,self.b.pz,Bool_hesh,File_hesh,False)
        except Exception as ex:
            if not Bool[0]:
                del self.b.Hesh[0][self.Hesh_now[0]]
            if not Bool[1]:
                del self.b.Hesh[1][self.Hesh_now[1]]
            if not Bool[2]:
                del self.b.Hesh[2][self.Hesh_now[2]]
            raise Exception(str(ex))

        self.MX = []
        Ekst = []
        Point = []
        D = {}
        for i in range(len(FiA)):
            self.MX.append(np.maximum(np.maximum(np.array(FiA[i],dtype=np.float64),np.array(FiB[i],dtype=np.float64)),\
                np.array(FiC[i],dtype=np.float64),(np.array(FiT1[i],dtype=np.float64))))
            Ekst.append(argrelextrema(self.MX[i],np.greater_equal)) # np.greater_equal
            Point+=[[self.MX[i][Ekst[i][0][j]],i,p[i][Ekst[i][0][j]],Ekst[i][0][j]] for j in range(np.shape(Ekst[i][0])[0])]

        for i in Point:
            if i[0] not in D:
                D[i[0]]=i[1:]
        Rez = []
        Volt=sorted(list(D.keys()),reverse=True)
        for i in range(len(Volt)):
            a = [Volt[i],self.per_list[D[Volt[i]][0]][0],self.per_list[D[Volt[i]][0]][1],D[Volt[i]][1],D[Volt[i]][0],D[Volt[i]][2]]
            Trig = True
            for j in range(len(self.spzy)):
                if a[1:4] == self.spzy[j][:3] and self.spzy[j][7] == "ПЗ":
                    Trig = False
                    break
            if Trig: Rez.append(a)
        
        V = 0
        for i in range(len(self.MX)):  
            V = max(V,max(self.MX[i]))             
        
        return Rez, self.b.sr_vel<V

    def stop( self ):
        print('stop')
        self.terminate()
        self.wait()
        return 'stop'

class AllSchemsCalculate(QThread):
    mysignal = QtCore.pyqtSignal(str)
    def __init__(self,b, parent=None):
        QtCore.QThread.__init__(self, parent)
        self.b=b
    def run(self):
        try:
            tab_ep =self.b.tablepo.currentText()
            if tab_ep =='Выводить таблицы и эпюры':
                te=''
            elif tab_ep =='Выводить таблицы':
                te=' таблицы'
            elif tab_ep =='Выводить эпюры':
                te=' эпюры'
            ivl, nm_ivl, km, per, yz, ivl_gr, yr_v, KL_sp = Kl.Sort_IVL(self.b.sp_ivl)
            zy = Kl.Sort_Zy(self.b.sp_zy,per,yz)
            #self.mysignal.emit('Расчет Схем №2')
            a=self.b.ValPZ1.value()-1
            b=self.b.ValPZ2.value()-1
            l_zy = len(self.b.sp_zy)
            if l_zy<2:
                raise Exception('Не верно введены заземляющие устройства')
            if a>l_zy-1 or b>l_zy-1:
                raise Exception('Не верно введены заземляющие устройства')
            if self.b.sp_zy[a][7] !='ПЗ' or self.b.sp_zy[b][7] != 'ПЗ':
                raise Exception('Не верно введены заземляющие устройства')

            sp_PZ = []
            for i in range(l_zy):
                if self.b.sp_zy[i][7] == 'ПЗ' and i!=a and i!=b:
                    sp_PZ.append(i)

            sp_SZ = []
            for i in range(l_zy):
                if self.b.sp_zy[i][7] == 'СЗ':
                    sp_SZ.append(i)
            
            comb_PZ = [thing for thing in itertools.product([0, 1], repeat=len(sp_PZ))]
            comb_SZ = [thing for thing in itertools.product([0, 1], repeat=len(sp_SZ))]
            comb_Base = [(1,1),(1,0),(0,1),(1,1)]
            spzy = copy.deepcopy(self.b.sp_zy)
            Ns = ['1','2','2','4']
            #print(comb_PZ)
            Text_sch = self.b.Sch_val.currentText()
            if Text_sch == "Все схемы":
                iterabl = [0,1,2,3]
            elif Text_sch == "Схема №1":
                iterabl = [0]
            elif Text_sch == "Схема №2":
                iterabl = [1,2]
            elif Text_sch == "Схема №4":
                iterabl = [3]
            elif Text_sch == "Схема №1,2":
                iterabl = [0,1,2]
            elif Text_sch == "Схема №1,4":
                iterabl = [0,3]
            elif Text_sch == "Схема №2,4":
                iterabl = [1,2,3]
            for i in iterabl:
                zy[a][2]=comb_Base[i][0]
                zy[b][2]=comb_Base[i][1]
                spzy[a][4] = '3ф с тр.' if comb_Base[i][0] else 'Разземлен'
                spzy[b][4] = '3ф с тр.' if comb_Base[i][1] else 'Разземлен'
                if 0<=i<=2:
                    if i != 2:
                        NV = 0
                    for j in comb_SZ:
                        for m in range(len(j)):
                            zy[sp_SZ[m]][2] = 1 if j[m] else 7
                            spzy[sp_SZ[m]][4] = '3ф с тр.' if j[m] else 'Разземлен'
                        break
                    for j in comb_PZ:
                        NV+=1
                        for m in range(len(j)):
                            zy[sp_PZ[m]][2] = 1 if j[m] else 7
                            spzy[sp_PZ[m]][4] = '3ф с тр.' if j[m] else 'Разземлен'
                        self.mysignal.emit('Схема №'+Ns[i]+"."+str(NV))
                        name_ins = self.b.name_vl.text()+" Cx. № "+Ns[i]+"."+str(NV)+te
                        fname = self.b.pytWRez+name_ins+'.docx'
                        if self.b.one_docx:                               
                            if self.b.dbl_vl:
                                if name_ins not in self.b.d_vl_onedocx:
                                    self.b.d_vl_onedocx.add(name_ins)
                                self.b.l_vl_onedocx.append(name_ins)
                            else:
                                if name_ins not in self.b.d_vl_onedocx:
                                    self.b.d_vl_onedocx.add(name_ins)
                                    self.b.l_vl_onedocx.append(name_ins)
                                else:
                                    self.mysignal.emit('Ранее уже добавлен')
                                    continue

                        self.OneR(zy,spzy,Ns[i],str(NV), fname)
                elif i == 3:
                    NV = 0
                    M = len(comb_PZ[0])
                    if M != 0:
                        M = len(comb_PZ)
                        for m in range(len(comb_PZ[M-1])):
                            zy[sp_PZ[m]][2] = 1 if comb_PZ[M-1][m] else 7
                            spzy[sp_PZ[m]][4] = '3ф с тр.' if comb_PZ[M-1][m] else 'Разземлен'
                    
                    M = len(comb_SZ)
                    for j in range(1,M):
                        if self.b.one_sch4 and M-1 != j: continue
                        NV+=1
                        for m in range(len(comb_SZ[j])):
                            zy[sp_SZ[m]][2] = 1 if comb_SZ[j][m] else 7
                            spzy[sp_SZ[m]][4] = '3ф с тр.' if comb_SZ[j][m] else 'Разземлен'
                        self.mysignal.emit('Схема №'+Ns[i]+"."+str(NV))
                        name_ins = self.b.name_vl.text()+" Cx. № "+Ns[i]+"."+str(NV)+te
                        fname = self.b.pytWRez+name_ins+'.docx'
                        if self.b.one_docx:                               
                            if self.b.dbl_vl:
                                if name_ins not in self.b.d_vl_onedocx:
                                    self.b.d_vl_onedocx.add(name_ins)
                                self.b.l_vl_onedocx.append(name_ins)
                            else:
                                if name_ins not in self.b.d_vl_onedocx:
                                    self.b.d_vl_onedocx.add(name_ins)
                                    self.b.l_vl_onedocx.append(name_ins)
                                else:
                                    self.mysignal.emit('Ранее уже добавлен')
                                    continue
                        self.OneR(zy,spzy,Ns[i],str(NV), fname)


            self.mysignal.emit('Good')
        except Exception as ex:
            self.mysignal.emit('mistake:'+str(ex))

    def OneR(self,zy,spzy,NS,NV, fname):
        ivl, nm_ivl, km, per, yz, ivl_gr, yr_v, KL_sp = Kl.Sort_IVL(self.b.sp_ivl)
        vvl = Kl.Sort_VVL(self.b.sp_vvl,per,yz)
        zytr = Kl.Sort_Zytr(self.b.sp_zytr,per,yz)
        zvl = Kl.Sort_Zvl(self.b.sp_zvl)
        #print(1)
        # Хэшируем запускаемый расчёт
        k_ivl = [tuple(i) for i in ivl]
        k_vvl1 = []
        k_vvl2 = []
        for i in vvl:
            if i[0]>0: 
                k_vvl1.append(tuple(i))
            else: 
                k_vvl2.append(tuple(i))
        k_zy = [tuple(i) for i in zy]
        k_zytr = [tuple(i) for i in zytr]
        k_zvl1 = [i[0] for i in ivl]
        k_zvl2 = [tuple(i[1:9]) for i in zvl]
        #print(2)
        Hesh_now = [tuple(k_ivl+k_vvl2+k_zvl2+[self.b.pz])]
        k_1 = self.b.Hesh[0].get(Hesh_now[0],len(self.b.Hesh[0]))
        Hesh_now.append(tuple(k_zy+k_zytr+k_zvl1+[k_1]))
        Hesh_now.append(tuple(k_vvl1+[k_1]))
        Kesh = [self.b.kesh_1,self.b.kesh_2,self.b.kesh_3]
        Bool = [Hesh_now[i] in self.b.Hesh[i] for i in range(3)]
        Bool_hesh = [Bool[i] and Kesh[i] for i in range(3)]
        File_hesh = [self.b.Hesh[i].setdefault(Hesh_now[i],len(self.b.Hesh[i])) for i in range(3)]
        print(Bool_hesh)
        print(File_hesh)
        #print(3)
        # Запускаем расчет, получаем результаты
        tic = time()
        try:
            FiA, FiB, FiC, FiT1, p\
            = rnn.Vvod_inf(ivl, vvl, zy, zytr, zvl, per, self.b.k_yvgr,self.b.pz,Bool_hesh,File_hesh,self.b.kesh_4)
        except Exception as ex:
            if not Bool[0]:
                del self.b.Hesh[0][Hesh_now[0]]
            if not Bool[1]:
                del self.b.Hesh[1][Hesh_now[1]]
            if not Bool[2]:
                del self.b.Hesh[2][Hesh_now[2]]
            raise Exception(str(ex))

        tac = time()
        #print(tac-tic)
        #print(4)
        if not self.b.FTRCh:
            for i in range(len(FiT1)):
                for j in range(len(FiT1[i])):
                    FiT1[i][j]=0
        
        #Проверяем обьеденяем ли ветви
        self.b.marker=True
        for i in yr_v:
            if i =='Нет':
                self.b.marker=False
                break
        
        # Если заданы уровни обьеденяем ветви
        yr_v_d = False
        if self.b.marker:
            yr_v_d=Kl.SpSoed(yr_v)
            
            FiA=Kl.SoedV(yr_v_d,FiA)
            FiB=Kl.SoedV(yr_v_d,FiB)
            FiC=Kl.SoedV(yr_v_d,FiC)
            FiT1=Kl.SoedV(yr_v_d,FiT1)
            p=Kl.SoedV(yr_v_d,p)
            km=Kl.SoedV(yr_v_d,km)
            
            nm_ivl=Kl.ReName(yr_v_d,nm_ivl)
        #print(5)
        # Подготавливаем исходники для схем заземления и рисуем их
        per_name = Kl.Yz_p(ivl_gr,yr_v_d)
        sp_pit = Kl.PodpSz(ivl_gr, yr_v_d)
        per_name = Kl.UserReOp(KL_sp,per_name,yr_v_d if self.b.marker else None,per if self.b.marker else None)
        
        Tiks = []
        gr_list, axi_list, otp_list, point_list =\
            Kl.PointGen(ivl_gr,yr_v_d,self.b.sp_vvl,self.b.marker,per,nm_ivl,spzy,km,per_name) # Вызываем медод для прорисовки характерных точек на графике
        #print(6)        
        for i in range(len(FiA)):
            F = plt.figure(dpi=75) # Создаём фигуру графика 

            ax = F.add_subplot(111) #
            # Добавляем на график кривую, и задаём цвет и тип линии
        
            if self.b.FACh: ax.plot(km[i],FiA[i],'y',label=self.b.FAName) 
            if self.b.FBCh: ax.plot(km[i],FiB[i],'g--',label=self.b.FBName) 
            if self.b.FCCh: ax.plot(km[i],FiC[i],'r-.',label=self.b.FCName)
            if self.b.FTRCh: ax.plot(km[i],FiT1[i],'b:',label=self.b.FTRName)

            ax.plot([km[i][0],km[i][(len(km[i]))-1]],[self.b.sr_vel,self.b.sr_vel],'k')
            
            mx=1.1*max(max(FiA[i])*int(self.b.FACh),max(FiB[i])*int(self.b.FBCh),\
                    max(FiC[i])*int(self.b.FCCh),max(FiT1[i])*int(self.b.FTRCh))
            
            ax.axis([km[i][0],km[i][(len(km[i]))-1],0,mx]) # Задаём область графика
            
            for j in point_list[i]:
                ax.text (j[0], mx, j[1], horizontalalignment='center', verticalalignment='bottom')
                ax.plot([j[0],j[0]],[0,mx],'k:', linewidth=1)
                
            
            ax.grid(True) # Включаем сетку
            ax.set_xlabel(u'Длина, км',fontsize=self.b.shr_gr) # Подпись оси х
            ax.set_ylabel(u'U, В',fontsize=self.b.shr_gr) # Подпись оси у
            ax.legend(frameon=False,fontsize=self.b.shr_gr) # Выводим легенду графика
            ax.tick_params(labelsize=self.b.shr_gr) # Размер подписей шкалы графика
            locator = matplotlib.ticker.MaxNLocator()
            ax.xaxis.set_major_locator(locator)

            # Делаем отображение граничных значений на эпюрах
            T = ax.get_xticks()
            TT = []
            t_len = abs(T[0]-T[1])
            if km[i][0] < km[i][(len(km[i]))-1]:
                TT.append(km[i][0])
                for h in range((np.shape(T))[0]):
                    if km[i][0] < T[h] < km[i][(len(km[i]))-1] and\
                        km[i][(len(km[i]))-1]-T[h] > t_len/3 and T[h]-km[i][0]>t_len/2:
                        TT.append(T[h])
                TT.append(km[i][(len(km[i]))-1])
            elif km[i][0] > km[i][(len(km[i]))-1]:
                TT.append(km[i][0])
                for h in range((np.shape(T))[0]):
                    if km[i][0] > T[h] > km[i][(len(km[i]))-1] and\
                        km[i][0]-T[h]>t_len/3 and T[h]-km[i][(len(km[i]))-1]>t_len/2:
                        TT.append(T[h])
                TT.append(km[i][(len(km[i]))-1])
            T = np.array(TT)
            Tiks.append(T)

            ax.set_xticks(T)

            ax.set_title("№ опоры\n",fontsize=self.b.shr_gr)
            

            for obj in F.findobj(matplotlib.text.Text):
                obj.set_fontproperties(arial_font)
                obj.set_fontsize(self.b.shr_gr)
            
            F.set_size_inches(17, 5,forward=True) # Изменяем размер сохраняемого графика
            F.savefig('images_grafik/'+str(i)+'.jpg', format='jpg', dpi=100) # Cохраняем графики
        #print(7)
        if self.b.marker:
            sp_b, per=Kl.RePer(yr_v_d,per)
            m, n = Kl.Otpaiki(sp_b,ivl_gr,nm_ivl)
            inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per),nm_ivl,Kl.ReZy(sp_b,spzy))

            f, f1, f2 = Kl.Zapret(FiA, FiB, FiC,FiT1,p,self.b.sr_vel,self.b.FACh,self.b.FBCh,self.b.FCCh,self.b.FTRCh)
            kl_list = Kl.KL_metki(KL_sp,yr_v_d)
            
            re_op_key = list(per_name.keys())

            for i in range(len(nm_ivl)):
                Sz.Sxeme(inf[i],f1[i],lbsz[i],m[i],n[i],self.b.grop,self.b.zpps,self.b.dz,per_name[re_op_key[i]],sp_pit[i],kl_list[i])
                Gr_sblizh.Sx_sb(i,gr_list[i],axi_list[i],self.b.shr_gr, Tiks[i], otp_list[i])
            
        else:
            inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per),nm_ivl,spzy)
            f, f1, f2 = Kl.Zapret(FiA, FiB, FiC,FiT1,p, self.b.sr_vel,self.b.FACh,self.b.FBCh,self.b.FCCh,self.b.FTRCh)
            kl_list = Kl.KL_metki(KL_sp)
            
            re_op_key = list(per_name.keys())
            for i in range(len(nm_ivl)):
                Sz.Sxeme(inf[i],f1[i],lbsz[i],[],[],self.b.grop,self.b.zpps,self.b.dz,per_name[re_op_key[i]],sp_pit[i],kl_list[i])
                Gr_sblizh.Sx_sb(i,gr_list[i],axi_list[i],self.b.shr_gr, Tiks[i], otp_list[i])
        #print(8)
        copy_docx = self.b.Docx_file
        copy_docx = OtchetW.Word(inf, lbsz, f, f2, fname, self.b.name_vl.text(),NS,NV,self.b.grop,self.b.zpps,per_name,copy_docx,self.b.one_docx,self.b.tablepo.currentText())
        if self.b.one_docx and copy_docx!=None:
            self.b.Docx_file = copy_docx
        #print(9)
    def stop( self ):
        print('stop')
        self.terminate()
        self.wait()
        return 'stop'


class SxemThree(QThread):
    mysignal = QtCore.pyqtSignal(str)
    def __init__(self,b, parent=None):
        QtCore.QThread.__init__(self, parent)
        self.b=b
    def run(self):            
        try:      
            ivl, nm_ivl, km, per, yz, ivl_gr, yr_v, KL_sp= Kl.Sort_IVL(self.b.sp_ivl)
            vvl = Kl.Sort_VVL(self.b.sp_vvl,per,yz)
            zy = Kl.Sort_Zy(self.b.sp_zy,per,yz)
            zytr = Kl.Sort_Zytr(self.b.sp_zytr,per,yz)
            zvl = Kl.Sort_Zvl(self.b.sp_zvl)
            # Хэшируем запускаемый расчёт
            k_ivl = [tuple(i) for i in ivl]
            k_vvl1 = []
            k_vvl2 = []
            for i in vvl:
                if i[0]>0: 
                    k_vvl1.append(tuple(i))
                else: 
                    k_vvl2.append(tuple(i))
            k_zy = [tuple(i) for i in zy]
            k_zytr = [tuple(i) for i in zytr]
            k_zvl1 = [i[0] for i in ivl]
            k_zvl2 = [tuple(i[1:9]) for i in zvl]

            Hesh_now = [tuple(k_ivl+k_vvl2+k_zvl2+[self.b.pz])]
            k_1 = self.b.Hesh[0].get(Hesh_now[0],len(self.b.Hesh[0]))
            Hesh_now.append(tuple(k_zy+k_zytr+k_zvl1+[k_1]))
            Hesh_now.append(tuple(k_vvl1+[k_1]))
            
            self.b.marker=True
            for i in yr_v:
                if i =='Нет':
                    self.b.marker=False
                    break
            
            if self.b.marker:
                yr_v_d=Kl.SpSoed(yr_v)
                per_name = Kl.Yz_p(ivl_gr,yr_v_d)
                sp_b, per1=Kl.RePer(yr_v_d,per)
                km = Kl.SoedV(yr_v_d,km)

            else:
                per_name = Kl.Yz_p(ivl_gr,False)
                per1=per

            per_name = Kl.UserReOp(KL_sp,per_name,yr_v_d if self.b.marker else None,per if self.b.marker else None)
            
            self.mysignal.emit('Расчет Схем №2')
            # Расчитываем возможные вторые схемы для для выбора ПС для заземления
            M=[]
            inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per),nm_ivl,self.b.sp_zy)
            for i in sp_PZ:
                for j in range(len(self.b.sp_zy)):
                    if str(i[0])==self.b.sp_zy[j][0] and str(i[1])==self.b.sp_zy[j][1] and str(i[2])==self.b.sp_zy[j][2]:
                        zy[j][2]=1
                        v=j    
                    else:
                        zy[j][2]=7

                FiA=[]
                FiB=[]
                FiC=[]
                p=[]

                # Хэшируем запускаемый расчёт

                k_zy = [tuple(i) for i in zy]
                
                Hesh_now[1] = tuple(k_zy+k_zytr+k_zvl1+[k_1])
                Kesh = [self.b.kesh_1,self.b.kesh_2,self.b.kesh_3]
                Bool = [Hesh_now[i] in self.b.Hesh[i] for i in range(3)]
                Bool_hesh = [Bool[i] and Kesh[i] for i in range(3)]
                File_hesh = [self.b.Hesh[i].setdefault(Hesh_now[i],len(self.b.Hesh[i])) for i in range(3)]

                try:
                    FiA, FiB, FiC, FiT1, p =\
                         rnn.Vvod_inf(ivl, vvl, zy, zytr, zvl, per,self.b.k_yvgr,self.b.pz,Bool_hesh,File_hesh,self.b.kesh_4)
                except Exception as ex:
                    if not Bool[0]:
                        del self.b.Hesh[0][Hesh_now[0]]
                    if not Bool[1]:
                        del self.b.Hesh[1][Hesh_now[1]]
                    if not Bool[2]:
                        del self.b.Hesh[2][Hesh_now[2]]
                    raise Exception(str(ex))
                
                if not self.b.FTRCh:
                    for i in range(len(FiT1)):
                        for j in range(len(FiT1[i])):
                            FiT1[i][j]=0
                if self.b.marker:
                    FiA=Kl.SoedV(yr_v_d,FiA)
                    FiB=Kl.SoedV(yr_v_d,FiB)
                    FiC=Kl.SoedV(yr_v_d,FiC)
                    FiT1=Kl.SoedV(yr_v_d,FiT1)

                
                M.append(Kl.MF(FiA, FiB, FiC, FiT1))

                zy[v][2]=0

            # Добавляем в ЗУ строку, которая будет характеризовать БЗ
            zy.append(['','','','',''])

            d_PZ=Kl.dict_PZ(sp_PZ)

            if self.b.marker:
                nm_ivl=Kl.ReName(yr_v_d,nm_ivl)
                inf, lbsz, sp_PZ = Kl.Szazeml(Kl.Kon(per1),nm_ivl,Kl.ReZy(sp_b,self.b.sp_zy))

            ky=list(per1.keys())
            ku1={}
            ku2={}
            for i in range(1,len(nm_ivl)+1):
                ku1[nm_ivl[i-1]]='Линия '+str(i)
                ku2[ky[i-1]]='Линия '+str(i)
            
            sp_op=[]
            sp_op1=[]
            sp_t=[]
            a=0
            
            for key1 in per1:
                a+=1
                b=0
                sp_t.append(ku2[key1])
                sp_op.append([ku2[key1]])
                for key2 in per1[key1]:
                    b+=1

                    u=False
                    for v in sp_PZ:
                        if key1==(v[0],v[1]) and v[2]==key2:
                            u=True
                            break
                    if u:
                        continue
                    """
                    QCoreApplication.processEvents()
                    if Ind.wasCanceled():
                        break
                    Ind.setLabelText('Ветвь '+str(key1[0])+'-'+str(key1[1])+', опора № '+str(key2))
                    Ind.setValue(b)
                    QCoreApplication.processEvents()
                    """
                    self.mysignal.emit('Ветвь '+str(key1[0])+'-'+str(key1[1])+', опора № '+str(key2))
                    
                    zy[len(zy)-1][0]=per1[key1][key2]
                    zy[len(zy)-1][1]=float(Kl.Tochka(self.b.zy_om.text()))
                    zy[len(zy)-1][2]=1
                    zy[len(zy)-1][3]=0
                    zy[len(zy)-1][4]=0

                    FiA=[]
                    FiB=[]
                    FiC=[]
                    p=[]
                    f=[]
                    f1=[]
                    f2=[]

                    # Хэшируем запускаемый расчёт
                    
                    k_zy = [tuple(i) for i in zy]
                    
                    Hesh_now[1] = tuple(k_zy+k_zytr+k_zvl1+[k_1])
                    Kesh = [self.b.kesh_1,self.b.kesh_2,self.b.kesh_3]
                    Bool = [Hesh_now[i] in self.b.Hesh[i] for i in range(3)]
                    Bool_hesh = [Bool[i] and Kesh[i] for i in range(3)]
                    File_hesh = [self.b.Hesh[i].setdefault(Hesh_now[i],len(self.b.Hesh[i])) for i in range(3)]

                    try:
                        FiA, FiB, FiC, FiT1, p =\
                            rnn.Vvod_inf(ivl, vvl, zy, zytr, zvl, per,self.b.k_yvgr,self.b.pz,Bool_hesh,File_hesh,self.b.kesh_4)
                    except Exception as ex:
                        if not Bool[0]:
                            del self.b.Hesh[0][Hesh_now[0]]
                        if not Bool[1]:
                            del self.b.Hesh[1][Hesh_now[1]]
                        if not Bool[2]:
                            del self.b.Hesh[2][Hesh_now[2]]
                        raise Exception(str(ex))

                    if not self.b.FTRCh:
                        for i in range(len(FiT1)):
                            for j in range(len(FiT1[i])):
                                FiT1[i][j]=0
                    if self.b.marker:
                        FiA=Kl.SoedV(yr_v_d,FiA)
                        FiB=Kl.SoedV(yr_v_d,FiB)
                        FiC=Kl.SoedV(yr_v_d,FiC)
                        FiT1=Kl.SoedV(yr_v_d,FiT1)
                        p=Kl.SoedV(yr_v_d,p)

                    f, f1, f2 = Kl.Zapret(FiA, FiB, FiC, FiT1, p, self.b.sr_vel,self.b.FACh,self.b.FBCh,self.b.FCCh,self.b.FTRCh)

                    for i in range(len(inf)):
                        if self.b.grop:
                            for j in range(len(f[i])):
                                if inf[i][0]==1 and abs(inf[i][2]-inf[i][3])>=3:
                                    if inf[i][2]==f[i][j][0][0]:
                                        if inf[i][2]<inf[i][3]:
                                            f[i][j][0][0]=inf[i][2]+1
                                        elif inf[i][2]>inf[i][3]:
                                            f[i][j][0][0]=inf[i][2]-1
                                    if inf[i][3]==f[i][j][0][1]:
                                        if inf[i][2]<inf[i][3]:
                                            f[i][j][0][1]=inf[i][3]-1
                                        elif inf[i][2]>inf[i][3]:
                                            f[i][j][0][1]=inf[i][3]+1
                                if inf[i][0]==2 and abs(inf[i][2]-inf[i][3])>=2:
                                    if inf[i][2]==f[i][j][0][0]:
                                        if inf[i][2]<inf[i][3]:
                                            f[i][j][0][0]=inf[i][2]+1
                                        elif inf[i][2]>inf[i][3]:
                                            f[i][j][0][0]=inf[i][2]-1
                                if inf[i][0]==3 and abs(inf[i][2]-inf[i][3])>=2:
                                        
                                    if inf[i][3]==f[i][j][0][1]:
                                        if inf[i][2]<inf[i][3]:
                                            f[i][j][0][1]=inf[i][3]-1
                                        elif inf[i][2]>inf[i][3]:
                                            f[i][j][0][1]=inf[i][3]+1
                        if self.b.zpps:
                            keyf2=list(f2[i].keys())
                            for j in keyf2:
                                f2[i][j]=True
                                
                                
                        

                    c = Kl.Zpz(M,a-1,b-1)
                    if not self.b.zpps:
                        for gg in range(len(f2)):
                            for kk in f2[gg]:
                                f2[gg][kk] = False
                    
                    
                    

                    for gg in range(len(f)):
                        if gg == a-1:
                            for kk in range(len(f[gg])):
                                if f[gg][kk][1] == False:
                                    if f[gg][kk][0][0] <= f[gg][kk][0][1]:
                                        if not f[gg][kk][0][0] <= key2 <= f[gg][kk][0][1]:
                                            f[gg][kk][1] = True
                                    elif f[gg][kk][0][0] > f[gg][kk][0][1]:
                                        if not f[gg][kk][0][0] >= key2 >= f[gg][kk][0][1]:
                                            f[gg][kk][1] = True
                        else:
                            f[gg] = []
                    


                    

                    sp_op.append(Kl.stTable(f, f2, c, sp_PZ, key2, ky, d_PZ, per_name,key1,len(sp_t)-1))

                    
            self.mysignal.emit('Оформление отчета')
            copy_docx = self.b.Docx_file
            copy_docx = SxT.WordSxT(sp_t,sp_op,d_PZ,ku1,self.b.fname3, self.b.name_vl.text(),copy_docx,self.b.one_docx)
            if self.b.one_docx and copy_docx!=None:
                self.b.Docx_file = copy_docx


            
            
            self.mysignal.emit('Good')
        except Exception as ex:
            self.mysignal.emit('mistake:'+str(ex))

    def stop( self ):
        print('stop')
        self.terminate()
        self.wait()
        return 'stop'
            
# Закрываем заставку
splash.close()
del app            
        

if __name__=='__main__':
    app=QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())
